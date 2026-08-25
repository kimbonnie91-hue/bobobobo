# -*- coding: utf-8 -*-
"""
일일 PUSH 발송성과 대시보드
─────────────────────────────────────────────────────────────
쿼리로 추출해 매일 갱신하는 "일일 PUSH 실적.xlsx"를 업로드하면 AF코드 기준으로
발송모수/UV/주문건수/거래액을 집계해 시각화한다.

· 이미 집계된 '누적_소재별'/'소재별 실적(당주)_N' 시트를 우선 읽고, 그 시트가 없거나 거기 없는
  일자·AF코드는 '발송 건수'+'RawNew' 원본 쿼리 시트에서 직접 (일자+AF코드) 단위로 집계해 채운다
  — 즉 쿼리 결과만 두 시트에 그대로 붙여넣고 올려도 동작한다.
· BPU별 실적, 발송유형별(기본/우수) 실적, 주차별 누적 추이는 "BPU별 실적"/"발송유형별 실적"
  피벗 시트를 그대로 읽지 않고, 소재별 실적 원본 로우 데이터에서 매번 다시 집계한다.
  (원본 피벗 시트의 셀 레이아웃이 바뀌어도 항상 정확한 값을 보장하기 위함)
· 업로드할 때마다 (일자+AF코드+시간대+타겟구분) 키로 누적 저장 — 구글시트(설정 시) ↔ 로컬 CSV 폴백.

데이터 로직(파싱/집계 함수)은 Streamlit 비의존 순수 함수이며 모듈 import 만으로 테스트 가능하다.
앱 UI 는 main() 안에 있고 `python -m streamlit run daily_push_perf_dashboard.py` 시에만 실행된다.
"""
import io, os, re, datetime, math
import numpy as np
import pandas as pd

# ══════════════════════════════════════════════════════════════════════
# 1. 순수 데이터 로직 (Streamlit 비의존)
# ══════════════════════════════════════════════════════════════════════

# 시트/구글시트 헤더가 "일자(8자리)" "타겟구분" "기획전No." 처럼 표기가 흔들려도
# 매칭되도록 괄호 안 내용을 지우고 영문/숫자/한글만 남겨 비교한다.
def _norm_h(s):
    s = str(s if s is not None else "")
    s = re.sub(r"\(.*?\)", "", s)
    s = re.sub(r"[^0-9A-Za-z가-힣]", "", s)
    return s


COLMAP_CANDIDATES = {
    "date":    ["일자(8자리)", "일자", "날짜"],
    "dow_k":   ["요일"],
    "hour":    ["시간대", "시간"],
    "target":  ["타겟구분", "타겟 구분", "타겟"],
    "stype":   ["발송유형"],
    "bpu":     ["BPU"],
    "prio":    ["우선순위"],
    "cat":     ["카테고리"],
    "attr":    ["속성"],
    "owner":   ["담당자"],
    "brand":   ["브랜드"],
    "af":      ["AF코드", "AF 코드"],
    "promo":   ["기획전No.", "기획전 No.", "기획전No", "기획전"],
    "send":    ["발송모수", "발송건수", "발송"],
    "uv":      ["UV"],
    "visit":   ["VISIT", "방문"],
    "cust":    ["고객수"],
    "oc":      ["주문건수"],
    "amt":     ["거래액", "주문금액", "GMV"],
    "infl_cr": ["유입전환율"],
    "ord_cr":  ["주문전환율"],
    "eff":     ["효율"],
}
NUM_COLS = ["prio", "hour", "send", "uv", "visit", "cust", "oc", "amt", "infl_cr", "ord_cr", "eff"]
TXT_COLS = ["dow_k", "target", "stype", "bpu", "cat", "attr", "owner", "brand", "promo"]


def _cand_norms(candidates):
    return {_norm_h(c) for cands in candidates.values() for c in cands}


_ALL_CAND_NORMS = _cand_norms(COLMAP_CANDIDATES)


def _norm_date(v):
    """엑셀 셀 값 → 'YYYYMMDD' 문자열 또는 None. datetime/문자열/구분자/실수(.0) 형태 모두 허용.
    엑셀 날짜 셀을 표시형식 그대로 복사-붙여넣기하면 '26/06/29'처럼 2자리 연도로 들어오기도
    해서(예: '발송일시' 컬럼), 6자리(YYMMDD)면 '20'을 붙여 8자리로 보정한다."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y%m%d")
    s = re.sub(r"\.0+$", "", str(v).strip())
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        return digits
    if len(digits) == 6:
        return "20" + digits
    return None


def _find_header_row(rows, candidates=None, max_scan=100, min_score=3):
    """rows 앞부분에서 candidates(기본 COLMAP_CANDIDATES)와 가장 많이 일치하는 행을 헤더로 판정."""
    norms = _ALL_CAND_NORMS if candidates is None else _cand_norms(candidates)
    best_i, best_score = None, min_score - 1
    for i, row in enumerate(rows[:max_scan]):
        cells = [_norm_h(v) for v in (row or [])]
        score = sum(1 for c in cells if c and c in norms)
        if score > best_score:
            best_score, best_i = score, i
    return best_i


def _map_columns(header_cells, candidates=None):
    """헤더 셀 리스트 → {표준키: 열 인덱스}."""
    candidates = COLMAP_CANDIDATES if candidates is None else candidates
    normed = [_norm_h(c) for c in header_cells]
    idx = {}
    for key, cands in candidates.items():
        for cand in cands:
            cn = _norm_h(cand)
            if cn in normed:
                idx[key] = normed.index(cn)
                break
    return idx


def parse_material_rows(rows):
    """소재별 실적 시트(2차원 rows) → AF코드 단위 레코드 DataFrame."""
    if not rows:
        return pd.DataFrame()
    hdr_i = _find_header_row(rows)
    if hdr_i is None:
        return pd.DataFrame()
    header = [("" if v is None else str(v)) for v in rows[hdr_i]]
    idx = _map_columns(header)
    if "af" not in idx:
        return pd.DataFrame()
    af_i = idx["af"]
    recs = []
    for row in rows[hdr_i + 1:]:
        if row is None:
            continue
        n = len(row)
        af_raw = row[af_i] if af_i < n else None
        af = "" if af_raw is None else str(af_raw).strip()
        if not af or af.lower() in ("nan", "none", "-"):
            continue
        rec = {key: (row[i] if i < n else None) for key, i in idx.items()}
        rec["af"] = af
        rec["date"] = _norm_date(rec.get("date"))
        recs.append(rec)
    return pd.DataFrame(recs)


# ── '발송 건수'/'RawNew' 원본 쿼리 시트 → AF코드 기준 자동 집계 ──────────────
# '소재별 실적(당주)' 류 시트가 아직 없어도, 매일 쿼리로 뽑아 붙여넣는 이 두 원본
# 시트만으로 (일자+AF코드) 단위 발송모수/UV/VISIT/고객수/주문건수/거래액을 만든다.
SEND_COUNT_CANDIDATES = {
    "sendtime": ["발송일시"],
    "camp_name": ["캠페인명"],
    "channel": ["발송채널"],
    "af": ["AF코드"],
    "camp_type": ["캠페인구분"],
    "send": ["모수"],
}
RAWNEW_CANDIDATES = {
    "std_dd": ["STD_DD"],
    "af": ["AF_CD"],
    "af_nm": ["AF_NM"],
    "chnl": ["CHNL_DTL_CD"],
    "uv": ["UV"],
    "visit": ["SV"],
    "cust": ["CUST_CNT"],
    "oc": ["ORD_CNT"],
    "amt": ["SALE_AMT"],
}
# '기획' 발송의 캠페인명은 "REAL_20260629_0800_기본발송_편성_112512" 형태(시간대_발송유형_BPU_기획전No)로
# 템플릿화되어 있어 브랜드명이 아니다 — 이 패턴이 매칭되면 브랜드로 쓰지 않고 메타로만 활용한다.
_CAMPNAME_RE = re.compile(r"^[^_]+_(\d{8})_(\d{3,4})_([^_]+)_([^_]+)_(\d+)$")


def parse_send_count_rows(rows):
    """'발송 건수' 원본 쿼리 시트(발송 로그 1건=1행) → (일자, AF코드, 캠페인명, 모수) 레코드."""
    if not rows:
        return pd.DataFrame()
    hdr_i = _find_header_row(rows, candidates=SEND_COUNT_CANDIDATES, min_score=2)
    if hdr_i is None:
        return pd.DataFrame()
    header = [("" if v is None else str(v)) for v in rows[hdr_i]]
    idx = _map_columns(header, candidates=SEND_COUNT_CANDIDATES)
    if "af" not in idx or "send" not in idx:
        return pd.DataFrame()
    recs = []
    for row in rows[hdr_i + 1:]:
        if row is None:
            continue
        n = len(row)
        af_i = idx["af"]
        af = row[af_i] if af_i < n else None
        af = "" if af is None else str(af).strip()
        if not af:
            continue
        rec = {key: (row[i] if i < n else None) for key, i in idx.items()}
        rec["af"] = af
        rec["date"] = _norm_date(rec.get("sendtime"))
        recs.append(rec)
    return pd.DataFrame(recs)


def parse_rawnew_rows(rows):
    """'RawNew' 원본 쿼리 시트(전사 채널 성과 원자료) → (일자, AF코드)별 UV/VISIT/고객수/주문건수/거래액 원자료.

    CHNL_DTL_CD가 여러 세부채널(A/M 등)로 쪼개져 있고 'Total' 행이 있으면 그 행만 써야
    중복 합산을 피한다 (Total 없는 그룹은 세부채널을 그대로 합산)."""
    if not rows:
        return pd.DataFrame()
    hdr_i = _find_header_row(rows, candidates=RAWNEW_CANDIDATES, min_score=3)
    if hdr_i is None:
        return pd.DataFrame()
    header = [("" if v is None else str(v)) for v in rows[hdr_i]]
    idx = _map_columns(header, candidates=RAWNEW_CANDIDATES)
    if "af" not in idx or "std_dd" not in idx:
        return pd.DataFrame()
    recs = []
    for row in rows[hdr_i + 1:]:
        if row is None:
            continue
        n = len(row)
        af_i = idx["af"]
        af = row[af_i] if af_i < n else None
        af = "" if af is None else str(af).strip()
        if not af:
            continue
        date = _norm_date(row[idx["std_dd"]] if idx["std_dd"] < n else None)
        if date is None:
            continue
        rec = {key: (row[i] if i < n else None) for key, i in idx.items()}
        rec["af"] = af
        rec["date"] = date
        rec["chnl"] = "" if rec.get("chnl") is None else str(rec["chnl"]).strip()
        recs.append(rec)
    return pd.DataFrame(recs)


def build_af_meta_lookup(df):
    """기존 소재별 실적 데이터 → AF코드별 최근 메타데이터(BPU/발송유형/카테고리/속성/담당자/
    브랜드/기획전/타겟구분/시간대) 사전. '발송 건수'+'RawNew'만으로는 알 수 없는 항목을 채우는 용도."""
    if df is None or df.empty:
        return {}
    d = df.sort_values("date") if "date" in df.columns else df
    lookup = {}
    for _, r in d.iterrows():
        af = str(r.get("af", "") or "").strip()
        if not af:
            continue
        lookup[af] = {k: (r.get(k) or "") for k in
                      ("bpu", "stype", "cat", "attr", "owner", "brand", "promo", "target", "hour")}
    return lookup


def _parse_campname_meta(name):
    """'REAL_20260629_0800_기본발송_편성_112512' → {hour, stype, bpu, promo}. 매칭 안 되면 {}."""
    if not name:
        return {}
    m = _CAMPNAME_RE.match(str(name).strip())
    if not m:
        return {}
    return {"hour": m.group(2), "stype": m.group(3), "bpu": m.group(4), "promo": m.group(5)}


_DOW_K = ["월", "화", "수", "목", "금", "토", "일"]


def build_from_raw_query_sheets(send_count_raw, rawnew_raw, meta_lookup=None):
    """'발송 건수'(발송로그) + 'RawNew'(성과 원자료) → (일자+AF코드) 단위 소재별 실적 행.

    meta_lookup: build_af_meta_lookup()의 결과 — 같은 통합문서의 기존 소재별 실적류 시트에서
    뽑은 AF코드→메타데이터. 없으면 캠페인명 패턴 인식 결과만으로 채우고 나머지는 빈 값."""
    meta_lookup = meta_lookup or {}
    has_send = send_count_raw is not None and not send_count_raw.empty
    has_perf = rawnew_raw is not None and not rawnew_raw.empty
    if not has_send and not has_perf:
        return pd.DataFrame()

    send_agg = pd.DataFrame(columns=["date", "af", "send", "camp_name"])
    if has_send:
        s = send_count_raw.dropna(subset=["date"]).copy()
        s["send"] = pd.to_numeric(s["send"], errors="coerce")
        send_agg = s.groupby(["date", "af"], as_index=False).agg(send=("send", "sum"), camp_name=("camp_name", "first"))

    perf_agg = pd.DataFrame(columns=["date", "af", "uv", "visit", "cust", "oc", "amt"])
    if has_perf:
        r = rawnew_raw.copy()
        if "sale_amt" in r.columns:
            r["amt"] = r["sale_amt"]
        for c in ("uv", "visit", "cust", "oc", "amt"):
            r[c] = pd.to_numeric(r[c], errors="coerce")
        r["chnl_norm"] = r["chnl"].astype(str).str.strip().str.lower()
        has_total = r.groupby(["date", "af"])["chnl_norm"].transform(lambda s: (s == "total").any())
        use = r[(~has_total) | (r["chnl_norm"] == "total")]
        perf_agg = use.groupby(["date", "af"], as_index=False).agg(
            uv=("uv", "sum"), visit=("visit", "sum"), cust=("cust", "sum"), oc=("oc", "sum"), amt=("amt", "sum"))

    if has_send:
        # '발송 건수'에 실제 발송 기록이 있는 (일자,AF코드)만 남긴다 — RawNew는 전사 전 채널
        # 원자료라 그날 발송이 없어도 잔여 귀속 클릭/주문이 잡힐 수 있어 노이즈가 된다.
        merged = pd.merge(send_agg, perf_agg, on=["date", "af"], how="left")
    else:
        merged = perf_agg.copy()
    if merged.empty:
        return merged
    for c in ("send", "uv", "visit", "cust", "oc", "amt"):
        if c not in merged.columns:
            merged[c] = np.nan
    if "camp_name" not in merged.columns:
        merged["camp_name"] = None

    rows = []
    for _, r in merged.iterrows():
        af = r["af"]
        meta = meta_lookup.get(af, {})
        parsed = _parse_campname_meta(r.get("camp_name"))
        dt = pd.to_datetime(r["date"], format="%Y%m%d", errors="coerce")
        brand = meta.get("brand", "") if parsed else (r.get("camp_name") or meta.get("brand", ""))
        rec = {c: None for c in STORE_COLS}
        rec.update({
            "date": r["date"], "af": af,
            "dow_k": _DOW_K[dt.dayofweek] if pd.notna(dt) else "",
            "hour": parsed.get("hour") or meta.get("hour", ""),
            "stype": parsed.get("stype") or meta.get("stype", ""),
            "bpu": parsed.get("bpu") or meta.get("bpu", ""),
            "promo": parsed.get("promo") or meta.get("promo", ""),
            "brand": brand or "",
            "cat": meta.get("cat", ""), "attr": meta.get("attr", ""),
            "owner": meta.get("owner", ""), "target": meta.get("target", ""),
            "send": r.get("send"), "uv": r.get("uv"), "visit": r.get("visit"),
            "cust": r.get("cust"), "oc": r.get("oc"), "amt": r.get("amt"),
        })
        rows.append(rec)
    return pd.DataFrame(rows)


def discover_raw_query_sheets(sheetnames):
    """워크북 시트명 목록 → ('발송 건수' 시트명, 'RawNew' 후보 시트명 리스트). 이름에 trailing space 등
    표기가 흔들려도(정규화 시 동일해짐) 후보를 모두 모아 두고, 여러 개면 날짜가 최신인 것을 고른다."""
    send_cnt = None
    rawnew_candidates = []
    for s in sheetnames:
        sn = _norm_h(s)
        if sn == "발송건수":
            send_cnt = send_cnt or s
        elif sn.lower() == "rawnew":
            rawnew_candidates.append(s)
    return send_cnt, rawnew_candidates


def _peek_max_date(wb, sheet_name, candidates, sample=1000):
    """시트 전체를 다 읽지 않고 앞부분 표본만으로 대략적인 최신 날짜를 추정
    (동일 이름의 여러 후보 시트 중 '최신 데이터가 든 것'을 빠르게 골라내기 위함)."""
    ws = wb[sheet_name]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return None
    idx = _map_columns([("" if v is None else str(v)) for v in header], candidates=candidates)
    date_key = "std_dd" if "std_dd" in idx else ("sendtime" if "sendtime" in idx else None)
    if date_key is None:
        return None
    ci = idx[date_key]
    best = None
    for row in ws.iter_rows(min_row=2, max_row=1 + sample, values_only=True):
        if ci < len(row):
            d = _norm_date(row[ci])
            if d and (best is None or d > best):
                best = d
    return best


def discover_perf_sheets(sheetnames):
    """워크북 시트명 목록 → (누적_소재별 시트명, 소재별 실적(당주) 시트명). 없으면 None."""
    cum = week = None
    for s in sheetnames:
        sn = re.sub(r"\s+", "", s)
        if "누적" in sn and "소재" in sn:
            cum = cum or s
        elif "소재별" in sn and "실적" in sn:
            week = week or s
    return cum, week


def load_excel_perf(file_bytes):
    """업로드 엑셀 바이트 → (합쳐진 DataFrame, 실제로 읽은 시트명 리스트).

    소스는 두 종류이고, 겹치면(같은 일자+AF코드) 사전 집계 시트 쪽 값을 우선한다:
    1) '누적_소재별' / '소재별 실적(당주)_*' — 이미 다 만들어진 소재별 실적(사전 집계, 우선순위 높음)
    2) '발송 건수' + 'RawNew' — 매일 쿼리로 붙여넣는 원본 로그. 소재별 실적 시트가 아직 없거나
       그 시트에 없는 일자+AF코드가 있으면, 이 두 시트만으로 직접 집계해서 채운다.
       (동명의 'RawNew' 시트가 여러 개면 표본을 훑어 날짜가 최신인 쪽을 고른다.)
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        cum_name, week_name = discover_perf_sheets(names)

        curated_frames, used = [], []
        for sname in (cum_name, week_name):
            if not sname:
                continue
            rows = list(wb[sname].iter_rows(values_only=True))
            d = parse_material_rows(rows)
            if not d.empty:
                curated_frames.append(d)
                used.append(sname)

        if not curated_frames:
            for s in names:
                head_rows = list(wb[s].iter_rows(values_only=True, max_row=10))
                hdr_i = _find_header_row(head_rows)
                if hdr_i is None:
                    continue
                header = [("" if v is None else str(v)) for v in head_rows[hdr_i]]
                idx = _map_columns(header)
                if "af" in idx and ("amt" in idx or "send" in idx):
                    rows = list(wb[s].iter_rows(values_only=True))
                    d = parse_material_rows(rows)
                    if not d.empty:
                        curated_frames.append(d)
                        used.append(s)

        curated = pd.DataFrame()
        if curated_frames:
            curated = pd.concat(curated_frames, ignore_index=True)
            key_cols = [c for c in ("date", "af", "hour", "target") if c in curated.columns]
            if key_cols:
                curated = curated.drop_duplicates(subset=key_cols, keep="last")

        send_cnt_name, rawnew_candidates = discover_raw_query_sheets(names)
        rawnew_name = None
        if len(rawnew_candidates) == 1:
            rawnew_name = rawnew_candidates[0]
        elif len(rawnew_candidates) > 1:
            scored = [(c, _peek_max_date(wb, c, RAWNEW_CANDIDATES)) for c in rawnew_candidates]
            scored = [(c, d) for c, d in scored if d]
            rawnew_name = max(scored, key=lambda x: x[1])[0] if scored else rawnew_candidates[0]

        derived = pd.DataFrame()
        derived_label = None
        if send_cnt_name or rawnew_name:
            send_raw = parse_send_count_rows(list(wb[send_cnt_name].iter_rows(values_only=True))) \
                if send_cnt_name else pd.DataFrame()
            rawnew_raw = parse_rawnew_rows(list(wb[rawnew_name].iter_rows(values_only=True))) \
                if rawnew_name else pd.DataFrame()
            meta_lookup = build_af_meta_lookup(curated) if not curated.empty else {}
            derived = build_from_raw_query_sheets(send_raw, rawnew_raw, meta_lookup)
            if not derived.empty:
                derived_label = " + ".join(n for n in (send_cnt_name, rawnew_name) if n) + " (자동 집계)"

        if not derived.empty and not curated.empty:
            covered = set(zip(curated["date"], curated["af"]))
            derived = derived[~derived.apply(lambda r: (r["date"], r["af"]) in covered, axis=1)]

        frames = [d for d in (derived, curated) if not d.empty]
        if not frames:
            return pd.DataFrame(), []
        if derived_label and not derived.empty:
            used.insert(0, derived_label)
        combined = pd.concat(frames, ignore_index=True)
        key_cols = [c for c in ("date", "af", "hour", "target") if c in combined.columns]
        if key_cols:
            combined = combined.drop_duplicates(subset=key_cols, keep="last")
        return combined, used
    finally:
        wb.close()


def _txt_val(v):
    """텍스트 컬럼 정리 — CSV 왕복 후 결측이 None 대신 NaN(float)으로 오거나, 숫자로만 된 값이
    112864.0처럼 float로 캐스팅돼 들어와도 "nan"/".0" 없이 깔끔한 문자열로 만든다."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _norm_bpu_group(v):
    """세부 BPU 코드(1BPU_B, 1BPU_C, 3BPU_B, 4BPU_A 등)를 대표 BPU(1BPU/3BPU/4BPU)로 묶는다.
    '마케팅'/'편성'/'브랜드컨텐츠'/'2BPU'처럼 접미사가 없는 값은 그대로 둔다."""
    s = str(v or "").strip()
    if not s:
        return s
    m = re.match(r"^(\d*BPU)_[A-Za-z0-9]+$", s, re.I)
    return m.group(1) if m else s


def _finalize(df):
    """숫자/텍스트 정리 + 파생지표(ctr/cvr/rps/aov) + 주차(월~일) 라벨 계산."""
    cols = list(COLMAP_CANDIDATES.keys())
    if df is None or df.empty:
        extra = ["dt", "week_start", "week_end", "week_label", "dow", "ctr", "cvr", "rps", "aov", "bpu_group"]
        return pd.DataFrame(columns=cols + extra)
    df = df.copy()
    df["af"] = df["af"].astype(str).str.strip()
    df["date"] = df["date"].astype(str).str.replace(r"\.0$", "", regex=True)
    for c in NUM_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce") if c in df else np.nan
    for c in TXT_COLS:
        if c in df:
            df[c] = df[c].apply(_txt_val)
        else:
            df[c] = ""
    df["bpu_group"] = df["bpu"].map(_norm_bpu_group)
    df["dt"] = pd.to_datetime(df["date"], format="%Y%m%d", errors="coerce")
    df = df.dropna(subset=["dt"]).reset_index(drop=True)
    wk = df["dt"].dt.to_period("W-SUN")
    df["week_start"] = wk.dt.start_time.dt.date
    df["week_end"] = wk.dt.end_time.dt.date
    df["week_label"] = df.apply(
        lambda r: f"{r['week_start'].month}/{r['week_start'].day}~{r['week_end'].month}/{r['week_end'].day}",
        axis=1)
    df["dow"] = df["dt"].dt.dayofweek
    send = df["send"].fillna(0); uv = df["uv"].fillna(0); oc = df["oc"].fillna(0); amt = df["amt"].fillna(0)
    df["ctr"] = np.where(send > 0, uv / send, 0.0)
    df["cvr"] = np.where(uv > 0, oc / uv, 0.0)
    df["rps"] = np.where(send > 0, amt / send, 0.0)
    df["aov"] = np.where(oc > 0, amt / oc, 0.0)
    return df


def agg_metrics(df, group_cols):
    """group_cols 기준 발송/UV/주문건수/거래액 합계 + 파생 효율지표."""
    if df is None or df.empty:
        return pd.DataFrame(columns=list(group_cols) + ["n", "send", "uv", "oc", "amt", "ctr", "cvr", "rps", "aov"])
    g = df.groupby(list(group_cols), dropna=False).agg(
        n=("af", "count"), send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum"),
    ).reset_index()
    g["ctr"] = np.where(g["send"] > 0, g["uv"] / g["send"], 0.0)
    g["cvr"] = np.where(g["uv"] > 0, g["oc"] / g["uv"], 0.0)
    g["rps"] = np.where(g["send"] > 0, g["amt"] / g["send"], 0.0)
    g["aov"] = np.where(g["oc"] > 0, g["amt"] / g["oc"], 0.0)
    return g


# ── 누적 저장소 — 업로드할 때마다 (일자+AF코드+시간대+타겟구분) 키로 병합 ─────
DATA_STORE = "daily_push_perf_store.csv"
STORE_COLS = list(COLMAP_CANDIDATES.keys())
STORE_KEY_COLS = ["date", "af", "hour", "target"]


def _norm_key_val(v):
    """병합 키 값 정규화 — None/NaN/빈문자열은 전부 ''로, 숫자형 문자열은 선행0·소수점(.0)을
    없앤 정수 표기로 맞춘다 ('0800'/'800.0'/800 모두 '800'). CSV 왕복 후 dtype이 바뀌어도
    (NaN↔'', '800'↔800.0) 같은 값으로 인식되게 하기 위함 — 안 하면 재업로드마다 중복 저장된다."""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", ""):
        return ""
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else str(f)
    except (ValueError, TypeError):
        return s


def store_key_frame(d):
    k = pd.DataFrame(index=d.index)
    for c in STORE_KEY_COLS:
        k[c] = d[c].map(_norm_key_val) if c in d.columns else ""
    return k


def merge_store(old, new):
    """기존 누적 + 신규 업로드 병합 — 같은 키는 신규 우선."""
    def _pick(d):
        if d is None or len(d) == 0:
            return pd.DataFrame(columns=STORE_COLS)
        return d[[c for c in STORE_COLS if c in d]].copy()
    both = pd.concat([_pick(old), _pick(new)], ignore_index=True)
    if both.empty:
        return both
    both["date"] = both["date"].astype(str).str.replace(r"\.0$", "", regex=True)
    keep = ~store_key_frame(both).duplicated(keep="last")
    return both.loc[keep].reset_index(drop=True)


def _blank_editor_df(cols, numeric_cols):
    """빈 입력행 1건 — 텍스트 컬럼은 빈 문자열, 숫자 컬럼은 NaN(빈 칸)으로 표시."""
    row = {c: (np.nan if c in numeric_cols else "") for c in cols}
    d = pd.DataFrame([row])
    for c in numeric_cols:
        d[c] = pd.to_numeric(d[c])
    return d


def blank_row_df():
    """직접 입력(최종 값)용 빈 입력행 1건."""
    return _blank_editor_df(STORE_COLS, NUM_COLS)


def clean_manual_rows(edited):
    """직접 입력/편집한 표(DataFrame) → 저장 가능한 레코드만 남긴다 (AF코드·일자 필수)."""
    d = edited.copy()
    for c in STORE_COLS:
        if c not in d.columns:
            d[c] = None
    d["af"] = d["af"].apply(lambda v: "" if v is None else str(v).strip())
    d = d[d["af"] != ""]
    d["date"] = d["date"].apply(_norm_date)
    d = d.dropna(subset=["date"])
    return d[STORE_COLS]


# ── '직접 입력' 페이지의 '발송 건수'/'RawNew' 원본 형태 입력 표 — 엑셀 시트의 컬럼과 1:1로 맞춘다.
# 실제 계산(build_from_raw_query_sheets)에는 date/af/send/camp_name, date/af/chnl/uv/visit/cust/oc/amt만
# 쓰이지만, 나머지(캠페인코드/캠페인구분/AF_NM/AF_LCTGR_NM 등)도 엑셀과 똑같이 눈에 보이게 둔다.
SEND_COUNT_ENTRY_COLS = ["date", "camp_name", "camp_code", "auto_code", "channel", "af", "camp_type", "send_kind", "send"]
SEND_COUNT_ENTRY_NUM = ["send"]
RAWNEW_ENTRY_COLS = ["std_ym", "date", "af", "af_nm", "lctgr", "mctgr", "sctgr", "chnl", "uv", "visit", "cust", "oc", "sale_amt", "amt"]
RAWNEW_ENTRY_NUM = ["uv", "visit", "cust", "oc", "sale_amt", "amt"]


def clean_send_count_entry(edited):
    """'발송 건수' 직접 입력 표 → build_from_raw_query_sheets에 넣을 수 있는 형태로 정리
    (실제 계산에는 date/af/send/camp_name만 쓰지만, 나머지 컬럼도 그대로 보존해 반환한다)."""
    d = edited.copy()
    for c in SEND_COUNT_ENTRY_COLS:
        if c not in d.columns:
            d[c] = None
    d["af"] = d["af"].apply(lambda v: "" if v is None else str(v).strip())
    d = d[d["af"] != ""]
    d["date"] = d["date"].apply(_norm_date)
    d = d.dropna(subset=["date"])
    d["send"] = pd.to_numeric(d["send"], errors="coerce")
    return d[SEND_COUNT_ENTRY_COLS]


def clean_rawnew_entry(edited):
    """'RawNew' 직접 입력 표 → build_from_raw_query_sheets에 넣을 수 있는 형태로 정리
    (실제 계산에는 date/af/chnl/uv/visit/cust/oc/amt만 쓰지만, STD_YM/AF_NM 등도 그대로 보존해 반환한다)."""
    d = edited.copy()
    for c in RAWNEW_ENTRY_COLS:
        if c not in d.columns:
            d[c] = None
    d["af"] = d["af"].apply(lambda v: "" if v is None else str(v).strip())
    d = d[d["af"] != ""]
    d["date"] = d["date"].apply(_norm_date)
    d = d.dropna(subset=["date"])
    d["chnl"] = d["chnl"].apply(lambda v: "" if v is None else str(v).strip())
    for c in RAWNEW_ENTRY_NUM:
        d[c] = pd.to_numeric(d[c], errors="coerce")
    return d[RAWNEW_ENTRY_COLS]


# 메인 입력 그리드에서 자동 조회(비활성화)로 두는 컬럼 — 발송모수(send)는 '자동 조회 결과' 표에서
# 직접 수정 가능하도록 여기서는 제외한다.
AUTO_LOOKUP_COLS = ["uv", "visit", "cust", "oc", "amt", "infl_cr", "ord_cr", "eff"]


def lookup_send_uv_etc(meta_df, send_log_df, rawnew_df):
    """메타데이터 표(날짜+AF코드 등)의 각 행에 대해 '발송 건수'/'RawNew' 표에서 (일자,AF코드) 기준으로
    UV/VISIT/고객수/주문건수/거래액을 찾아 채운다. 엑셀의
    =SUMIFS(RawNew!UV, RawNew!일자, 일자, RawNew!AF코드, AF코드, RawNew!CHNL_DTL_CD,"Total") 와 동일한 방식
    (일자+AF코드 합계, RawNew는 CHNL_DTL_CD='Total' 행이 있으면 그것만 사용).

    발송모수(send)는 meta_df에 직접 입력된 값이 있으면 그 값을 그대로 쓰고, 비어 있을 때만
    '발송 건수' 표에서 조회해 채운다(직접 기입 우선).

    유입전환율=UV/발송, 주문전환율=고객수/UV, 효율=거래액/발송 으로 계산한다
    (모두 실제 소재별 실적 데이터를 역산해 확인한 공식). AF코드가 없거나 값을 못 찾으면 NaN으로 남긴다."""
    d = meta_df.copy()
    manual_send = pd.to_numeric(d["send"], errors="coerce") if "send" in d.columns else pd.Series(np.nan, index=d.index)
    for c in AUTO_LOOKUP_COLS:
        d[c] = np.nan
    if d.empty:
        d["send"] = manual_send
        return d

    send_map = {}
    if send_log_df is not None and not send_log_df.empty:
        s = send_log_df.dropna(subset=["date"]).copy()
        s = s[s["af"].astype(str).str.strip() != ""]
        if not s.empty:
            s["send"] = pd.to_numeric(s["send"], errors="coerce")
            send_map = s.groupby(["date", "af"])["send"].sum().to_dict()

    perf_map = {}
    if rawnew_df is not None and not rawnew_df.empty:
        r = rawnew_df.dropna(subset=["date"]).copy()
        r = r[r["af"].astype(str).str.strip() != ""]
        if not r.empty:
            if "sale_amt" in r.columns:
                r["amt"] = r["sale_amt"]
            for c in ("uv", "visit", "cust", "oc", "amt"):
                r[c] = pd.to_numeric(r[c], errors="coerce")
            r["chnl_norm"] = r["chnl"].astype(str).str.strip().str.lower()
            has_total = r.groupby(["date", "af"])["chnl_norm"].transform(lambda x: (x == "total").any())
            use = r[(~has_total) | (r["chnl_norm"] == "total")]
            g = use.groupby(["date", "af"]).agg(uv=("uv", "sum"), visit=("visit", "sum"), cust=("cust", "sum"),
                                                oc=("oc", "sum"), amt=("amt", "sum"))
            perf_map = g.to_dict("index")

    keys = list(zip(d["date"], d["af"].astype(str).str.strip()))
    looked_up_send = pd.Series([send_map.get(k, np.nan) for k in keys], index=d.index)
    d["send"] = manual_send.where(manual_send.notna(), looked_up_send)
    for col in ("uv", "visit", "cust", "oc", "amt"):
        d[col] = [(perf_map[k][col] if k in perf_map else np.nan) for k in keys]

    return compute_ratio_cols(d)


def compute_ratio_cols(d):
    """발송모수/UV/고객수/거래액이 채워진 표 → 유입전환율(UV/발송)/주문전환율(고객수/UV)/
    효율(거래액/발송)을 (다시) 계산해 넣는다. '자동 조회 결과'에서 발송모수를 직접 고쳤을 때
    비율을 다시 맞추는 용도로도 재사용한다."""
    d = d.copy()
    send_n = pd.to_numeric(d["send"], errors="coerce")
    uv_n = pd.to_numeric(d["uv"], errors="coerce")
    cust_n = pd.to_numeric(d["cust"], errors="coerce")
    amt_n = pd.to_numeric(d["amt"], errors="coerce")
    d["infl_cr"] = np.where(send_n.fillna(0) > 0, uv_n / send_n, np.nan)
    d["ord_cr"] = np.where(uv_n.fillna(0) > 0, cust_n / uv_n, np.nan)
    d["eff"] = np.where(send_n.fillna(0) > 0, amt_n / send_n, np.nan)
    return d


def load_store():
    if os.path.exists(DATA_STORE):
        try:
            return pd.read_csv(DATA_STORE, encoding="utf-8-sig", dtype={"date": str, "af": str})
        except Exception:
            pass
    return pd.DataFrame(columns=STORE_COLS)


def save_store(df):
    df[[c for c in STORE_COLS if c in df]].to_csv(DATA_STORE, index=False, encoding="utf-8-sig")


# ── 구글시트 영속 저장 (선택) — 미설정 시 로컬 CSV 폴백 ──────────────────
GS_TITLE = "daily_push_perf_store"


def _fix_pem(pk):
    """Secrets에서 깨진 private_key 복구 — 어떤 형태든 base64 본문을 추출해 표준 PEM으로 재구성."""
    if not isinstance(pk, str):
        return pk
    s = pk.strip().strip('"').strip("'")
    s = s.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\r\n", "\n").replace("\r", "\n")
    mt = re.search(r"-----BEGIN ([A-Z0-9 ]+?)-----(.*?)-----END \1-----", s, re.S)
    if mt:
        header, body = mt.group(1).strip(), re.sub(r"\s+", "", mt.group(2))
    else:
        header = "PRIVATE KEY"
        body = re.sub(r"\s+", "", s)
    if not body:
        return s + "\n"
    wrapped = "\n".join(body[i:i + 64] for i in range(0, len(body), 64))
    return f"-----BEGIN {header}-----\n{wrapped}\n-----END {header}-----\n"


def _pem_diag(pk):
    if not isinstance(pk, str) or not pk.strip():
        return "키가 비어있음(미입력)"
    has_b = "-----BEGIN" in pk
    has_e = "-----END" in pk
    body = re.sub(r"\s+", "", re.sub(r"-----[A-Z0-9 ]+-----", "", pk))
    n = len(body)
    tip = "정상 길이" if n >= 1500 else "너무 짧음 → 잘린 듯"
    return f"BEGIN:{'있음' if has_b else '없음'} END:{'있음' if has_e else '없음'} 본문:{n}자({tip})"


def gs_open(creds_dict, spreadsheet):
    """서비스 계정 자격으로 스프레드시트 열기 (URL/키/제목 모두 허용)."""
    import gspread
    from google.oauth2.service_account import Credentials
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    info = dict(creds_dict)
    info["private_key"] = _fix_pem(info.get("private_key"))
    try:
        creds = Credentials.from_service_account_info(info, scopes=scopes)
    except Exception as e:
        raise ValueError(f"서비스계정 private_key 문제 — {_pem_diag(info.get('private_key'))}. "
                         f"(원본오류: {str(e)[:50]})")
    gc = gspread.authorize(creds)
    sp = str(spreadsheet).strip()
    if sp.startswith("http"):
        return gc.open_by_url(sp)
    if "/" not in sp and " " not in sp and len(sp) >= 30:
        return gc.open_by_key(sp)
    try:
        return gc.open(sp)
    except Exception:
        return gc.open_by_key(sp)


def gs_read_ws(sh, title, cols):
    """워크시트가 아예 없으면(최초 사용) 빈 표를 돌려준다 — 이건 정상 케이스.
    반면 API 오류(쿼터/네트워크 등)는 여기서 삼키지 않고 그대로 올려보낸다 — 호출부가 이걸
    '데이터 없음'으로 착각해 병합·덮어쓰기하면 기존 누적 데이터가 통째로 지워지기 때문."""
    import gspread
    try:
        ws = sh.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        return pd.DataFrame(columns=cols)
    vals = ws.get_all_values()
    if not vals or len(vals) < 2:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(vals[1:], columns=vals[0])
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df[cols]


def gs_write_ws(sh, title, df, cols):
    out = df.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    out = out[cols].fillna("").astype(str)
    try:
        ws = sh.worksheet(title)
    except Exception:
        ws = sh.add_worksheet(title=title, rows=max(len(out) + 10, 100), cols=max(len(cols), 10))
    ws.clear()
    ws.update(values=[cols] + out.values.tolist(), range_name="A1")


def gs_clear_ws(sh, title, cols):
    try:
        ws = sh.worksheet(title)
        ws.clear()
        ws.update(values=[cols], range_name="A1")
    except Exception:
        pass


# ── '발송 계획' 구글시트 불러오기 (예: APP PUSH 발송 시트의 '기본 발송' 탭) ─────
# 시트 컬럼(일자/요일/시간대/타겟구분/발송유형/BPU/우선순위/카테고리/속성/담당자/브랜드/
# AF코드/기획전No./발송모수)이 COLMAP_CANDIDATES와 1:1로 맞아 parse_material_rows를 그대로 쓴다.
PLAN_SHEET_URL = "https://docs.google.com/spreadsheets/d/1xqlaRnHa5HMLz3ASUn-H7AvMkUsvQw6l7aw1rWLqfjw/edit?gid=147510377#gid=147510377"
PLAN_SHEET_GID = "147510377"
# 요일/시간대는 (병합된 셀 등으로) 비어 있을 수 있어 제외한다. 발송모수는 나중에 '자동 조회'로
# 채워지는 실적값이라 계획 단계에서는 비어 있는 게 정상이라 마찬가지로 제외한다.
PLAN_SHEET_REQUIRED = ["target", "stype", "bpu", "prio", "cat", "attr", "owner", "brand", "af", "promo"]


def _is_blank(v):
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() in ("nan", "none", "-")


def filter_complete_plan_rows(df, required=PLAN_SHEET_REQUIRED):
    """요일/시간대/발송모수 외 계획 항목이 하나라도 빈 칸인 행은 자동 불러오기에서 제외한다."""
    if df.empty:
        return df
    mask = pd.Series(True, index=df.index)
    for col in required:
        mask &= (~df[col].map(_is_blank)) if col in df.columns else False
    return df[mask].reset_index(drop=True)


def gs_worksheet_by_gid(sh, gid):
    """gid(워크시트 ID)로 특정 탭을 찾는다. 못 찾으면 None."""
    gid = str(gid)
    for ws in sh.worksheets():
        if str(ws.id) == gid:
            return ws
    return None


def load_plan_rows_from_gsheet(creds_dict, spreadsheet, gid):
    """발송 계획 시트에서 (필터 적용 전) 헤더 인식·컬럼 매핑까지 마친 원본 행을 읽어온다.
    필터는 filter_complete_plan_rows에서 별도로 건다 (진단 메시지에서 필터 전/후를 구분해 보여주기 위함).
    반환값: (파싱된 DataFrame, API가 실제로 돌려준 원본 행(rows) 리스트) — rows는 파싱이 완전히
    비었을 때(헤더를 못 찾음/시트가 진짜 비어있음) 원인을 구분해 보여주는 진단용."""
    sh = gs_open(creds_dict, spreadsheet)
    ws = gs_worksheet_by_gid(sh, gid) if gid else sh.sheet1
    if ws is None:
        raise ValueError(f"gid={gid} 탭을 찾을 수 없어요. 시트 URL의 gid를 확인해주세요.")
    rows = ws.get_all_values()
    return parse_material_rows(rows), rows


def plan_rows_blank_summary(df, required=PLAN_SHEET_REQUIRED):
    """필수 항목별로 빈 칸인 행이 몇 개인지 — 어떤 컬럼 때문에 필터링되는지 진단용."""
    n = len(df)
    recs = []
    for col in required:
        blanks = n if (n == 0 or col not in df.columns) else int(df[col].map(_is_blank).sum())
        recs.append({"컬럼": col, "빈칸 행수": blanks, "전체 행수": n})
    return pd.DataFrame(recs)


def plan_rows_to_editor_df(parsed):
    """불러온 계획 행 → 직접입력 표(STORE_COLS 전체)에 넣을 수 있는 형태로 정리."""
    d = parsed.copy()
    for c in STORE_COLS:
        if c not in d.columns:
            d[c] = np.nan if c in NUM_COLS else ""
    for c in NUM_COLS:
        d[c] = pd.to_numeric(d[c], errors="coerce")
    for c in TXT_COLS + ["af", "date"]:
        d[c] = d[c].apply(lambda v: "" if v is None else str(v).strip())
    return d[STORE_COLS].reset_index(drop=True)


# ── '요청문구' 시트 (기본값은 발송계획과 같은 스프레드시트, 다른 탭) — 기획전No.(promo) 기준으로
# 제목/내용을 매칭. 사용자가 "일자별 실적" 페이지에서 다른 스프레드시트로 바꿔 연결할 수도 있다.
REQUEST_COPY_GID = "1864635721"
REQUEST_COPY_CANDIDATES = {
    "date": ["일자(8자리)", "일자", "날짜"],
    "hour": ["시간대", "시간"],
    "promo": ["기획전번호", "기획전 번호", "우선순위기획전번호", "우선순위 기획전 번호"],
    "title": ["제목"],
    "content": ["내용"],
}


def _parse_sheet_url_gid(url):
    """구글시트 URL에서 gid를 추출한다 (예: '...edit?gid=123#gid=123' → '123'). 없으면 None."""
    m = re.search(r"[#&?]gid=(\d+)", str(url or ""))
    return m.group(1) if m else None


def _merge_header_rows(rows):
    """'요청문구' 시트처럼 헤더가 병합 셀 때문에 여러 행에 걸쳐 나뉜 경우(위 행 '요청문구'가
    제목/내용 두 칸에 걸쳐 병합, '기획전 번호'는 두 행에 걸쳐 병합 등), 각 컬럼별로 더 아래쪽
    (더 구체적인) 행의 텍스트를 우선하고, 그 자리가 비어 있을 때만 위쪽 행 텍스트로 채운다."""
    if not rows:
        return []
    width = max(len(r or []) for r in rows)
    merged = [""] * width
    for row in reversed(rows):
        row = row or []
        for i in range(width):
            v = row[i] if i < len(row) else None
            if v and not merged[i]:
                merged[i] = str(v).strip()
    return merged


def _norm_promo_val(v):
    """'J_C_1_112026'처럼 접두어가 붙은 기획전 번호도 뒤쪽 숫자만 뽑아 우리 데이터의
    기획전No.(promo, 예: '112026')와 같은 형식으로 맞춘다. 숫자로 안 끝나면 원래 텍스트 그대로."""
    s = _txt_val(v)
    if not s:
        return ""
    m = re.search(r"(\d+)$", s)
    return m.group(1) if m else s


def _norm_hour_val(v):
    """시간대 값을 정수 문자열로 맞춘다 ('1600'/'1600.0'/1600 모두 '1600')."""
    s = _txt_val(v)
    if not s:
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s


def _hhmm(h):
    """HHMM 인코딩된 시간대 값(예: 800/1350/1920) → 'HH:MM' 문자열(예: '08:00'/'13:50'/'19:20')."""
    s = f"{int(h):04d}"
    return f"{s[:2]}:{s[2:]}"


# REQUEST_COPY 조회 키 우선순위 — 시트에 있는 컬럼만큼만 써서 매칭한다(모두 있으면 3개 다 사용,
# 일자/시간대가 없으면 기획전 번호만으로 매칭 — parse_request_copy_rows의 반환값 key_cols 참고).
REQUEST_COPY_KEY_ORDER = ["date", "hour", "promo"]


def parse_request_copy_rows(rows):
    """'요청문구' 시트(일자/시간대/기획전 번호/제목/내용) → 조회용 DataFrame과, 실제로 시트에서
    찾은 매칭 키 컬럼 목록(key_cols, 예: ['date','hour','promo'] 또는 ['promo']만)을 함께 돌려준다."""
    empty = pd.DataFrame(columns=["date", "hour", "promo", "title", "content"])
    if not rows:
        return empty, []
    best_idx, best_hdr_rows, best_score = None, 0, -1
    for hdr_rows in (1, 2, 3):
        if hdr_rows > len(rows):
            break
        merged = _merge_header_rows(rows[:hdr_rows])
        idx = _map_columns(merged, candidates=REQUEST_COPY_CANDIDATES)
        if len(idx) > best_score:
            best_score, best_idx, best_hdr_rows = len(idx), idx, hdr_rows
    if best_idx is None or "promo" not in best_idx:
        return empty, []
    idx = best_idx
    key_cols = [c for c in REQUEST_COPY_KEY_ORDER if c in idx]
    recs = []
    for row in rows[best_hdr_rows:]:
        if row is None:
            continue
        n = len(row)
        promo = _norm_promo_val(row[idx["promo"]] if idx["promo"] < n else None)
        if not promo:
            continue
        rec = {"promo": promo,
               "title": _txt_val(row[idx["title"]]) if "title" in idx and idx["title"] < n else "",
               "content": _txt_val(row[idx["content"]]) if "content" in idx and idx["content"] < n else ""}
        if "date" in idx:
            rec["date"] = _norm_date(row[idx["date"]] if idx["date"] < n else None) or ""
        if "hour" in idx:
            rec["hour"] = _norm_hour_val(row[idx["hour"]] if idx["hour"] < n else None)
        recs.append(rec)
    if not recs:
        return empty, []
    out = pd.DataFrame(recs).drop_duplicates(subset=key_cols, keep="last")
    return out, key_cols


# ══════════════════════════════════════════════════════════════════════
# 2. Streamlit 앱
# ══════════════════════════════════════════════════════════════════════
METRIC_LABELS = {"send": "발송모수", "uv": "UV", "oc": "주문건수", "amt": "거래액"}
RATE_LABELS = {"ctr": "CTR(UV/발송)", "cvr": "주문전환율(주문/UV)", "rps": "RPS(발송당거래액)", "aov": "객단가(거래액/주문)"}


def main():
    import streamlit as st
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    st.set_page_config(page_title="일일 PUSH 발송성과 대시보드", layout="wide", initial_sidebar_state="expanded")
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"]{background:#f8f9fc}
    [data-testid="stSidebar"]{background:#ffffff;border-right:1px solid #e2e8f0}
    [data-testid="stMetric"]{background:#ffffff;border-radius:8px;padding:12px 16px;border:1px solid #e2e8f0;box-shadow:0 1px 3px rgba(0,0,0,0.06)}
    [data-testid="stMetricLabel"]{color:#64748b!important;font-size:12px!important}
    [data-testid="stMetricValue"]{color:#1e293b!important;font-size:20px!important}
    h1,h2,h3{color:#1e293b}
    </style>""", unsafe_allow_html=True)

    def won(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "–"
        if abs(v) >= 1e8:
            return f"{v/1e8:.2f}억"
        if abs(v) >= 1e4:
            return f"{v/1e4:,.0f}만"
        return f"{v:,.0f}"

    def base_layout(h=340, title=""):
        return dict(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#475569", size=11), height=h,
                    margin=dict(l=10, r=10, t=(72 if title else 20), b=10),
                    title=dict(text=title, font=dict(color="#94a3b8", size=13), x=0, xanchor="left",
                               y=0.99, yanchor="top"),
                    legend=dict(orientation="h", yanchor="bottom", y=1.1, xanchor="left", x=0),
                    xaxis=dict(gridcolor="rgba(0,0,0,0)", linecolor="#e2e8f0"),
                    yaxis=dict(gridcolor="#f1f5f9", linecolor="#e2e8f0"))

    # ── AI 대화 (Gemini) — 지표를 근거로 논의하는 채팅 기능에서 공용으로 사용 ──
    # 모델명은 구글이 수시로 세대교체·구버전 폐지를 하므로(예: gemini-2.5-pro가 신규 키에는
    # 막히고 gemini-3.1-pro-preview로 대체된 사례), ai_chat_generate가 404/NOT_FOUND 에러 메시지에
    # 적힌 대체 모델명을 자동으로 파싱해 한 번 더 시도한다 — 하드코딩한 이름이 나중에 또 막혀도
    # 코드를 안 고쳐도 되게 하기 위함.
    AI_MODELS = {
        "Gemini 2.5 Flash (균형)": "gemini-2.5-flash",
        "Gemini 3.1 Pro Preview (최고 품질·무료 사용량 없을 수 있음)": "gemini-3.1-pro-preview",
        "Gemini 2.5 Flash-Lite (빠름·저렴)": "gemini-2.5-flash-lite",
    }
    AI_FREE_TIER_SAFE_MODEL = "gemini-2.5-flash"  # Pro 계열은 무료 쿼터가 0인 경우가 많아 폴백 대상으로 씀

    def gemini_key():
        for k in ("GEMINI_API_KEY", "GOOGLE_API_KEY"):
            try:
                if k in st.secrets:
                    return st.secrets[k]
            except Exception:
                pass
            v = os.environ.get(k)
            if v:
                return v
        return None

    def ai_chat_generate(system, history, model):
        """history: [(role, text), ...] role은 'user' 또는 'assistant'.
        반환값: (텍스트, 에러메시지, 실제로 사용된 모델명)."""
        key = gemini_key()
        if not key:
            return None, "GEMINI_API_KEY 미설정 — Streamlit Secrets 또는 환경변수에 추가하세요.", model
        try:
            from google import genai
            from google.genai import types
        except ImportError:
            return None, "google-genai 패키지가 없습니다. requirements.txt 반영 후 재배포하세요.", model

        client = genai.Client(api_key=key)
        contents = [
            types.Content(role=("user" if role == "user" else "model"), parts=[types.Part(text=text)])
            for role, text in history
        ]

        def _call(m):
            resp = client.models.generate_content(
                model=m, contents=contents,
                config=types.GenerateContentConfig(system_instruction=system, max_output_tokens=4000),
            )
            return (resp.text or "").strip()

        def _friendly_error(msg):
            if "RESOURCE_EXHAUSTED" in msg or "429" in msg:
                wait = re.search(r"retry in ([\d.]+)s", msg)
                wait_txt = f" 약 {round(float(wait.group(1)))}초 후 다시 시도해보세요." if wait else " 잠시 후 다시 시도해보세요."
                return (f"⏳ 이 모델의 무료 사용량 한도를 초과했어요.{wait_txt} 계속 막히면 더 가벼운 모델"
                       "(Flash/Flash-Lite)로 바꾸거나, Google AI Studio에서 결제를 연결해보세요.")
            return f"생성 오류: {msg}"

        try:
            text = _call(model)
            return (text or None), (None if text else "빈 응답"), model
        except Exception as e:
            msg = str(e)
            # 429(RESOURCE_EXHAUSTED)면 재시도해봐야 소용없는 경우가 많음(무료 쿼터 자체가 0인 모델,
            # 특히 Pro 계열) — 재시도 대신 무료 쿼터가 보통 열려 있는 Flash로 한 번 자동 전환해본다.
            if ("RESOURCE_EXHAUSTED" in msg or "429" in msg) and model != AI_FREE_TIER_SAFE_MODEL:
                try:
                    text = _call(AI_FREE_TIER_SAFE_MODEL)
                    return (text or None), (None if text else "빈 응답"), AI_FREE_TIER_SAFE_MODEL
                except Exception as e2:
                    return None, _friendly_error(str(e2)), model
            # 404(NOT_FOUND)는 보통 "model models/gemini-2.5-pro is no longer available ...
            # use models/gemini-3.1-pro-preview" 처럼 막힌 모델명이 먼저, 대체 모델명이 나중에 나온다 —
            # 그래서 마지막으로 언급된 모델명을 대체 후보로 쓴다(첫 번째로 하면 막힌 모델을 또 시도하게 됨).
            if "NOT_FOUND" in msg or "404" in msg:
                alts = re.findall(r"models/([\w.\-]+)", msg)
                alt_model = alts[-1] if alts else None
                if alt_model and alt_model != model:
                    try:
                        text = _call(alt_model)
                        return (text or None), (None if text else "빈 응답"), alt_model
                    except Exception as e2:
                        return None, _friendly_error(str(e2)), model
            return None, _friendly_error(msg), model

    # ── 표 입력용 컬럼 설정 — 직접 입력 / 데이터 편집 화면에서 공용으로 사용 ──
    TXT_LABELS = {"date": "일자(YYYYMMDD)*", "dow_k": "요일", "target": "타겟구분", "stype": "발송유형",
                  "bpu": "BPU", "cat": "카테고리", "attr": "속성", "owner": "담당자", "brand": "브랜드",
                  "af": "AF코드*", "promo": "기획전"}
    NUM_LABELS = {"hour": "시간대", "prio": "우선순위", "send": "발송모수", "uv": "UV", "visit": "VISIT",
                  "cust": "고객수", "oc": "주문건수", "amt": "거래액", "infl_cr": "유입전환율",
                  "ord_cr": "주문전환율", "eff": "효율"}

    PCT_NUM_COLS = {"infl_cr", "ord_cr"}

    def store_column_config(disabled=()):
        cfg = {}
        for k, label in TXT_LABELS.items():
            cfg[k] = st.column_config.TextColumn(label, disabled=(k in disabled))
        for k, label in NUM_LABELS.items():
            fmt = "percent" if k in PCT_NUM_COLS else None
            cfg[k] = st.column_config.NumberColumn(label, disabled=(k in disabled), format=fmt)
        return cfg

    # ── 저장소 백엔드: 구글시트(설정 시) ↔ 로컬 CSV(폴백) ──
    @st.cache_resource(show_spinner=False)
    def _get_sh(_email, spreadsheet):
        return gs_open(st.secrets["gcp_service_account"], spreadsheet)

    @st.cache_data(ttl=300, show_spinner=False)
    def _load_request_copy_cached(_email, spreadsheet, gid):
        sh = _get_sh(_email, spreadsheet)
        ws = gs_worksheet_by_gid(sh, gid) if gid else sh.sheet1
        if ws is None:
            return pd.DataFrame(columns=["date", "hour", "promo", "title", "content"]), []
        return parse_request_copy_rows(ws.get_all_values())

    def request_copy_source():
        """세션에서 사용자가 직접 연결한 시트가 있으면 그걸, 없으면 기본 시트를 돌려준다.
        직접 연결한 시트인데 URL에 gid가 없으면(gid=None) 그 시트의 첫 번째 탭을 쓴다 —
        기본 시트 전용 gid(REQUEST_COPY_GID)로 잘못 폴백하면 안 되므로 분기를 분리한다."""
        is_custom = "request_copy_url" in st.session_state
        if is_custom:
            return st.session_state.get("request_copy_url") or PLAN_SHEET_URL, \
                   st.session_state.get("request_copy_gid"), True
        return PLAN_SHEET_URL, REQUEST_COPY_GID, False

    def plan_sheet_source():
        """발송 계획 시트도 요청문구 시트와 같은 방식으로 세션에서 직접 연결한 시트를 우선한다.
        발송 계획은 스프레드시트 하나에 주차별로 탭이 나뉘어 있어(예: '7월 1주차(7/6~7/12)'),
        기본 gid 하나만으로는 최신 주차를 못 읽어올 수 있어 사용자가 원하는 주차 탭 URL로 바꿀 수 있게 한다."""
        is_custom = "plan_sheet_url" in st.session_state
        if is_custom:
            return st.session_state.get("plan_sheet_url") or PLAN_SHEET_URL, \
                   st.session_state.get("plan_sheet_gid"), True
        return PLAN_SHEET_URL, PLAN_SHEET_GID, False

    def load_request_copy_map():
        """'요청문구' 시트를 읽어 {키튜플: {title, content}} 형태와, 실제로 매칭에 쓸 키 컬럼
        목록(예: ['date','hour','promo'] — 시트에 일자/시간대가 없으면 ['promo']만)을 함께 돌려준다.
        시크릿 미설정/조회 실패 시 빈 dict + 에러 메시지(있으면)를 함께 준다."""
        try:
            has_creds = "gcp_service_account" in st.secrets
        except Exception:
            has_creds = False
        if not has_creds:
            return {}, [], None
        url, gid, _ = request_copy_source()
        try:
            rc_df, key_cols = _load_request_copy_cached(
                st.secrets["gcp_service_account"].get("client_email", ""), url, gid)
        except Exception as e:
            return {}, [], (str(e)[:120].strip() or type(e).__name__)
        if rc_df.empty or not key_cols:
            return {}, [], None
        keys = list(rc_df[key_cols].itertuples(index=False, name=None))
        values = rc_df[["title", "content"]].to_dict("records")
        return dict(zip(keys, values)), key_cols, None

    def init_storage():
        try:
            has = "gcp_service_account" in st.secrets
        except Exception:
            has = False
        if not has:
            return {"mode": "local", "status": "💾 로컬 CSV로 저장해요 (구글시트 미설정)"}
        try:
            sp = None
            if "gsheets" in st.secrets:
                sp = st.secrets["gsheets"].get("spreadsheet")
            sp = sp or st.secrets.get("gsheets_spreadsheet")
            if not sp:
                return {"mode": "local", "status": "⚠️ gsheets.spreadsheet 미설정 → 로컬에 저장해요"}
            sh = _get_sh(st.secrets["gcp_service_account"].get("client_email", ""), sp)
            return {"mode": "gsheets", "sh": sh, "status": "☁️ 구글시트에 연결됐어요"}
        except Exception as e:
            msg = str(e)[:80].strip() or f"{type(e).__name__} (메시지 없음 — 시트 공유/ID를 확인하세요)"
            return {"mode": "local", "status": f"⚠️ 구글시트 연결 실패 → 로컬에 저장해요 ({msg})"}

    def storage_load(bk):
        """gsheets 모드에서 읽기 실패는 여기서 감추지 않고 그대로 예외를 던진다 — 호출부가
        '빈 데이터'로 착각해 병합 후 저장하면 기존 누적 데이터를 지워버릴 수 있기 때문.
        호출부에서 반드시 실패 시 저장을 중단하도록 처리해야 한다."""
        if bk["mode"] == "gsheets":
            return gs_read_ws(bk["sh"], GS_TITLE, STORE_COLS)
        return load_store()

    def storage_save(bk, df):
        if bk["mode"] == "gsheets":
            try:
                gs_write_ws(bk["sh"], GS_TITLE, df, STORE_COLS)
                return
            except Exception as e:
                save_store(df)
                msg = str(e)[:200].strip() or type(e).__name__
                st.warning(f"⚠️ 구글시트 저장 실패 → 로컬 CSV에 저장했어요 ({msg}). "
                          "구글시트 API 사용량 제한(잠시 후 재시도)이거나 시트 공유 권한 문제일 수 있어요.")
                return
        save_store(df)

    def storage_clear(bk):
        if bk["mode"] == "gsheets":
            gs_clear_ws(bk["sh"], GS_TITLE, STORE_COLS)
        elif os.path.exists(DATA_STORE):
            os.remove(DATA_STORE)

    BK = init_storage()

    @st.cache_data(show_spinner=False)
    def cached_load_excel(b):
        return load_excel_perf(b)

    # ══════════════════════════════════════════════════════════
    # 사이드바 — 업로드 / 필터
    # ══════════════════════════════════════════════════════════
    st.sidebar.markdown("### 📲 일일 PUSH 발송성과")
    st.sidebar.caption(BK["status"])

    uploaded = st.sidebar.file_uploader(
        "📂 일일 PUSH 실적.xlsx 업로드", type=["xlsx"], accept_multiple_files=True, key="perf_up",
        help="'누적_소재별'/'소재별 실적(당주)_2' 시트를 우선 찾아 읽고, 없거나 그 시트에 없는 "
             "일자·AF코드는 '발송 건수'+'RawNew' 원본 쿼리 시트에서 직접 집계해요. "
             "RawNew는 용량이 크면(수십만 행) 반영에 10~20초 걸릴 수 있어요.")

    c1, c2 = st.sidebar.columns(2)
    apply_clicked = c1.button("📥 반영하기", use_container_width=True, disabled=not uploaded)
    clear_clicked = c2.button("🗑️ 초기화", use_container_width=True)

    if clear_clicked:
        st.session_state.confirm_clear = True

    if st.session_state.get("confirm_clear"):
        st.sidebar.warning("⚠️ 저장된 데이터를 전부 지울까요? 되돌릴 수 없어요.")
        cc1, cc2 = st.sidebar.columns(2)
        if cc1.button("✅ 네, 삭제", use_container_width=True, key="confirm_clear_yes"):
            storage_clear(BK)
            st.session_state.confirm_clear = False
            st.sidebar.success("저장된 데이터를 초기화했어요.")
            st.rerun()
        if cc2.button("취소", use_container_width=True, key="confirm_clear_no"):
            st.session_state.confirm_clear = False
            st.rerun()

    if apply_clicked and uploaded:
        try:
            old = storage_load(BK)
        except Exception as e:
            st.sidebar.error(f"⚠️ 기존 누적 데이터를 불러오지 못해 반영을 중단했어요 ({str(e)[:150].strip()}). "
                             "이 상태로 반영하면 기존 데이터가 사라질 수 있어요 — 새로고침 후 다시 시도해 주세요.")
        else:
            frames, all_used = [], []
            with st.sidebar.status("엑셀 읽는 중... (RawNew가 크면 시간이 걸려요)", expanded=False) as _status:
                for f in uploaded:
                    d, used = cached_load_excel(f.getvalue())
                    if not d.empty:
                        frames.append(d)
                        all_used.extend([f"{f.name} · {u}" for u in used])
                _status.update(label="완료", state="complete")
            if not frames:
                st.sidebar.error("인식 가능한 시트를 찾지 못했어요. '누적_소재별'/'소재별 실적(당주)' 시트가 있는지 확인해 주세요.")
            else:
                new = pd.concat(frames, ignore_index=True)
                merged = merge_store(old, new)
                storage_save(BK, merged)
                st.sidebar.success(f"{len(new):,}건 반영 완료 (누적 {len(merged):,}건)")
                with st.sidebar.expander("읽은 시트"):
                    for u in all_used:
                        st.caption(u)

    try:
        raw = storage_load(BK)
    except Exception as e:
        st.error(f"⚠️ 구글시트에서 데이터를 불러오지 못했어요 ({str(e)[:200].strip()}). "
                 "이 상태에서 저장하면 기존 누적 데이터가 사라질 수 있어 안전을 위해 여기서 멈춰요. "
                 "새로고침해서 다시 시도해 주세요 — 계속 실패하면 구글시트 API 사용량 제한이거나 "
                 "시트 공유 권한 문제일 수 있어요.")
        st.stop()
    df = _finalize(raw)

    st.sidebar.markdown("---")
    st.sidebar.markdown("**필터**")
    if not df.empty:
        dmin, dmax = df["dt"].min().date(), df["dt"].max().date()
        default_end = min(dmax, datetime.date.today())
        if default_end < dmin:
            default_end = dmax
        date_range = st.sidebar.date_input("기간", value=(dmin, default_end), min_value=dmin, max_value=dmax)
        bpu_sel = st.sidebar.multiselect("BPU", sorted([b for b in df["bpu"].unique() if b]))
        stype_sel = st.sidebar.multiselect("발송유형", sorted([s for s in df["stype"].unique() if s]))
        search = st.sidebar.text_input("검색 (브랜드/AF코드/담당자)", "")

        f = df.copy()
        if isinstance(date_range, tuple) and len(date_range) == 2:
            f = f[(f["dt"].dt.date >= date_range[0]) & (f["dt"].dt.date <= date_range[1])]
        if bpu_sel:
            f = f[f["bpu"].isin(bpu_sel)]
        if stype_sel:
            f = f[f["stype"].isin(stype_sel)]
        if search:
            q = search.lower()
            hay = (f["brand"] + " " + f["af"] + " " + f["owner"]).str.lower()
            f = f[hay.str.contains(re.escape(q), na=False)]
        st.sidebar.caption(f"필터 후 {len(f):,}건 · 전체 {len(df):,}건")
    else:
        st.sidebar.caption("아직 데이터가 없어요. '✍️ 직접 입력'에서 바로 추가하거나 엑셀을 업로드해 보세요.")
        f = df.copy()

    pages = ["주간보고", "종합요약", "UV 유입 분석", "BPU별 실적", "발송유형별 실적", "캠페인별 실적", "일자별 실적",
             "주차별 누적 추이", "우수발송 대시보드", "✍️ 직접 입력", "데이터"]
    page = st.sidebar.radio("페이지", pages, index=(pages.index("✍️ 직접 입력") if df.empty else 0))

    if df.empty and page not in ("✍️ 직접 입력", "데이터"):
        st.title("일일 PUSH 발송성과 대시보드")
        st.info("왼쪽에서 '✍️ 직접 입력'으로 바로 기록하거나, '일일 PUSH 실적.xlsx'를 업로드하고 **반영하기**를 눌러 주세요.")
        with st.expander("기대하는 파일 형식"):
            st.markdown("""
**우선순위 1 — 이미 집계된 시트가 있으면 그대로 사용**
- 시트명에 **'누적_소재별'** 또는 **'소재별 실적(당주)'**가 포함된 시트를 찾습니다 (`_2` 등 접미사 있어도 인식).
- 헤더 컬럼: 일자(8자리)/요일/시간대/타겟구분/발송유형/BPU/우선순위/카테고리/속성/담당자/브랜드/AF코드/기획전No. +
  발송(모수)/UV/VISIT/고객수/주문건수/거래액(주문금액)/유입전환율/주문전환율/효율

**우선순위 2 — 위 시트가 없거나 거기 없는 일자·AF코드는 원본 쿼리에서 직접 집계**
- **'발송 건수'** 시트: 발송일시/캠페인명/AF코드/모수 컬럼 → (일자,AF코드)별 발송모수로 집계
- **'RawNew'** 시트: STD_DD/AF_CD/CHNL_DTL_CD/UV/SV/CUST_CNT/ORD_CNT/SALE_AMT 컬럼 → (일자,AF코드)별
  UV/VISIT/고객수/주문건수/거래액으로 집계 (CHNL_DTL_CD='Total' 행이 있으면 그것만 사용해 중복 합산 방지)
- 같은 이름의 RawNew 시트가 여러 개 있어도(예: 트레일링 스페이스) 날짜가 최신인 쪽을 자동으로 골라요.
- BPU/발송유형/기획전No.는 캠페인명이 `이름_일자_시간_발송유형_BPU_기획전No` 패턴이면 자동 추출하고,
  아니면 기존 소재별 실적 시트에 있는 같은 AF코드의 최근 값을 참고해요. 둘 다 없으면 빈 값으로 남아요.
- AF코드가 비어있는 행은 자동으로 제외돼요.
            """)
        return

    # ══════════════════════════════════════════════════════════
    # 주간보고
    # ══════════════════════════════════════════════════════════
    if page == "주간보고":
        st.title("주간보고")
        st.caption("기준 주차(월~일) 실적을 전주·전월 동주·전년 동주와 비교하고, 월 누계(MTD)는 전월·전년 같은 "
                  "기간과 비교해요. 모든 값은 합산(가중) 기준이에요. 사이드바 필터가 반영돼요.")

        weeks_sorted = f[["week_start", "week_label"]].drop_duplicates().sort_values("week_start", ascending=False)
        week_options = weeks_sorted["week_label"].tolist()
        if not week_options:
            st.info("표시할 주차 데이터가 없어요.")
        else:
            pick = st.selectbox("기준 주차", week_options, key="wr_week_pick",
                                help=("**비교 기준**\n\n"
                                      "- **전주**: 기준 주차 바로 앞 1주(월~일)\n"
                                      "- **전월 동주**: 기준 주차 시작일(월요일)의 한 달 전 날짜가 속한 주(월~일)\n"
                                      "- **전년 동주**: 기준 주차와 요일이 맞도록 364일(52주) 전 같은 주\n"
                                      "- 해당 기간에 데이터가 없으면 '-'로 표시돼요\n"
                                      "- **✱**: CTR·주문CR의 증감이 통계적으로 유의미함(이표본 비율 z검정, p<0.05)"))
            cur_start = weeks_sorted.loc[weeks_sorted["week_label"] == pick, "week_start"].iloc[0]
            cur_end = cur_start + pd.Timedelta(days=6)

            def _period_bounds(base_start, weeks_back=None, months_back=None):
                base_ts = pd.Timestamp(base_start)
                if weeks_back is not None:
                    s = (base_ts - pd.Timedelta(days=7 * weeks_back)).date()
                else:
                    d = base_ts - pd.DateOffset(months=months_back)
                    s = (d - pd.Timedelta(days=d.weekday())).date()
                return s, s + pd.Timedelta(days=6)

            prev_start, prev_end = _period_bounds(cur_start, weeks_back=1)
            pmonth_start, pmonth_end = _period_bounds(cur_start, months_back=1)
            pyear_start, pyear_end = _period_bounds(cur_start, weeks_back=52)

            def _agg(s, e):
                sub = f[(f["dt"].dt.date >= s) & (f["dt"].dt.date <= e)]
                if sub.empty:
                    return None
                send, uv, oc, amt = sub["send"].sum(), sub["uv"].sum(), sub["oc"].sum(), sub["amt"].sum()
                return {
                    "n_camp": sub[["af", "brand"]].drop_duplicates().shape[0],
                    "send": send, "uv": uv, "oc": oc, "amt": amt,
                    "ctr": (uv / send if send else 0.0), "cr": (oc / uv if uv else 0.0),
                    "rps": (amt / send if send else 0.0), "aov": (amt / oc if oc else 0.0),
                }

            cur = _agg(cur_start, cur_end)
            prev = _agg(prev_start, prev_end)
            pmonth = _agg(pmonth_start, pmonth_end)
            pyear = _agg(pyear_start, pyear_end)

            def _ztest_pvalue(x1, n1, x2, n2):
                if not n1 or not n2:
                    return None
                p1, p2 = x1 / n1, x2 / n2
                p_pool = (x1 + x2) / (n1 + n2)
                var = p_pool * (1 - p_pool) * (1 / n1 + 1 / n2)
                if var <= 0:
                    return None
                z = (p1 - p2) / (var ** 0.5)
                return 2 * (1 - 0.5 * (1 + math.erf(abs(z) / (2 ** 0.5))))

            def _sig(a, b, num_key, den_key):
                if not a or not b:
                    return False
                p = _ztest_pvalue(a[num_key], a[den_key], b[num_key], b[den_key])
                return p is not None and p < 0.05

            def _delta_pct_txt(cur_v, prior_v):
                if cur_v is None or not prior_v:
                    return ":gray[–]"
                pct = (cur_v - prior_v) / prior_v * 100
                color = "red" if pct > 0 else ("blue" if pct < 0 else "gray")
                arrow = "▲" if pct > 0 else ("▽" if pct < 0 else "-")
                return f":{color}[{arrow}{abs(pct):.1f}%]"

            def _delta_pp_txt(cur_v, prior_v, sig=False):
                if cur_v is None or prior_v is None:
                    return ":gray[–]"
                pp = (cur_v - prior_v) * 100
                color = "red" if pp > 0 else ("blue" if pp < 0 else "gray")
                sign = "+" if pp >= 0 else ""
                mark = " ✱" if sig else ""
                return f":{color}[{sign}{pp:.2f}%p{mark}]"

            cards = [
                ("발송", (f"{cur['send']:,.0f}" if cur else "-"), "send", False),
                ("UV", (f"{cur['uv']:,.0f}" if cur else "-"), "uv", False),
                ("CTR", (f"{cur['ctr']:.2%}" if cur else "-"), "ctr", True),
                ("주문CR", (f"{cur['cr']:.2%}" if cur else "-"), "cr", True),
                ("거래액", (won(cur['amt']) if cur else "-"), "amt", False),
                ("RPS", (f"{cur['rps']:,.0f}" if cur else "-"), "rps", False),
            ]
            cols = st.columns(len(cards))
            for col, (label, val_str, key, is_pp) in zip(cols, cards):
                col.markdown(f"<div style='color:#64748b;font-size:12px'>{label}</div>", unsafe_allow_html=True)
                col.markdown(f"<div style='font-size:26px;font-weight:700;color:#1e293b'>{val_str}</div>",
                            unsafe_allow_html=True)
                if cur and is_pp:
                    sig_prev = _sig(cur, prev, "uv" if key == "ctr" else "oc", "send" if key == "ctr" else "uv")
                    sig_year = _sig(cur, pyear, "uv" if key == "ctr" else "oc", "send" if key == "ctr" else "uv")
                    col.caption(_delta_pp_txt(cur[key], prev[key] if prev else None, sig_prev) + " 전주 대비")
                    col.caption(_delta_pp_txt(cur[key], pyear[key] if pyear else None, sig_year) + " 전년 대비")
                elif cur:
                    col.caption(_delta_pct_txt(cur[key], prev[key] if prev else None) + " 전주 대비")
                    col.caption(_delta_pct_txt(cur[key], pyear[key] if pyear else None) + " 전년 대비")

            st.markdown("#### 📋 주요 지표 현황")
            ROW_SPECS = [
                ("n_camp", "캠페인수", False, lambda v: f"{v:,.0f}"),
                ("send", "발송", False, lambda v: f"{v:,.0f}"),
                ("uv", "UV", False, lambda v: f"{v:,.0f}"),
                ("oc", "주문건수", False, lambda v: f"{v:,.0f}"),
                ("amt", "거래액", False, won),
                ("ctr", "CTR", True, lambda v: f"{v:.2%}"),
                ("cr", "주문CR", True, lambda v: f"{v:.2%}"),
                ("rps", "RPS", False, lambda v: f"{v:,.0f}"),
                ("aov", "객단가", False, won),
            ]

            def _cell(v, fmt):
                return fmt(v) if v is not None else "-"

            rows_html = []
            for key, label, is_pp, fmt in ROW_SPECS:
                cur_v = cur[key] if cur else None
                prev_v = prev[key] if prev else None
                pmonth_v = pmonth[key] if pmonth else None
                pyear_v = pyear[key] if pyear else None
                if is_pp:
                    sig_prev = _sig(cur, prev, "uv" if key == "ctr" else "oc", "send" if key == "ctr" else "uv")
                    sig_pmonth = _sig(cur, pmonth, "uv" if key == "ctr" else "oc", "send" if key == "ctr" else "uv")
                    sig_pyear = _sig(cur, pyear, "uv" if key == "ctr" else "oc", "send" if key == "ctr" else "uv")
                    d_prev = _delta_pp_txt(cur_v, prev_v, sig_prev)
                    d_pmonth = _delta_pp_txt(cur_v, pmonth_v, sig_pmonth)
                    d_pyear = _delta_pp_txt(cur_v, pyear_v, sig_pyear)
                else:
                    d_prev = _delta_pct_txt(cur_v, prev_v)
                    d_pmonth = _delta_pct_txt(cur_v, pmonth_v)
                    d_pyear = _delta_pct_txt(cur_v, pyear_v)
                rows_html.append({
                    "지표": label, "전주": _cell(prev_v, fmt), "기준주": _cell(cur_v, fmt), "전주비": d_prev,
                    "전월 동주": _cell(pmonth_v, fmt), "전월비": d_pmonth,
                    "전년 동주": _cell(pyear_v, fmt), "전년비": d_pyear,
                })
            report_df = pd.DataFrame(rows_html)

            def _md_color(s):
                m = re.match(r":(\w+)\[(.*)\]", s)
                return f"<span style='color:{m.group(1)}'>{m.group(2)}</span>" if m else s

            html_cols = ["지표", "전주", "기준주", "전주비", "전월 동주", "전월비", "전년 동주", "전년비"]
            thead = "".join(
                f"<th style='padding:6px 12px;text-align:{'left' if c == '지표' else 'right'};"
                f"border-bottom:2px solid #e2e8f0'>{c}</th>" for c in html_cols)
            trs = []
            for _, r in report_df.iterrows():
                tds = []
                for c in html_cols:
                    align = "left" if c == "지표" else "right"
                    val = _md_color(r[c]) if c in ("전주비", "전월비", "전년비") else r[c]
                    tds.append(f"<td style='padding:6px 12px;text-align:{align};border-bottom:1px solid #f1f5f9'>{val}</td>")
                trs.append(f"<tr>{''.join(tds)}</tr>")
            table_html = (f"<table style='width:100%;border-collapse:collapse;font-size:14px'>"
                          f"<thead><tr>{thead}</tr></thead><tbody>{''.join(trs)}</tbody></table>")
            st.markdown(table_html, unsafe_allow_html=True)

            st.caption((f"기준주 {cur_start.strftime('%Y')}년 {pick} · 전년 동주 {pyear_start}~{pyear_end} — "
                       "해당 기간에 데이터가 없으면 '-'로 표시돼요. 전월비는 기준주 시작일의 한 달 전 날짜가 속한 "
                       "주(전월 동주)와 비교해요. ✱ = CTR·주문CR의 증감이 통계적으로 유의(p<0.05).").replace("~", "\\~"))

            st.markdown("---")
            st.markdown("#### 🔍 주간 트렌드 리뷰")
            st.caption("목표: 발송량은 줄이면서(피로도↓) UV·CTR은 높이는 것 — 최근 주차 흐름을 이 기준으로 짚어봐요.")

            trend_n = 8
            wk_trend = f[f["week_start"] <= cur_start].groupby(
                ["week_start", "week_label"], dropna=False
            ).agg(send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum"),
                  n_camp=("af", "nunique")).reset_index().sort_values("week_start").tail(trend_n)
            wk_trend["ctr"] = np.where(wk_trend["send"] > 0, wk_trend["uv"] / wk_trend["send"], np.nan)

            if len(wk_trend) < 2:
                st.info("흐름을 분석하려면 최소 2주 이상의 데이터가 필요해요.")
            else:
                def _week_trend(vals):
                    vals = [v for v in vals if v is not None and not (isinstance(v, float) and np.isnan(v))]
                    if len(vals) < 2:
                        return None
                    diffs = np.diff(vals)
                    ups, downs = int((diffs > 0).sum()), int((diffs < 0).sum())
                    pct = (vals[-1] - vals[0]) / vals[0] * 100 if vals[0] else None
                    if ups and downs and abs(ups - downs) <= 1:
                        pattern = "등락 반복"
                    elif ups >= downs:
                        pattern = "상승"
                    else:
                        pattern = "하락"
                    return {"first": vals[0], "last": vals[-1], "pct": pct, "pattern": pattern}

                t_send = _week_trend(wk_trend["send"].tolist())
                t_uv = _week_trend(wk_trend["uv"].tolist())
                t_ctr = _week_trend(wk_trend["ctr"].tolist())
                t_camp = _week_trend(wk_trend["n_camp"].tolist())
                span = (f"{wk_trend.iloc[0]['week_label']} ~ {wk_trend.iloc[-1]['week_label']} "
                       f"({len(wk_trend)}주)").replace("~", "\\~")

                def _fmt_pct(t):
                    return f"{t['pct']:+.1f}%" if t and t["pct"] is not None else "–"

                st.markdown(f"**{span} 흐름**")

                def _goal_color(t, good_pattern):
                    if not t:
                        return "#94a3b8"
                    if t["pattern"] == good_pattern:
                        return "#22a55e"
                    if t["pattern"] == "등락 반복":
                        return "#f59e0b"
                    return "#ef4444"

                send_c = _goal_color(t_send, "하락")
                uv_c = _goal_color(t_uv, "상승")
                ctr_c = _goal_color(t_ctr, "상승")

                trend_fig = make_subplots(rows=1, cols=3,
                                          subplot_titles=["발송량 (↓ 목표)", "UV (↑ 목표)", "CTR (↑ 목표)"])
                trend_fig.add_trace(go.Bar(x=wk_trend["week_label"], y=wk_trend["send"],
                                           marker_color=send_c, showlegend=False), row=1, col=1)
                trend_fig.add_trace(go.Scatter(x=wk_trend["week_label"], y=wk_trend["uv"], mode="lines+markers",
                                               line=dict(color=uv_c, width=2), marker=dict(size=6),
                                               showlegend=False), row=1, col=2)
                trend_fig.add_trace(go.Scatter(x=wk_trend["week_label"], y=wk_trend["ctr"], mode="lines+markers",
                                               line=dict(color=ctr_c, width=2), marker=dict(size=6),
                                               showlegend=False), row=1, col=3)
                trend_fig.update_yaxes(tickformat=".1%", row=1, col=3)
                trend_fig.update_layout(height=260, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                        font=dict(color="#475569", size=11), margin=dict(l=10, r=10, t=40, b=10))
                trend_fig.update_xaxes(gridcolor="rgba(0,0,0,0)", linecolor="#e2e8f0", tickangle=-30)
                trend_fig.update_yaxes(gridcolor="#f1f5f9", linecolor="#e2e8f0")
                for ann in trend_fig["layout"]["annotations"]:
                    ann["font"] = dict(color="#94a3b8", size=13)
                st.plotly_chart(trend_fig, use_container_width=True)
                st.caption("🟢 초록 = 목표 방향(발송↓/UV·CTR↑), 🟠 주황 = 등락 반복, 🔴 빨강 = 목표 반대 방향.")

                if t_send:
                    good = t_send["pattern"] == "하락"
                    icon = "✅" if good else ("⚠️" if t_send["pattern"] == "등락 반복" else "❌")
                    comment = ("목표대로 줄이는 흐름이에요." if good else
                              "일정하지 않게 오르내려요." if t_send["pattern"] == "등락 반복" else
                              "오히려 늘고 있어요 — 감축이 필요해요.")
                    st.markdown(f"{icon} **발송량**: {t_send['first']:,.0f}건 → {t_send['last']:,.0f}건 "
                               f"({_fmt_pct(t_send)}, {t_send['pattern']}) — {comment}")
                if t_camp:
                    st.markdown(f"　· 캠페인수(발송빈도): {t_camp['first']:,.0f}건 → {t_camp['last']:,.0f}건 "
                               f"({_fmt_pct(t_camp)}, {t_camp['pattern']}) — 피로도와 직결되는 보조 지표예요.")
                if t_uv:
                    good = t_uv["pattern"] == "상승"
                    icon = "✅" if good else ("⚠️" if t_uv["pattern"] == "등락 반복" else "❌")
                    comment = ("목표대로 오르는 흐름이에요." if good else
                              "등락이 있어 아직 뚜렷한 개선은 아니에요." if t_uv["pattern"] == "등락 반복" else
                              "줄고 있어 주의가 필요해요.")
                    st.markdown(f"{icon} **UV**: {t_uv['first']:,.0f} → {t_uv['last']:,.0f} "
                               f"({_fmt_pct(t_uv)}, {t_uv['pattern']}) — {comment}")
                if t_ctr:
                    good = t_ctr["pattern"] == "상승"
                    icon = "✅" if good else ("⚠️" if t_ctr["pattern"] == "등락 반복" else "❌")
                    ctr_first_pp, ctr_last_pp = t_ctr["first"] * 100, t_ctr["last"] * 100
                    comment = ("메시지 반응률이 개선되는 흐름이에요." if good else
                              "눈에 띄는 개선은 아직이에요." if t_ctr["pattern"] == "등락 반복" else
                              "반응률이 떨어지고 있어 피로도 신호일 수 있어요.")
                    st.markdown(f"{icon} **CTR**: {ctr_first_pp:.2f}% → {ctr_last_pp:.2f}% "
                               f"({ctr_last_pp - ctr_first_pp:+.2f}%p, {t_ctr['pattern']}) — {comment}")

                send_good = bool(t_send and t_send["pattern"] == "하락")
                uv_good = bool(t_uv and t_uv["pattern"] == "상승")
                ctr_good = bool(t_ctr and t_ctr["pattern"] == "상승")
                flags = [("발송량 감축", send_good), ("UV 상승", uv_good), ("CTR 상승", ctr_good)]
                good_list = [name for name, ok in flags if ok]
                bad_list = [name for name, ok in flags if not ok]
                n_good = len(good_list)

                if n_good == 3:
                    verdict = ("🟢 **목표에 부합하는 이상적인 흐름이에요** — 발송은 줄이면서 UV·CTR은 함께 오르고 있어, "
                              "더 적은 발송으로 더 좋은 반응을 끌어내고 있어요.")
                elif n_good == 0:
                    verdict = (f"🔴 **목표와 반대 방향으로 가고 있어요** — {', '.join(bad_list)} 모두 아쉬운 흐름이에요. "
                              "발송량과 타겟팅·오퍼를 함께 재점검해볼 시점이에요.")
                else:
                    verdict = (f"🟡 **부분적으로 목표에 부합해요** — {', '.join(good_list)}은(는) 잘 되고 있지만, "
                              f"{', '.join(bad_list)}은(는) 더 필요해요.")
                st.markdown(verdict)
                st.caption("판정 기준: 각 지표가 조회 구간 내내 순감소/순증가면 하락/상승, 오르내림이 섞이면 '등락 반복'으로 봐요 "
                          "(등락 반복은 목표 미달로 간주). 구간이 짧으면(2~3주) 노이즈일 수 있으니 참고용으로만 봐주세요.")

    # ══════════════════════════════════════════════════════════
    # 종합요약
    # ══════════════════════════════════════════════════════════
    elif page == "종합요약":
        st.title("종합 요약")
        tot = agg_metrics(f, ["stype"]).drop(columns=["stype"]).sum(numeric_only=True) if not f.empty else None
        send_t = f["send"].sum(); uv_t = f["uv"].sum(); oc_t = f["oc"].sum(); amt_t = f["amt"].sum()
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("발송모수", f"{send_t:,.0f}")
        c2.metric("UV", f"{uv_t:,.0f}")
        c3.metric("주문건수", f"{oc_t:,.0f}")
        c4.metric("거래액", won(amt_t))
        c5, c6, c7 = st.columns(3)
        c5.metric("CTR(UV/발송)", f"{(uv_t/send_t if send_t else 0):.2%}")
        c6.metric("주문전환율(주문/UV)", f"{(oc_t/uv_t if uv_t else 0):.2%}")
        c7.metric("RPS(발송당거래액)", f"{(amt_t/send_t if send_t else 0):,.0f}")

        st.markdown("---")
        st.markdown("#### 추이 지표 — 한눈에 보기")
        TREND_METRIC_LABELS = {**METRIC_LABELS, "ctr": "CTR(UV/발송)", "cvr": "CR(주문/UV)"}
        granularity = st.radio("단위", ["주차별", "일자별"], horizontal=True, key="trend_grid_granularity")

        if granularity == "주차별":
            gdf = f.groupby(["week_start", "week_label"], dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index()
            gdf = gdf.sort_values("week_start").rename(columns={"week_label": "x"})
            x_mode = "lines+markers"
        else:
            gdf = f.groupby(f["dt"].dt.date, dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index(names="x")
            gdf = gdf.sort_values("x")
            x_mode = "lines"
        gdf["ctr"] = np.where(gdf["send"] > 0, gdf["uv"] / gdf["send"], np.nan)
        gdf["cvr"] = np.where(gdf["uv"] > 0, gdf["oc"] / gdf["uv"], np.nan)

        grid_metrics = list(TREND_METRIC_LABELS)
        fig = make_subplots(rows=2, cols=3, subplot_titles=[TREND_METRIC_LABELS[m] for m in grid_metrics])
        for i, m in enumerate(grid_metrics):
            r, c = i // 3 + 1, i % 3 + 1
            fig.add_trace(go.Scatter(x=gdf["x"], y=gdf[m], mode=x_mode,
                                     line=dict(color="#4f8fff", width=2), showlegend=False), row=r, col=c)
            if m in ("ctr", "cvr"):
                fig.update_yaxes(tickformat=".1%", row=r, col=c)
        fig.update_layout(height=520, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                          font=dict(color="#475569", size=11), margin=dict(l=10, r=10, t=40, b=10))
        fig.update_xaxes(gridcolor="rgba(0,0,0,0)", linecolor="#e2e8f0")
        fig.update_yaxes(gridcolor="#f1f5f9", linecolor="#e2e8f0")
        for ann in fig["layout"]["annotations"]:
            ann["font"] = dict(color="#94a3b8", size=13)
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("#### 전주 대비 (월~일 기준)")
        this_start = pd.Timestamp(f["week_start"].max())
        this_end = this_start + pd.Timedelta(days=6)
        prev_start = this_start - pd.Timedelta(days=7)
        prev_end = this_start - pd.Timedelta(days=1)
        cur = f[(f["dt"] >= this_start) & (f["dt"] <= this_end)]
        prev = f[(f["dt"] >= prev_start) & (f["dt"] <= prev_end)]
        this_label = f"{this_start.month}/{this_start.day}~{this_end.month}/{this_end.day}"
        prev_label = f"{prev_start.month}/{prev_start.day}~{prev_end.month}/{prev_end.day}"
        rows = []
        for k, label in METRIC_LABELS.items():
            cv, pv = cur[k].sum(), prev[k].sum()
            delta = (cv - pv) / pv * 100 if pv else np.nan
            rows.append({"지표": label, this_label: cv, prev_label: pv, "증감%": delta})
        n_count_rows = len(rows)
        cur_send, prev_send = cur["send"].sum(), prev["send"].sum()
        cur_uv, prev_uv = cur["uv"].sum(), prev["uv"].sum()
        cur_oc, prev_oc = cur["oc"].sum(), prev["oc"].sum()
        ctr_cur = cur_uv / cur_send if cur_send else np.nan
        ctr_prev = prev_uv / prev_send if prev_send else np.nan
        cr_cur = cur_oc / cur_uv if cur_uv else np.nan
        cr_prev = prev_oc / prev_uv if prev_uv else np.nan
        rows.append({"지표": "CTR(UV/발송)", this_label: ctr_cur, prev_label: ctr_prev,
                    "증감%": ((ctr_cur - ctr_prev) / ctr_prev * 100 if prev_send and ctr_prev else np.nan)})
        rows.append({"지표": "CR(주문/UV)", this_label: cr_cur, prev_label: cr_prev,
                    "증감%": ((cr_cur - cr_prev) / cr_prev * 100 if prev_uv and cr_prev else np.nan)})
        show = pd.DataFrame(rows)
        sty = show.style.format({this_label: "{:,.0f}", prev_label: "{:,.0f}"},
                                subset=pd.IndexSlice[0:n_count_rows - 1, [this_label, prev_label]])
        sty = sty.format({this_label: "{:.2%}", prev_label: "{:.2%}"},
                         subset=pd.IndexSlice[n_count_rows:n_count_rows + 1, [this_label, prev_label]])
        sty = sty.format({"증감%": "{:+.1f}%"}, subset=pd.IndexSlice[:, ["증감%"]])
        st.dataframe(sty, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════
    # UV 유입 분석 — 발송모수 대비 UV 효율을 뜯어서 유효타겟 규모를 가늠하고
    # UV 개선을 위한 긴급 대안 후보를 자동으로 찾아주는 진단 페이지
    # ══════════════════════════════════════════════════════════
    elif page == "UV 유입 분석":
        st.title("UV 유입 분석")
        st.caption("발송모수 대비 UV(신규 유입) 효율을 뜯어봐서, 실제 유효 타겟 규모를 가늠하고 "
                  "UV를 끌어올릴 긴급 대안을 찾기 위한 페이지예요.")

        if f.empty:
            st.info("표시할 데이터가 없어요.")
        else:
            send_t, uv_t = f["send"].sum(), f["uv"].sum()
            ctr_t = uv_t / send_t if send_t else 0.0

            this_start = pd.Timestamp(f["week_start"].max())
            this_end = this_start + pd.Timedelta(days=6)
            prev_start = this_start - pd.Timedelta(days=7)
            prev_end = this_start - pd.Timedelta(days=1)
            cur_w = f[(f["dt"] >= this_start) & (f["dt"] <= this_end)]
            prev_w = f[(f["dt"] >= prev_start) & (f["dt"] <= prev_end)]
            cur_send, cur_uv = cur_w["send"].sum(), cur_w["uv"].sum()
            prev_send, prev_uv = prev_w["send"].sum(), prev_w["uv"].sum()
            cur_ctr = cur_uv / cur_send if cur_send else np.nan
            prev_ctr = prev_uv / prev_send if prev_send else np.nan
            ctr_delta_pp = (cur_ctr - prev_ctr) * 100 if pd.notna(cur_ctr) and pd.notna(prev_ctr) else np.nan

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("누적 발송모수", f"{send_t:,.0f}")
            c2.metric("누적 UV", f"{uv_t:,.0f}")
            c3.metric("전체 CTR(UV/발송)", f"{ctr_t:.2%}")
            c4.metric("이번 주 CTR", f"{(cur_ctr if pd.notna(cur_ctr) else 0):.2%}",
                      delta=(f"{ctr_delta_pp:+.2f}%p 전주 대비" if pd.notna(ctr_delta_pp) else None))

            st.markdown("---")
            st.markdown("#### 주차별 발송량 × CTR 추이")
            st.caption("발송량은 늘어나는데 CTR이 같이 떨어지면 '같은 타겟에 반복 발송 → 피로도·타겟 소진' 신호예요.")
            wk = f.groupby(["week_start", "week_label"], dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum")).reset_index().sort_values("week_start")
            wk["ctr"] = np.where(wk["send"] > 0, wk["uv"] / wk["send"], np.nan)

            if len(wk) < 2:
                st.info("추이를 보려면 최소 2주 이상의 데이터가 필요해요.")
            else:
                fig = go.Figure()
                fig.add_bar(x=wk["week_label"], y=wk["send"], name="발송모수", marker_color="#cbd5e1", yaxis="y")
                fig.add_trace(go.Scatter(x=wk["week_label"], y=wk["ctr"], name="CTR", yaxis="y2",
                                         mode="lines+markers", line=dict(color="#e53e3e", width=2)))
                layout = base_layout(title="주차별 발송모수 · CTR 추이")
                layout["yaxis2"] = dict(overlaying="y", side="right", showgrid=False, tickformat=".1%")
                fig.update_layout(**layout)
                st.plotly_chart(fig, use_container_width=True)

                vals = wk["ctr"].dropna().tolist()
                if len(vals) >= 2:
                    first, last = vals[0], vals[-1]
                    pct = (last - first) / first * 100 if first else None
                    send_first, send_last = wk["send"].iloc[0], wk["send"].iloc[-1]
                    send_pct = (send_last - send_first) / send_first * 100 if send_first else None
                    wk_first_label = str(wk.iloc[0]["week_label"]).replace("~", "\\~")
                    wk_last_label = str(wk.iloc[-1]["week_label"]).replace("~", "\\~")
                    if pct is not None and pct <= -10 and (send_pct is None or send_pct >= -5):
                        st.warning(f"🔴 CTR이 {wk_first_label} {first:.2%} → {wk_last_label} "
                                  f"{last:.2%}({pct:+.1f}%)로 떨어지는 동안 발송량은 "
                                  f"{'유지·증가' if (send_pct is None or send_pct >= 0) else f'{send_pct:+.1f}%'}했어요 "
                                  "— 같은 타겟에 반복 발송하며 반응이 둔화되는 '타겟 소진' 신호일 수 있어요.")
                    elif pct is not None and pct >= 10:
                        st.success(f"🟢 CTR이 {first:.2%} → {last:.2%}({pct:+.1f}%)로 개선되는 흐름이에요.")
                    else:
                        st.info(f"CTR이 {first:.2%} → {last:.2%}로 큰 변화 없이 유지되고 있어요.")

            st.markdown("---")
            st.markdown("#### 일자별 총 UV 추이")
            st.caption("일 단위로 더 세밀하게 봐요. 주차별보다 변동은 크지만, 특정 요일·이벤트의 영향을 잡아내기 좋아요.")
            dd = f.groupby(f["dt"].dt.date, dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum")).reset_index(names="date").sort_values("date")
            if len(dd) < 2:
                st.info("추이를 보려면 최소 2일 이상의 데이터가 필요해요.")
            else:
                fig = go.Figure(go.Scatter(x=dd["date"], y=dd["uv"], mode="lines+markers", name="UV",
                                           line=dict(color="#4f8fff", width=2), fill="tozeroy",
                                           fillcolor="rgba(79,143,255,0.08)",
                                           hovertemplate="%{x}<br>UV: %{y:,.0f}<extra></extra>"))
                fig.update_layout(**base_layout(title="일자별 총 UV 추이"))
                st.plotly_chart(fig, use_container_width=True)
                st.caption(f"조회 기간 {len(dd)}일 · 일평균 UV {dd['uv'].mean():,.0f} · "
                          f"최고 {dd.loc[dd['uv'].idxmax(), 'date']}({dd['uv'].max():,.0f}) · "
                          f"최저 {dd.loc[dd['uv'].idxmin(), 'date']}({dd['uv'].min():,.0f})")

                st.markdown("##### 전주·전월 동일자 대비")
                st.caption("각 날짜를 정확히 7일 전(전주 동일 요일)·1개월 전 같은 날짜(전월 동일자)와 비교해요. "
                          "비교 대상 날짜가 현재 선택한 기간보다 앞이면 '-'로 표시돼요 — 필요하면 기간을 넓혀보세요.")
                comp = dd.copy()
                comp["date_ts"] = pd.to_datetime(comp["date"])
                uv_by_date = comp.set_index("date_ts")["uv"]

                def _lookup(ts, offset):
                    return uv_by_date.get(ts - offset, np.nan)

                comp["전주 동일요일"] = comp["date_ts"].apply(lambda d: _lookup(d, pd.Timedelta(days=7)))
                comp["전월 동일자"] = comp["date_ts"].apply(lambda d: _lookup(d, pd.DateOffset(months=1)))
                comp["전주비"] = np.where(comp["전주 동일요일"] > 0,
                                        (comp["uv"] - comp["전주 동일요일"]) / comp["전주 동일요일"] * 100, np.nan)
                comp["전월비"] = np.where(comp["전월 동일자"] > 0,
                                        (comp["uv"] - comp["전월 동일자"]) / comp["전월 동일자"] * 100, np.nan)

                latest = comp.sort_values("date_ts").iloc[-1]
                lc1, lc2, lc3 = st.columns(3)
                lc1.metric(f"{latest['date']} UV", f"{latest['uv']:,.0f}")
                lc2.metric("전주 동일요일 UV",
                          (f"{latest['전주 동일요일']:,.0f}" if pd.notna(latest["전주 동일요일"]) else "-"),
                          delta=(f"{latest['전주비']:+.1f}%" if pd.notna(latest["전주비"]) else None))
                lc3.metric("전월 동일자 UV",
                          (f"{latest['전월 동일자']:,.0f}" if pd.notna(latest["전월 동일자"]) else "-"),
                          delta=(f"{latest['전월비']:+.1f}%" if pd.notna(latest["전월비"]) else None))

                show_comp = comp.sort_values("date_ts", ascending=False)[
                    ["date", "uv", "전주 동일요일", "전주비", "전월 동일자", "전월비"]
                ].rename(columns={"date": "일자", "uv": "UV"})
                st.dataframe(show_comp.style.format({
                    "UV": "{:,.0f}", "전주 동일요일": "{:,.0f}", "전월 동일자": "{:,.0f}",
                    "전주비": "{:+.1f}%", "전월비": "{:+.1f}%",
                }, na_rep="-"), use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("#### 발송모수 구간별 평균 CTR — 유효타겟 규모 가늠")
            st.caption("발송을 많이 한 건(대량 살포)일수록 CTR이 뚜렷하게 낮다면, 그만큼 실제로 반응하는 "
                      "유효 타겟 풀이 작다는 뜻이에요 — 많이 보내도 '덜 원하는 사람'까지 긁어야 하니까요.")
            vf = f[f["send"] > 0].copy()
            if len(vf) < 4:
                st.info("구간을 나누기엔 데이터가 부족해요.")
            else:
                try:
                    cats, bins = pd.qcut(vf["send"], q=4, duplicates="drop", retbins=True)
                except ValueError:
                    cats, bins = None, None
                if cats is None or len(bins) < 3:
                    st.info("발송모수 값이 너무 비슷해서 구간을 나눌 수 없어요.")
                else:
                    labels = [f"{int(bins[i]):,}~{int(bins[i + 1]):,}" for i in range(len(bins) - 1)]
                    vf["발송구간"] = cats.cat.rename_categories(labels)
                    qg = vf.groupby("발송구간", observed=True).agg(
                        send=("send", "sum"), uv=("uv", "sum"), n=("af", "count")).reset_index()
                    qg["ctr"] = np.where(qg["send"] > 0, qg["uv"] / qg["send"], 0.0)
                    fig = go.Figure(go.Bar(x=qg["발송구간"].astype(str), y=qg["ctr"], marker_color="#4f8fff",
                                           text=[f"{v:.2%} ({n}건)" for v, n in zip(qg["ctr"], qg["n"])],
                                           textposition="outside"))
                    layout = base_layout(title="발송모수 구간별 CTR")
                    layout["yaxis"]["tickformat"] = ".1%"
                    fig.update_layout(**layout)
                    st.plotly_chart(fig, use_container_width=True)

            st.markdown("---")
            st.markdown("#### 세그먼트별 CTR 랭킹")
            SEG_DIMS = {"bpu_group": "BPU", "brand": "브랜드", "cat": "카테고리",
                       "stype": "발송유형", "target": "타겟구분"}
            seg_dim = st.selectbox("기준", list(SEG_DIMS), format_func=lambda k: SEG_DIMS[k], key="uv_seg_dim")
            seg = f[f[seg_dim].astype(str).str.strip() != ""].groupby(seg_dim, dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum")).reset_index()
            if seg.empty:
                st.info(f"{SEG_DIMS[seg_dim]} 정보가 있는 데이터가 없어요.")
            else:
                seg["ctr"] = np.where(seg["send"] > 0, seg["uv"] / seg["send"], 0.0)
                seg["발송비중"] = seg["send"] / seg["send"].sum()
                seg = seg.sort_values("ctr", ascending=False)

                fig = go.Figure(go.Bar(x=seg["ctr"], y=seg[seg_dim].astype(str), orientation="h",
                                       marker_color="#4f8fff"))
                fig.update_layout(**base_layout(h=max(260, 30 * len(seg)), title=f"{SEG_DIMS[seg_dim]}별 CTR"))
                fig.update_xaxes(tickformat=".1%")
                fig.update_yaxes(autorange="reversed")
                st.plotly_chart(fig, use_container_width=True)

                show_seg = seg.rename(columns={seg_dim: SEG_DIMS[seg_dim], "send": "발송모수", "uv": "UV",
                                               "ctr": "CTR"})
                st.dataframe(show_seg[[SEG_DIMS[seg_dim], "발송모수", "UV", "CTR", "발송비중"]].style.format({
                    "발송모수": "{:,.0f}", "UV": "{:,.0f}", "CTR": "{:.2%}", "발송비중": "{:.1%}",
                }), use_container_width=True, hide_index=True)

                st.markdown("##### 🚨 긴급 대안 후보")
                MIN_SEND_SHARE = 0.02  # 표본이 너무 작은 세그먼트는 노이즈가 커서 제외
                cand = seg[seg["발송비중"] >= MIN_SEND_SHARE]
                avg_ctr = ctr_t
                median_share = cand["발송비중"].median() if not cand.empty else 0
                if avg_ctr > 0:
                    expand = cand[(cand["ctr"] > avg_ctr * 1.2) & (cand["발송비중"] < median_share)].sort_values(
                        "ctr", ascending=False).head(3)
                    review = cand[(cand["ctr"] < avg_ctr * 0.7) & (cand["발송비중"] >= median_share)].sort_values(
                        "ctr").head(3)
                else:
                    expand = review = cand.iloc[0:0]

                ec1, ec2 = st.columns(2)
                with ec1:
                    st.markdown("**📈 확대 후보** (효율 좋은데 비중은 낮음)")
                    if expand.empty:
                        st.caption("조건에 맞는 후보가 없어요.")
                    else:
                        for _, r in expand.iterrows():
                            st.markdown(f"- **{r[seg_dim]}**: CTR {r['ctr']:.2%} "
                                       f"(전체 평균의 {r['ctr'] / avg_ctr:.1f}배) · 발송비중 {r['발송비중']:.1%}")
                with ec2:
                    st.markdown("**⚠️ 재검토 후보** (비중은 큰데 효율은 낮음)")
                    if review.empty:
                        st.caption("조건에 맞는 후보가 없어요.")
                    else:
                        for _, r in review.iterrows():
                            st.markdown(f"- **{r[seg_dim]}**: CTR {r['ctr']:.2%} "
                                       f"(전체 평균의 {r['ctr'] / avg_ctr:.1f}배) · 발송비중 {r['발송비중']:.1%}")
                st.caption(f"기준: 전체 평균 CTR {avg_ctr:.2%}의 1.2배 초과 + 발송비중 중앙값({median_share:.1%}) 미만 "
                          "→ 확대 후보 / 0.7배 미만 + 발송비중 중앙값 이상 → 재검토 후보. 발송비중 "
                          f"{MIN_SEND_SHARE:.0%} 미만인 소표본 세그먼트는 제외했어요. "
                          "이 페이지는 캠페인 단위 집계라 실제 중복 없는 순수 회원수(고유 UV)까지는 알 수 없다는 "
                          "점은 참고해 주세요 — 같은 사람이 여러 캠페인에 잡힐 수 있어요.")

            st.markdown("---")
            st.markdown("#### 🤖 AI와 지표 논의하기")
            st.caption("위에서 본 지표들을 근거로 AI에게 물어보세요. 예: '지금 가장 시급한 문제가 뭐야?', "
                      "'신규 타겟 비중을 얼마나 더 늘려야 할까?', '휴면 타겟 발송을 줄이면 거래액에 영향이 클까?'")

            def _build_uv_ai_facts():
                lines = [f"전체 기간: 누적 발송 {send_t:,.0f} · 누적 UV {uv_t:,.0f} · 전체 CTR {ctr_t:.2%}"]
                if pd.notna(cur_ctr):
                    lines.append(f"이번 주 CTR {cur_ctr:.2%}"
                                 + (f" (전주 대비 {ctr_delta_pp:+.2f}%p)" if pd.notna(ctr_delta_pp) else ""))
                fw = f.groupby(["week_start", "week_label"], dropna=False).agg(
                    send=("send", "sum"), uv=("uv", "sum")).reset_index().sort_values("week_start")
                fw["ctr"] = np.where(fw["send"] > 0, fw["uv"] / fw["send"], np.nan)
                lines.append("\n주차별 발송·CTR (월~일):")
                for _, r in fw.iterrows():
                    if pd.notna(r["ctr"]):
                        lines.append(f"- {r['week_label']}: 발송 {r['send']:,.0f} · UV {r['uv']:,.0f} · "
                                     f"CTR {r['ctr']:.2%}")
                for dim_key, dim_label in SEG_DIMS.items():
                    sg = f[f[dim_key].astype(str).str.strip() != ""].groupby(dim_key, dropna=False).agg(
                        send=("send", "sum"), uv=("uv", "sum")).reset_index()
                    if sg.empty:
                        continue
                    sg["ctr"] = np.where(sg["send"] > 0, sg["uv"] / sg["send"], 0.0)
                    sg["발송비중"] = sg["send"] / sg["send"].sum()
                    sg = sg.sort_values("ctr", ascending=False)
                    lines.append(f"\n{dim_label}별 CTR (CTR 내림차순):")
                    for _, r in sg.iterrows():
                        lines.append(f"- {r[dim_key]}: CTR {r['ctr']:.2%} · 발송비중 {r['발송비중']:.1%} · "
                                     f"발송 {r['send']:,.0f}")
                return "\n".join(lines)

            ai_hist_key = "uv_ai_chat_history"
            if ai_hist_key not in st.session_state:
                st.session_state[ai_hist_key] = []

            ac1, ac2 = st.columns([3, 1])
            ai_model_name = ac1.selectbox("AI 모델", list(AI_MODELS), key="uv_ai_model",
                                          index=list(AI_MODELS).index("Gemini 2.5 Flash (균형)"))
            nominal_model = AI_MODELS[ai_model_name]
            # 이전 대화에서 이 모델이 막혀서 다른 모델로 자동 전환된 적 있으면, 이번에도 그 모델부터 씀
            # (막힌 걸 알면서 매번 한 번 더 실패시키지 않기 위함)
            ai_model_map = st.session_state.setdefault("uv_ai_working_model", {})
            ai_model = ai_model_map.get(nominal_model, nominal_model)
            if ac2.button("🗑️ 대화 초기화", use_container_width=True, key="uv_ai_reset"):
                st.session_state[ai_hist_key] = []
                st.rerun()

            for role, text in st.session_state[ai_hist_key]:
                with st.chat_message("user" if role == "user" else "assistant"):
                    st.markdown(text)

            user_q = st.chat_input("지표에 대해 질문해보세요", key="uv_ai_input")
            if user_q:
                st.session_state[ai_hist_key].append(("user", user_q))
                with st.chat_message("user"):
                    st.markdown(user_q)
                system = ("당신은 LF몰 CRM PUSH 발송 데이터 분석가입니다. 아래 [데이터]에 있는 수치만 근거로 "
                          "사용자와 한국어로 대화하세요. 데이터에 없는 값은 지어내지 말고 모른다고 답하세요. "
                          "실무자가 바로 실행할 수 있도록 구체적이고 간결하게 답변하세요.\n\n"
                          f"[데이터]\n{_build_uv_ai_facts()}")
                with st.chat_message("assistant"):
                    with st.spinner("생각 중…"):
                        txt, err, used_model = ai_chat_generate(system, st.session_state[ai_hist_key], ai_model)
                    if err:
                        st.warning(err)
                        st.session_state[ai_hist_key].pop()
                    else:
                        if used_model != nominal_model:
                            ai_model_map[nominal_model] = used_model
                            st.caption(f"ℹ️ '{ai_model_name}' 모델은 더 이상 지원되지 않아 "
                                      f"`{used_model}`로 자동 전환해서 답했어요.")
                        st.markdown(txt)
                        st.session_state[ai_hist_key].append(("assistant", txt))

    # ══════════════════════════════════════════════════════════
    # BPU별 실적
    # ══════════════════════════════════════════════════════════
    elif page == "BPU별 실적":
        st.title("BPU별 실적")
        st.caption("소재별 실적 원본 데이터를 BPU 기준으로 실시간 집계한 화면이에요 "
                  "(1BPU_B/1BPU_C처럼 세부 코드가 붙은 BPU는 1BPU로 묶어서 보여줘요).")
        g = agg_metrics(f, ["bpu_group"])
        g = g[g["bpu_group"] != ""].sort_values("amt", ascending=False)
        metric = st.selectbox("정렬/차트 지표", list(METRIC_LABELS), format_func=lambda k: METRIC_LABELS[k], key="bpu_metric")
        gs = g.sort_values(metric, ascending=False)
        fig = go.Figure(go.Bar(x=gs[metric], y=gs["bpu_group"], orientation="h", marker_color="#4f8fff"))
        fig.update_layout(**base_layout(h=max(300, 28 * len(gs)), title=f"BPU별 {METRIC_LABELS[metric]}"))
        fig.update_yaxes(autorange="reversed")
        st.plotly_chart(fig, use_container_width=True)

        show = gs.rename(columns={**METRIC_LABELS, **RATE_LABELS, "bpu_group": "BPU", "n": "캠페인수"})
        st.dataframe(show.style.format({
            "발송모수": "{:,.0f}", "UV": "{:,.0f}", "주문건수": "{:,.0f}", "거래액": "{:,.0f}",
            "CTR(UV/발송)": "{:.2%}", "주문전환율(주문/UV)": "{:.2%}", "RPS(발송당거래액)": "{:,.0f}", "객단가(거래액/주문)": "{:,.0f}",
        }), use_container_width=True, hide_index=True)

        st.markdown("#### 주차별 BPU 랭킹")
        st.caption("선택한 주차의 1~4BPU 발송건수·거래액·발송건당 거래액을 비교해 순위를 매겨요 "
                  "(동률은 같은 순위, 마케팅/편성/브랜드컨텐츠 등은 제외).",
                  help=(
                      "**집계 기준**\n\n"
                      "- **대상 주차**: 위 '주차 선택'에서 고른 한 주(월~일)의 데이터만 집계해요.\n"
                      "- **BPU**: 세부 코드(1BPU_B/1BPU_C 등)는 대표 BPU(1BPU/2BPU/3BPU/4BPU)로 합쳐서 보여줘요. "
                      "마케팅·편성·브랜드컨텐츠처럼 번호가 없는 BPU는 랭킹에서 제외해요.\n"
                      "- **거래액**: 선택한 주차 동안 해당 BPU로 발송된 캠페인들의 거래액 합계예요.\n"
                      "- **전주 PUSH발송건수**: 선택한 주차의 발송모수(발송건수) 합계예요. "
                      "(컬럼명은 원본 리포트 표기를 따른 것으로, 실제로는 '전주'가 아니라 선택한 주차 자체의 발송건수예요.)\n"
                      "- **발송건당 거래액**: 거래액 ÷ 전주 PUSH발송건수예요.\n"
                      "- **RANK**: 발송건당 거래액이 높은 순서로 매긴 순위예요. 값이 같으면 동일 순위를 공유해요(예: 공동 1위가 있으면 다음 순위는 2위가 아닌 3위).\n"
                      "- **계** 행: 1~4BPU 전체 합계이며 순위 매기기 대상에서는 제외돼요(RANK는 '-')."
                  ))
        weeks_sorted = f[["week_start", "week_label"]].drop_duplicates().sort_values("week_start", ascending=False)
        week_options = weeks_sorted["week_label"].tolist()
        if not week_options:
            st.info("표시할 주차 데이터가 없어요.")
        else:
            week_pick = st.selectbox("주차 선택", week_options, key="bpu_rank_week")
            wf = f[f["week_label"] == week_pick]
            rank_g = wf.groupby("bpu_group", dropna=False).agg(send=("send", "sum"), amt=("amt", "sum")).reset_index()
            rank_g = rank_g[rank_g["bpu_group"].str.match(r"^\d+BPU$", na=False)]
            rank_g["rps"] = np.where(rank_g["send"] > 0, rank_g["amt"] / rank_g["send"], 0.0)
            rank_g["RANK"] = rank_g["rps"].rank(method="min", ascending=False).astype(int)
            rank_g = rank_g.sort_values("RANK")
            total_send, total_amt = rank_g["send"].sum(), rank_g["amt"].sum()
            total_row = pd.DataFrame([{
                "bpu_group": "계", "send": total_send, "amt": total_amt,
                "rps": (total_amt / total_send if total_send > 0 else 0.0), "RANK": "-",
            }])
            rank_table = pd.concat([total_row, rank_g], ignore_index=True)
            show_rank = rank_table.rename(columns={"bpu_group": "BPU", "amt": "거래액",
                                                    "send": "전주 PUSH발송건수", "rps": "발송건당 거래액"})
            st.dataframe(show_rank[["BPU", "거래액", "전주 PUSH발송건수", "발송건당 거래액", "RANK"]].style.format({
                "거래액": "{:,.0f}", "전주 PUSH발송건수": "{:,.0f}", "발송건당 거래액": "{:,.0f}",
            }), use_container_width=True, hide_index=True)

        st.markdown("#### BPU 드릴다운")
        pick = st.selectbox("BPU 선택", gs["bpu_group"].tolist())
        picked_df = f[f["bpu_group"] == pick]
        drill = agg_metrics(picked_df, ["brand", "cat"]).sort_values("amt", ascending=False).head(15)
        st.dataframe(drill.rename(columns={**METRIC_LABELS, "brand": "브랜드", "cat": "카테고리", "n": "캠페인수"}),
                     use_container_width=True, hide_index=True)

        st.markdown("#### 요일 × 시간대 효율 히트맵")
        st.caption(f"'{pick}' BPU를 요일·시간대로 쪼개서 비교해요. 진할수록 값이 높아요 — 어느 요일·시간대에 "
                  "보내는 게 효율이 좋은지 한눈에 볼 수 있어요.")
        DOW_LABELS = ["월", "화", "수", "목", "금", "토", "일"]
        heat_metric = st.selectbox("지표", list(RATE_LABELS), format_func=lambda k: RATE_LABELS[k],
                                   index=list(RATE_LABELS).index("rps"), key="bpu_heat_metric")
        hf = picked_df.dropna(subset=["dow", "hour"])
        if hf.empty:
            st.info("표시할 데이터가 없어요.")
        else:
            hg = hf.groupby(["dow", "hour"], dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index()
            hg["ctr"] = np.where(hg["send"] > 0, hg["uv"] / hg["send"], 0.0)
            hg["cvr"] = np.where(hg["uv"] > 0, hg["oc"] / hg["uv"], 0.0)
            hg["rps"] = np.where(hg["send"] > 0, hg["amt"] / hg["send"], 0.0)
            hg["aov"] = np.where(hg["oc"] > 0, hg["amt"] / hg["oc"], 0.0)
            hg["dow_k"] = hg["dow"].astype(int).map(lambda d: DOW_LABELS[d])
            is_pct = heat_metric in ("ctr", "cvr")

            pivot = hg.pivot_table(index="dow_k", columns="hour", values=heat_metric, fill_value=0)
            pivot = pivot.reindex(DOW_LABELS)
            x_labels = [_hhmm(h) for h in pivot.columns]
            text = [[(f"{v:.1%}" if is_pct else f"{v:,.0f}") for v in row] for row in pivot.values]
            fig = go.Figure(go.Heatmap(z=pivot.values, x=x_labels, y=pivot.index.tolist(), colorscale="Blues",
                                       text=text, texttemplate="%{text}", textfont=dict(size=10),
                                       hovertemplate="%{y} %{x}<br>" + RATE_LABELS[heat_metric] + ": %{text}<extra></extra>"))
            fig.update_layout(height=340, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                              font=dict(color="#475569", size=11), margin=dict(l=10, r=10, t=20, b=10))
            st.plotly_chart(fig, use_container_width=True)

            st.markdown("##### 효율 좋은 요일 × 시간대 TOP 10")
            top = hg.sort_values(heat_metric, ascending=False).head(10).copy()
            top["시간대"] = top["hour"].map(_hhmm)
            show_top = top.rename(columns={"dow_k": "요일", "send": "발송모수", "uv": "UV",
                                           "oc": "주문건수", "amt": "거래액", **RATE_LABELS})
            metric_label = RATE_LABELS[heat_metric]
            cols = ["요일", "시간대", "발송모수", "UV", "주문건수", "거래액", RATE_LABELS["ctr"], RATE_LABELS["cvr"]]
            if metric_label not in cols:
                cols.append(metric_label)
            fmt = {"발송모수": "{:,.0f}", "UV": "{:,.0f}", "주문건수": "{:,.0f}", "거래액": "{:,.0f}",
                  RATE_LABELS["ctr"]: "{:.2%}", RATE_LABELS["cvr"]: "{:.2%}"}
            fmt[metric_label] = "{:.2%}" if is_pct else "{:,.0f}"
            st.dataframe(show_top[cols].style.format(fmt), use_container_width=True, hide_index=True)

        st.markdown("#### 🗓️ 컨틴구좌 배정 추천 (월~금 16시)")
        st.caption("정책 변경으로 신설된 월~금 16시 컨틴전시 구좌에 어떤 BPU·브랜드를 배정하면 좋을지, "
                  "'효율(그 시간대 실적)'과 '형평성(마지막 배정 이후 경과 주차)'을 함께 봐서 추천해요. "
                  "실제 발송 기록에서 16시(hour=1600) + 월~금 발송분을 컨틴 배정 이력의 대리 지표로 사용해요.")

        contin_hour = 1600
        contin_dow = [0, 1, 2, 3, 4]
        cf_contin = f[(f["hour"] == contin_hour) & (f["dow"].isin(contin_dow))]

        st.markdown("##### 1) 요일 × BPU 매칭 추천")
        st.caption("각 BPU가 월~금 16시 중 어느 요일에 가장 효율이 좋은지 보여줘요. 진하게 표시된 셀이 "
                  "그 BPU의 추천 요일이에요.")
        match_metric = st.selectbox("지표", list(RATE_LABELS), format_func=lambda k: RATE_LABELS[k],
                                    index=list(RATE_LABELS).index("rps"), key="contin_match_metric")
        bpu_main = cf_contin[cf_contin["bpu_group"].str.match(r"^\d+BPU$", na=False)]
        if bpu_main.empty:
            st.info("월~금 16시 발송 이력이 없어서 매칭표를 만들 수 없어요.")
        else:
            mg = bpu_main.groupby(["bpu_group", "dow"], dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index()
            mg["ctr"] = np.where(mg["send"] > 0, mg["uv"] / mg["send"], 0.0)
            mg["cvr"] = np.where(mg["uv"] > 0, mg["oc"] / mg["uv"], 0.0)
            mg["rps"] = np.where(mg["send"] > 0, mg["amt"] / mg["send"], 0.0)
            mg["aov"] = np.where(mg["oc"] > 0, mg["amt"] / mg["oc"], 0.0)
            mg["dow_k"] = mg["dow"].astype(int).map(lambda d: DOW_LABELS[d])
            match_pivot = mg.pivot_table(index="bpu_group", columns="dow_k", values=match_metric, fill_value=0)
            match_pivot = match_pivot.reindex(columns=[DOW_LABELS[d] for d in contin_dow]).sort_index()
            is_pct_m = match_metric in ("ctr", "cvr")
            fmt_m = "{:.2%}" if is_pct_m else "{:,.0f}"
            st.dataframe(match_pivot.style.highlight_max(axis=1, color="#c7e6d0").format(fmt_m),
                        use_container_width=True)

            best_day = match_pivot.idxmax(axis=1)
            rec_lines = [f"- **{bpu}** → 추천 요일 **{best_day[bpu]}**" for bpu in match_pivot.index]
            st.markdown("\n".join(rec_lines))
            dup = best_day[best_day.duplicated(keep=False)]
            if not dup.empty:
                for d in dup.unique():
                    conflicted = dup[dup == d].index.tolist()
                    st.warning(f"⚠️ **{d}요일**을 {', '.join(conflicted)}가 동시에 원해요 — 차선책 요일과의 "
                              "격차가 더 작은 쪽이 양보하는 식으로 조정해보세요.")

        st.markdown("##### 2) 카테고리 → 브랜드 선정 (입점처럼 여러 브랜드가 경쟁하는 BPU용)")
        st.caption("① 그 요일·16시에 카테고리별 평균 효율을 먼저 보고, ② 고른 카테고리 안에서 브랜드별로 "
                  "'효율'과 '마지막 컨틴 배정 이후 경과 주차(형평성)'를 합산 점수로 랭킹해요.")
        bpu_options_all = sorted(f["bpu_group"].unique().tolist())
        cb1, cb2, cb3 = st.columns([1, 1, 1.4])
        sel_bpu = cb1.selectbox("BPU 선택", bpu_options_all, key="contin_sel_bpu")
        sel_dow_label = cb2.selectbox("요일 선택", [DOW_LABELS[d] for d in contin_dow], key="contin_sel_dow")
        eff_weight = cb3.slider("효율 vs 형평성 가중치 (효율 쪽)", 0.0, 1.0, 0.5, 0.1, key="contin_eff_weight")
        sel_dow = contin_dow[[DOW_LABELS[d] for d in contin_dow].index(sel_dow_label)]

        slot_f = f[(f["bpu_group"] == sel_bpu) & (f["hour"] == contin_hour) & (f["dow"] == sel_dow)]
        if slot_f.empty:
            st.info(f"'{sel_bpu}'의 {sel_dow_label}요일 16시 발송 이력이 없어서 카테고리 랭킹을 만들 수 없어요.")
        else:
            cat_g = slot_f.groupby("cat", dropna=False).agg(
                send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index()
            cat_g = cat_g[cat_g["cat"] != ""]
            if cat_g.empty:
                st.info("카테고리 정보가 있는 데이터가 없어요.")
            else:
                cat_g["rps"] = np.where(cat_g["send"] > 0, cat_g["amt"] / cat_g["send"], 0.0)
                cat_g["ctr"] = np.where(cat_g["send"] > 0, cat_g["uv"] / cat_g["send"], 0.0)
                cat_g = cat_g.sort_values("rps", ascending=False)
                show_cat = cat_g.rename(columns={"cat": "카테고리", "send": "발송모수", "uv": "UV",
                                                 "oc": "주문건수", "amt": "거래액", "rps": "RPS", "ctr": "CTR"})
                st.dataframe(show_cat[["카테고리", "발송모수", "UV", "주문건수", "거래액", "RPS", "CTR"]].style.format({
                    "발송모수": "{:,.0f}", "UV": "{:,.0f}", "주문건수": "{:,.0f}", "거래액": "{:,.0f}",
                    "RPS": "{:,.0f}", "CTR": "{:.2%}",
                }), use_container_width=True, hide_index=True)

                sel_cat = st.selectbox("카테고리 선택 (브랜드 랭킹 보기)", cat_g["cat"].tolist(), key="contin_sel_cat")
                cat_slot_f = slot_f[slot_f["cat"] == sel_cat]
                cat_send_sum = cat_slot_f["send"].sum()
                cat_avg_rps = cat_slot_f["amt"].sum() / cat_send_sum if cat_send_sum else 0.0

                brand_g = cat_slot_f.groupby("brand", dropna=False).agg(
                    send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index()
                brand_g = brand_g[brand_g["brand"] != ""]
                if brand_g.empty:
                    st.info("이 카테고리에 브랜드 정보가 있는 데이터가 없어요.")
                else:
                    brand_g["rps"] = np.where(brand_g["send"] > 0, brand_g["amt"] / brand_g["send"], 0.0)
                    brand_g["vs_cat_avg"] = np.where(cat_avg_rps > 0, brand_g["rps"] / cat_avg_rps, 0.0)

                    ref_week_start = f.loc[f["dt"].idxmax(), "week_start"]
                    hist_f = cf_contin[cf_contin["bpu_group"] == sel_bpu]
                    last_assigned = hist_f.groupby("brand")["week_start"].max()

                    def _elapsed_weeks(brand):
                        if brand not in last_assigned.index or pd.isna(last_assigned[brand]):
                            return 999
                        return max(0, (ref_week_start - last_assigned[brand]).days // 7)

                    brand_g["경과주차"] = brand_g["brand"].map(_elapsed_weeks)

                    eff_vals = brand_g["rps"]
                    eff_min, eff_max = eff_vals.min(), eff_vals.max()
                    brand_g["eff_norm"] = ((eff_vals - eff_min) / (eff_max - eff_min)) if eff_max > eff_min else 1.0
                    fair_vals = brand_g["경과주차"].clip(upper=52)
                    f_min, f_max = fair_vals.min(), fair_vals.max()
                    brand_g["fair_norm"] = ((fair_vals - f_min) / (f_max - f_min)) if f_max > f_min else 1.0

                    brand_g["최종점수"] = brand_g["eff_norm"] * eff_weight + brand_g["fair_norm"] * (1 - eff_weight)
                    brand_g = brand_g.sort_values("최종점수", ascending=False).reset_index(drop=True)

                    show_brand = brand_g.rename(columns={"brand": "브랜드", "rps": "RPS",
                                                         "vs_cat_avg": "카테고리 평균 대비"})
                    show_brand["추천"] = ["🏆" if i == 0 else "" for i in range(len(show_brand))]
                    cols_b = ["추천", "브랜드", "RPS", "카테고리 평균 대비", "경과주차", "최종점수"]
                    st.dataframe(show_brand[cols_b].style.format({
                        "RPS": "{:,.0f}", "카테고리 평균 대비": "{:.0%}", "최종점수": "{:.2f}",
                    }), use_container_width=True, hide_index=True)
                    st.caption("경과주차 999 = 이 BPU에서 월~금 16시 발송 이력이 아직 없는 브랜드(최우선 후보)예요. "
                              "카테고리 평균 대비 70% 미만이면 품질 가드레일 차원에서 후순위 검토를 권장해요.")

    # ══════════════════════════════════════════════════════════
    # 발송유형별 실적
    # ══════════════════════════════════════════════════════════
    elif page == "발송유형별 실적":
        st.title("발송유형별 실적")
        st.caption("소재별 실적 원본 데이터를 발송유형 기준으로 실시간 집계한 화면이에요.")
        has_basic = f["stype"].str.contains("기본", na=False).any()
        has_good = f["stype"].str.contains("우수", na=False).any()
        if has_basic and has_good:
            basic = f[f["stype"].str.contains("기본", na=False)]
            good = f[f["stype"].str.contains("우수", na=False)]
            st.markdown("#### 기본발송 vs 우수발송")
            cols = st.columns(4)
            for i, (k, label) in enumerate(METRIC_LABELS.items()):
                bv, gv = basic[k].sum(), good[k].sum()
                cols[i].metric(label, f"{bv:,.0f} / {gv:,.0f}", help="기본발송 / 우수발송")

        g = agg_metrics(f, ["stype"])
        g = g[g["stype"] != ""].sort_values("amt", ascending=False)
        metric = st.selectbox("정렬/차트 지표", list(METRIC_LABELS), format_func=lambda k: METRIC_LABELS[k], key="stype_metric")
        gs = g.sort_values(metric, ascending=False)
        fig = go.Figure(go.Bar(x=gs["stype"], y=gs[metric], marker_color="#48bb78"))
        fig.update_layout(**base_layout(title=f"발송유형별 {METRIC_LABELS[metric]}"))
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("#### 발송유형 × BPU")
        pivot = f.pivot_table(index="bpu_group", columns="stype", values=metric, aggfunc="sum", fill_value=0)
        pivot.index.name = "BPU"
        st.dataframe(pivot.style.format("{:,.0f}"), use_container_width=True)

    # ══════════════════════════════════════════════════════════
    # 캠페인별 실적 — BPU 안에서 발송유형×카테고리로 캠페인을 구분해서 본다
    # ══════════════════════════════════════════════════════════
    elif page == "캠페인별 실적":
        st.title("캠페인별 실적")
        st.caption("같은 BPU 안에서도 캠페인마다 발송유형·카테고리가 달라요. 발송유형/카테고리를 "
                  "여러 개 선택하면 하나로 합쳐서 봐요 — 예를 들어 '럭키먼데이' 캠페인이 카테고리 "
                  "'장바구니'·'럭키먼데이' 두 군데에 걸쳐 있다면 둘 다 선택해서 합쳐 보면 돼요.")

        bpu_options = sorted([b for b in f["bpu_group"].unique() if b])
        if not bpu_options:
            st.info("표시할 데이터가 없어요.")
        else:
            default_bpu = "마케팅" if "마케팅" in bpu_options else bpu_options[0]
            bpu_pick = st.selectbox("BPU", bpu_options, index=bpu_options.index(default_bpu))
            bf = f[f["bpu_group"] == bpu_pick]

            c1, c2 = st.columns(2)
            stype_pick = c1.multiselect("발송유형", sorted([s for s in bf["stype"].unique() if s]),
                                        key="camp_stype_pick")
            cat_pick = c2.multiselect("카테고리", sorted([c for c in bf["cat"].unique() if c]),
                                      key="camp_cat_pick")

            cf = bf.copy()
            if stype_pick:
                cf = cf[cf["stype"].isin(stype_pick)]
            if cat_pick:
                cf = cf[cf["cat"].isin(cat_pick)]

            if cf.empty:
                st.info("조건에 맞는 데이터가 없어요. 발송유형/카테고리를 선택해 보세요.")
            else:
                st.markdown("#### 발송유형 × 카테고리 요약")
                summary = agg_metrics(cf, ["stype", "cat"]).sort_values("amt", ascending=False)
                show_summary = summary.rename(columns={"stype": "발송유형", "cat": "카테고리", "n": "건수",
                                                        **METRIC_LABELS, **RATE_LABELS})
                st.dataframe(show_summary[["발송유형", "카테고리", "건수", "발송모수", "UV", "주문건수", "거래액",
                                          *RATE_LABELS.values()]].style.format({
                    "발송모수": "{:,.0f}", "UV": "{:,.0f}", "주문건수": "{:,.0f}", "거래액": "{:,.0f}",
                    RATE_LABELS["ctr"]: "{:.2%}", RATE_LABELS["cvr"]: "{:.2%}",
                    RATE_LABELS["rps"]: "{:,.0f}", RATE_LABELS["aov"]: "{:,.0f}",
                }), use_container_width=True, hide_index=True)

                st.markdown("#### 주차별 비교")
                wk = cf.groupby(["week_start", "week_label"], dropna=False).agg(
                    send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index()
                wk = wk.sort_values("week_start")
                wk["ctr"] = np.where(wk["send"] > 0, wk["uv"] / wk["send"], 0.0)
                wk["cvr"] = np.where(wk["uv"] > 0, wk["oc"] / wk["uv"], 0.0)
                metric = st.selectbox("차트 지표", list(METRIC_LABELS), format_func=lambda k: METRIC_LABELS[k],
                                      key="camp_week_metric")
                fig = go.Figure(go.Bar(x=wk["week_label"], y=wk[metric], marker_color="#7b5bc0"))
                fig.update_layout(**base_layout(title=f"주차별 {METRIC_LABELS[metric]}"))
                st.plotly_chart(fig, use_container_width=True)
                show_wk = wk.rename(columns={"week_label": "주차", "ctr": "CTR", "cvr": "CR", **METRIC_LABELS})
                st.dataframe(show_wk[["주차", "발송모수", "UV", "주문건수", "거래액", "CTR", "CR"]].style.format({
                    "발송모수": "{:,.0f}", "UV": "{:,.0f}", "주문건수": "{:,.0f}", "거래액": "{:,.0f}",
                    "CTR": "{:.2%}", "CR": "{:.2%}",
                }), use_container_width=True, hide_index=True)

                st.markdown("#### 일자별 상세")
                weeks_sorted = cf[["week_start", "week_label"]].drop_duplicates().sort_values("week_start")
                hours_sorted = sorted(cf["hour"].dropna().unique())
                d1, d2, d3 = st.columns(3)
                brand_pick = d1.multiselect("브랜드", sorted([b for b in cf["brand"].unique() if b]),
                                            key="camp_detail_brand_pick")
                week_pick = d2.multiselect("주차", weeks_sorted["week_label"].tolist(),
                                           key="camp_detail_week_pick")
                hour_pick = d3.multiselect("시간대", hours_sorted, format_func=_hhmm, key="camp_detail_hour_pick")
                detail_f = cf
                if brand_pick:
                    detail_f = detail_f[detail_f["brand"].isin(brand_pick)]
                if week_pick:
                    detail_f = detail_f[detail_f["week_label"].isin(week_pick)]
                if hour_pick:
                    detail_f = detail_f[detail_f["hour"].isin(hour_pick)]

                detail = detail_f.sort_values("dt", ascending=False)[
                    ["date", "hour", "week_label", "cat", "brand", "send", "uv", "oc", "amt", "ctr", "cvr", "rps"]].copy()
                detail["hour"] = detail["hour"].map(lambda h: _hhmm(h) if pd.notna(h) else "")
                show_detail = detail.rename(columns={"date": "일자", "hour": "시간대", "week_label": "주차",
                                                     "cat": "카테고리", "brand": "브랜드", "ctr": "CTR",
                                                     "cvr": "CR", "rps": "효율", **METRIC_LABELS})
                st.caption(f"{len(show_detail):,}행")
                st.dataframe(show_detail.style.format({
                    "발송모수": "{:,.0f}", "UV": "{:,.0f}", "주문건수": "{:,.0f}", "거래액": "{:,.0f}",
                    "CTR": "{:.2%}", "CR": "{:.2%}", "효율": "{:,.0f}",
                }), use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════
    # 일자별 실적 — 왼쪽 필터가 적용된 데이터를 행 단위(일자+시간대+AF코드)로 그대로 본다
    # ══════════════════════════════════════════════════════════
    elif page == "일자별 실적":
        st.title("일자별 실적")
        st.caption("왼쪽 필터(기간/BPU/발송유형/검색)가 적용된 데이터를 일자 단위 행으로 봐요.")

        cur_url, cur_gid, is_custom = request_copy_source()
        with st.expander("🔗 요청문구 구글시트 연결", expanded=is_custom):
            st.caption(("현재 **직접 연결한 시트**를 쓰고 있어요." if is_custom else
                       "현재 **기본 시트**(발송 계획과 같은 스프레드시트의 '요청문구' 탭)를 쓰고 있어요.") +
                      " 다른 구글시트를 연결하려면 URL을 붙여넣어 주세요 (URL에 `#gid=...`가 있으면 그 "
                      "탭을, 없으면 첫 번째 탭을 읽어요). 기획전No./제목/내용 컬럼이 있어야 인식돼요. "
                      "서비스 계정 이메일에 뷰어 권한 공유가 필요해요.")
            url_input = st.text_input("구글시트 URL", value=cur_url, key="request_copy_url_input")
            b1, b2 = st.columns(2)
            if b1.button("🔗 연결", key="request_copy_connect", use_container_width=True):
                st.session_state.request_copy_url = url_input.strip()
                st.session_state.request_copy_gid = _parse_sheet_url_gid(url_input)
                _load_request_copy_cached.clear()
                st.rerun()
            if b2.button("↺ 기본 시트로 되돌리기", key="request_copy_reset", use_container_width=True,
                        disabled=not is_custom):
                st.session_state.pop("request_copy_url", None)
                st.session_state.pop("request_copy_gid", None)
                _load_request_copy_cached.clear()
                st.rerun()

        if f.empty:
            st.info("표시할 데이터가 없어요.")
        else:
            c0, c1, c2, c3 = st.columns(4)
            c4, c5, c6 = st.columns(3)
            dmin_d, dmax_d = f["dt"].min().date(), f["dt"].max().date()
            date_pick = c0.date_input("일자", value=(dmin_d, dmax_d), min_value=dmin_d, max_value=dmax_d,
                                      key="daily_date_pick")
            bpu_pick = c1.multiselect("BPU", sorted([v for v in f["bpu_group"].unique() if v]), key="daily_bpu_pick")
            cat_pick = c2.multiselect("카테고리", sorted([v for v in f["cat"].unique() if v]), key="daily_cat_pick")
            brand_pick = c3.multiselect("브랜드", sorted([v for v in f["brand"].unique() if v]), key="daily_brand_pick")
            owner_pick = c4.multiselect("담당자", sorted([v for v in f["owner"].unique() if v]), key="daily_owner_pick")
            promo_pick = c5.multiselect("기획전", sorted([v for v in f["promo"].unique() if v]), key="daily_promo_pick")
            af_pick = c6.multiselect("AF코드", sorted([v for v in f["af"].unique() if v]), key="daily_af_pick")

            df_detail = f
            if isinstance(date_pick, tuple) and len(date_pick) == 2:
                df_detail = df_detail[(df_detail["dt"].dt.date >= date_pick[0]) & (df_detail["dt"].dt.date <= date_pick[1])]
            if bpu_pick:
                df_detail = df_detail[df_detail["bpu_group"].isin(bpu_pick)]
            if cat_pick:
                df_detail = df_detail[df_detail["cat"].isin(cat_pick)]
            if brand_pick:
                df_detail = df_detail[df_detail["brand"].isin(brand_pick)]
            if owner_pick:
                df_detail = df_detail[df_detail["owner"].isin(owner_pick)]
            if promo_pick:
                df_detail = df_detail[df_detail["promo"].isin(promo_pick)]
            if af_pick:
                df_detail = df_detail[df_detail["af"].isin(af_pick)]

            request_copy_map, rc_key_cols, rc_err = load_request_copy_map()
            if rc_err:
                st.caption(f"⚠️ 요청문구 시트를 불러오지 못했어요 ({rc_err}) — 제목/내용 없이 표시해요.")
            df_detail = df_detail.copy()
            if rc_key_cols:
                if rc_key_cols == ["promo"]:
                    st.caption("ℹ️ 요청문구 시트에 일자/시간대 컬럼이 없어서 기획전No.만으로 매칭했어요.")

                def _rc_key(row, cols=rc_key_cols):
                    vals = []
                    for c in cols:
                        if c == "hour":
                            vals.append(_norm_hour_val(row["hour"]))
                        elif c == "promo":
                            vals.append(_norm_promo_val(row["promo"]))
                        else:
                            vals.append(_txt_val(row[c]))
                    return tuple(vals)

                rc_keys = df_detail.apply(_rc_key, axis=1)
                df_detail["title"] = rc_keys.map(lambda k: request_copy_map.get(k, {}).get("title", ""))
                df_detail["content"] = rc_keys.map(lambda k: request_copy_map.get(k, {}).get("content", ""))
            else:
                df_detail["title"] = ""
                df_detail["content"] = ""

            detail_cols = ["date", "hour", "bpu_group", "cat", "brand", "owner", "promo", "title", "content",
                          "send", "uv", "visit", "cust", "oc", "amt", "cvr", "ctr", "rps"]
            detail = df_detail.sort_values("dt", ascending=False)[detail_cols]
            show = detail.rename(columns={"date": "일자", "hour": "시간대", "bpu_group": "BPU", "cat": "카테고리",
                                          "brand": "브랜드", "owner": "담당자", "promo": "기획전",
                                          "title": "제목", "content": "내용",
                                          "send": "발송모수", "uv": "UV", "visit": "VISIT", "cust": "고객수",
                                          "oc": "주문건수", "amt": "거래액", "cvr": "CR", "ctr": "CTR", "rps": "효율"})
            st.caption(f"{len(show):,}행")
            st.dataframe(show.style.format({
                "시간대": "{:.0f}", "발송모수": "{:,.0f}", "UV": "{:,.0f}", "VISIT": "{:,.0f}", "고객수": "{:,.0f}",
                "주문건수": "{:,.0f}", "거래액": "{:,.0f}", "CR": "{:.2%}", "CTR": "{:.2%}", "효율": "{:,.0f}",
            }), use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════
    # 주차별 누적 추이
    # ══════════════════════════════════════════════════════════
    elif page == "주차별 누적 추이":
        st.title("주차별 누적 추이")
        st.caption("월~일 기준 주차로 묶어 집계했어요 (누적_소재별 시트 기반).")
        g = f.groupby(["week_start", "week_label"], dropna=False).agg(
            send=("send", "sum"), uv=("uv", "sum"), oc=("oc", "sum"), amt=("amt", "sum")).reset_index()
        g = g.sort_values("week_start")
        g["ctr"] = np.where(g["send"] > 0, g["uv"] / g["send"], 0.0)
        g["cvr"] = np.where(g["uv"] > 0, g["oc"] / g["uv"], 0.0)
        TREND_METRIC_LABELS = {**METRIC_LABELS, "ctr": "CTR(UV/발송)", "cvr": "CR(주문/UV)"}
        metric = st.selectbox("지표", list(TREND_METRIC_LABELS), format_func=lambda k: TREND_METRIC_LABELS[k],
                              key="week_metric")
        is_pct = metric in ("ctr", "cvr")
        g["누적"] = g[metric].cumsum()
        g["WoW%"] = g[metric].pct_change() * 100

        fig = go.Figure()
        fig.add_bar(x=g["week_label"], y=g[metric], name=f"주간 {TREND_METRIC_LABELS[metric]}", marker_color="#4f8fff")
        fig.add_trace(go.Scatter(x=g["week_label"], y=g["누적"], name="누적", yaxis="y2",
                                 mode="lines+markers", line=dict(color="#ed8936", width=2)))
        layout = base_layout(title=f"주차별 {TREND_METRIC_LABELS[metric]} · 누적 추이")
        layout["yaxis2"] = dict(overlaying="y", side="right", showgrid=False)
        if is_pct:
            layout["yaxis"]["tickformat"] = ".1%"
        fig.update_layout(**layout)
        st.plotly_chart(fig, use_container_width=True)

        show = g.rename(columns={"week_label": "주차", "ctr": "CTR", "cvr": "CR", **METRIC_LABELS})
        st.dataframe(show[["주차", "발송모수", "UV", "주문건수", "거래액", "CTR", "CR", "누적", "WoW%"]]
                     .style.format({"발송모수": "{:,.0f}", "UV": "{:,.0f}", "주문건수": "{:,.0f}",
                                    "거래액": "{:,.0f}", "CTR": "{:.2%}", "CR": "{:.2%}",
                                    "누적": ("{:.2%}" if is_pct else "{:,.0f}"), "WoW%": "{:+.1f}%"}),
                     use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════
    # 우수발송 대시보드 — premium_push_dashboard.py의 render_premium_dashboard를
    # 그대로 호출한다. 컬럼명만 그쪽이 기대하는 한글 이름으로 맞춰서 넘긴다.
    # ══════════════════════════════════════════════════════════
    elif page == "우수발송 대시보드":
        from premium_push_dashboard import render_premium_dashboard, WEEKDAY_ORDER as PREMIUM_DOW
        raw_premium = f.rename(columns={
            "date": "날짜", "stype": "발송유형", "prio": "우선순위", "cat": "카테고리",
            "attr": "속성", "brand": "브랜드", "promo": "기획전", "send": "발송",
            "uv": "UV", "visit": "VISIT", "cust": "고객수", "oc": "주문건수", "amt": "주문금액",
        }).copy()
        raw_premium["요일"] = f["dow"].map(lambda d: PREMIUM_DOW[int(d)] if pd.notna(d) else "")
        raw_premium["시간대"] = f["hour"]
        raw_premium["BPU"] = f["bpu_group"]
        raw_premium["주차"] = f["week_label"]
        render_premium_dashboard(raw_premium)

    # ══════════════════════════════════════════════════════════
    # 직접 입력 — 엑셀 업로드 없이 표에서 바로 기록
    # ══════════════════════════════════════════════════════════
    elif page == "✍️ 직접 입력":
        st.title("직접 입력")
        tab_raw, tab_final = st.tabs(["📋 발송 건수 + RawNew (조회용 원본)", "🧮 최종 값 직접 입력 (자동 조회)"])

        with tab_raw:
            st.caption("엑셀의 **'발송 건수'**·**'RawNew'** 시트와 같은 원본 형태로 입력해 두면, 옆 **'최종 값 직접 입력'** "
                      "탭에서 일자+AF코드만 넣어도 여기 있는 내용을 바탕으로 자동 조회돼요. **따로 계산 버튼을 누를 필요 없어요.**")
            st.markdown("**발송 건수** (발송로그)")
            if "raw_send_ver" not in st.session_state:
                st.session_state.raw_send_ver = 0
            send_edit = st.data_editor(
                _blank_editor_df(SEND_COUNT_ENTRY_COLS, SEND_COUNT_ENTRY_NUM), num_rows="dynamic",
                use_container_width=True, height=260,
                column_config={
                    "date": st.column_config.TextColumn("발송일시*", help="YYYYMMDD, 예: 20260706"),
                    "camp_name": st.column_config.TextColumn("캠페인명"),
                    "camp_code": st.column_config.TextColumn("캠페인코드"),
                    "auto_code": st.column_config.TextColumn("자동캠페인코드"),
                    "channel": st.column_config.TextColumn("발송채널"),
                    "af": st.column_config.TextColumn("AF코드*"),
                    "camp_type": st.column_config.TextColumn("캠페인구분"),
                    "send_kind": st.column_config.TextColumn("발송구분"),
                    "send": st.column_config.NumberColumn("모수*"),
                }, key=f"raw_send_editor_{st.session_state.raw_send_ver}")

            st.markdown("**RawNew** (성과 원자료)")
            if "raw_perf_ver" not in st.session_state:
                st.session_state.raw_perf_ver = 0
            perf_edit = st.data_editor(
                _blank_editor_df(RAWNEW_ENTRY_COLS, RAWNEW_ENTRY_NUM), num_rows="dynamic",
                use_container_width=True, height=260,
                column_config={
                    "std_ym": st.column_config.TextColumn("STD_YM"),
                    "date": st.column_config.TextColumn("STD_DD*", help="YYYYMMDD, 예: 20260706"),
                    "af": st.column_config.TextColumn("AF_CD*"),
                    "af_nm": st.column_config.TextColumn("AF_NM"),
                    "lctgr": st.column_config.TextColumn("AF_LCTGR_NM"),
                    "mctgr": st.column_config.TextColumn("AF_MCTGR_NM"),
                    "sctgr": st.column_config.TextColumn("AF_SCTGR_NM"),
                    "chnl": st.column_config.TextColumn("CHNL_DTL_CD", help="여러 세부채널이 있으면 'Total' 행을 넣는 걸 권장"),
                    "uv": st.column_config.NumberColumn("UV"),
                    "visit": st.column_config.NumberColumn("SV"),
                    "cust": st.column_config.NumberColumn("CUST_CNT"),
                    "oc": st.column_config.NumberColumn("ORD_CNT"),
                    "sale_amt": st.column_config.NumberColumn("SALE_AMT"),
                    "amt": st.column_config.NumberColumn("ORD_AMT"),
                }, key=f"raw_perf_editor_{st.session_state.raw_perf_ver}")

            if st.button("↺ 두 표 비우기", key="raw_calc_clear"):
                st.session_state.raw_send_ver += 1
                st.session_state.raw_perf_ver += 1
                st.rerun()

        send_lookup = clean_send_count_entry(send_edit)
        perf_lookup = clean_rawnew_entry(perf_edit)

        with tab_final:
            st.caption("**AF코드**와 **일자(YYYYMMDD)** 를 입력하면 왼쪽 **'발송 건수 + RawNew'** 탭에 입력된 내용을 바탕으로 "
                      "**UV/VISIT/고객수/주문건수/거래액**과 **유입전환율/주문전환율/효율**이 엑셀의 SUMIFS 수식과 같은 방식으로 "
                      "자동 조회돼요 (회색으로 비활성화된 컬럼). **발송모수**는 직접 입력한 값이 있으면 그 값을 우선 쓰고, "
                      "비워두면 '발송 건수' 표에서 자동 조회해요 — 값을 바꾸면 유입전환율/주문전환율/효율도 그에 맞게 다시 계산돼요. "
                      "행 추가는 표 맨 아래 + 를 누르면 돼요.")

            plan_url, plan_gid, plan_is_custom = plan_sheet_source()
            with st.expander("☁️ 구글 발송 계획 시트에서 불러오기", expanded=plan_is_custom):
                st.caption("요일·시간대는 비어 있어도 불러오지만, 그 외 항목(타겟구분/발송유형/BPU/우선순위/"
                          "카테고리/속성/담당자/브랜드/AF코드/기획전No.) 중 하나라도 빈 칸인 행은 불러오지 않아요. "
                          "발송모수는 이후 자동 조회로 채워지는 값이라 확인하지 않아요.")
                st.caption(("현재 **직접 연결한 시트**(주차 탭)를 쓰고 있어요." if plan_is_custom else
                           "현재 **기본 시트**를 쓰고 있어요 — 발송 계획은 주차마다 탭이 나뉘어 있어서, "
                           "원하는 주차의 탭 URL을 붙여넣으면 그 주차를 불러올 수 있어요.") +
                          " (URL에 `#gid=...`가 있으면 그 탭을 읽어요.)")
                plan_url_input = st.text_input("구글시트 URL", value=plan_url, key="plan_sheet_url_input")
                pb1, pb2 = st.columns(2)
                if pb1.button("🔗 연결", key="plan_sheet_connect", use_container_width=True):
                    st.session_state.plan_sheet_url = plan_url_input.strip()
                    st.session_state.plan_sheet_gid = _parse_sheet_url_gid(plan_url_input)
                    st.rerun()
                if pb2.button("↺ 기본 시트로 되돌리기", key="plan_sheet_reset", use_container_width=True,
                             disabled=not plan_is_custom):
                    st.session_state.pop("plan_sheet_url", None)
                    st.session_state.pop("plan_sheet_gid", None)
                    st.rerun()
                if st.button("📥 불러오기", key="plan_sheet_import"):
                    try:
                        has_creds = "gcp_service_account" in st.secrets
                    except Exception:
                        has_creds = False
                    if not has_creds:
                        st.error("구글시트 연동이 설정돼 있지 않아요 (Secrets에 gcp_service_account가 필요해요).")
                    else:
                        try:
                            raw_rows, sheet_rows = load_plan_rows_from_gsheet(
                                st.secrets["gcp_service_account"], plan_url, plan_gid)
                        except Exception as e:
                            msg = str(e)[:150].strip() or type(e).__name__
                            st.error(f"구글시트에서 불러오기 실패: {msg}")
                        else:
                            if raw_rows.empty:
                                if not sheet_rows:
                                    st.warning("이 탭 자체가 완전히 비어 있어요. gid가 맞는 탭을 가리키는지 "
                                              "확인해주세요.")
                                else:
                                    st.warning(f"시트에서 총 {len(sheet_rows):,}행을 받았지만, 그 안에서 "
                                              "'일자'·'AF코드' 같은 헤더 행을 찾지 못했어요 (위쪽 100행 안에서 찾아요). "
                                              "아래 원본 미리보기에서 실제 헤더가 몇 번째 줄에 있는지 확인해주세요.")
                                    with st.expander("읽어온 원본 미리보기 (최대 100행)"):
                                        st.dataframe(pd.DataFrame(sheet_rows[:100]), use_container_width=True)
                            else:
                                new_rows = filter_complete_plan_rows(raw_rows)
                                if new_rows.empty:
                                    st.warning(f"시트에서 {len(raw_rows):,}행을 읽었지만, 요일/시간대/발송모수를 뺀 "
                                              "나머지 항목이 모두 채워진 행이 없어요. 항목별로 빈 칸인 행이 몇 개인지 "
                                              "아래에서 확인해 보세요.")
                                    st.dataframe(plan_rows_blank_summary(raw_rows), hide_index=True,
                                                use_container_width=True)
                                    with st.expander("읽어온 원본 미리보기 (최대 10행)"):
                                        st.dataframe(raw_rows.head(10), use_container_width=True)
                                else:
                                    editor_rows = plan_rows_to_editor_df(new_rows)
                                    existing = st.session_state.get("plan_import_rows")
                                    combined = (pd.concat([existing, editor_rows], ignore_index=True)
                                               if existing is not None and not existing.empty else editor_rows)
                                    combined = combined[~store_key_frame(combined).duplicated(keep="last")].reset_index(drop=True)
                                    st.session_state.plan_import_rows = combined
                                    st.session_state.manual_editor_ver += 1
                                    st.success(f"{len(editor_rows):,}건을 불러왔어요 (누적 {len(combined):,}건).")
                                    st.rerun()

            if "manual_editor_ver" not in st.session_state:
                st.session_state.manual_editor_ver = 0
            imported_rows = st.session_state.get("plan_import_rows")
            initial_rows = imported_rows if imported_rows is not None and not imported_rows.empty else blank_row_df()
            edited = st.data_editor(initial_rows, num_rows="dynamic", use_container_width=True, height=280,
                                    column_config=store_column_config(disabled=AUTO_LOOKUP_COLS),
                                    key=f"manual_editor_{st.session_state.manual_editor_ver}")

            if st.button("🔄 새로고침 (발송 건수+RawNew 반영)", key="lookup_refresh"):
                st.rerun()

            meta_clean = clean_manual_rows(edited)
            final_rows = None
            if meta_clean.empty:
                st.info("AF코드와 일자를 입력하면 조회 결과가 여기 표시돼요.")
            else:
                looked_up = lookup_send_uv_etc(meta_clean, send_lookup, perf_lookup)
                edit_cols = ["af", "bpu", "stype", "brand", "send", "uv", "visit", "cust", "oc", "amt"]
                row_id = tuple(zip(looked_up["date"], looked_up["af"]))

                st.markdown("**🔎 자동 조회 결과** — 조회가 안 됐거나 값이 다르면 발송모수/UV/VISIT/고객수/주문건수/거래액을 "
                           "직접 고칠 수 있어요. **여러 칸을 고친 뒤 아래 '✅ 수정 반영'을 한 번만 누르면** 한꺼번에 적용돼요 "
                           "(칸마다 바로바로 재실행되지 않아 빨라요).")
                with st.form(key=f"results_form_{hash(row_id)}"):
                    results_edited = st.data_editor(
                        looked_up[edit_cols], use_container_width=True, hide_index=True, num_rows="fixed",
                        column_config={
                            "af": st.column_config.TextColumn("AF코드", disabled=True),
                            "bpu": st.column_config.TextColumn("BPU", disabled=True),
                            "stype": st.column_config.TextColumn("발송유형", disabled=True),
                            "brand": st.column_config.TextColumn("브랜드", disabled=True),
                            "send": st.column_config.NumberColumn("발송모수"),
                            "uv": st.column_config.NumberColumn("UV"),
                            "visit": st.column_config.NumberColumn("VISIT"),
                            "cust": st.column_config.NumberColumn("고객수"),
                            "oc": st.column_config.NumberColumn("주문건수"),
                            "amt": st.column_config.NumberColumn("거래액"),
                        }, key=f"results_editor_{hash(row_id)}")
                    applied = st.form_submit_button("✅ 수정 반영")

                if applied:
                    st.session_state.results_override = results_edited.copy()
                    st.session_state.results_override_id = row_id

                override = st.session_state.get("results_override")
                if override is not None and st.session_state.get("results_override_id") == row_id:
                    final_rows = looked_up.copy()
                    for c in ("send", "uv", "visit", "cust", "oc", "amt"):
                        final_rows[c] = override[c].values
                    final_rows = compute_ratio_cols(final_rows)
                else:
                    final_rows = looked_up

                st.markdown("**최종 계산값** (저장하기를 누르면 이 값이 저장돼요)")
                show = final_rows.rename(columns={**METRIC_LABELS, "af": "AF코드", "bpu": "BPU", "stype": "발송유형",
                                                  "brand": "브랜드", "visit": "VISIT", "cust": "고객수",
                                                  "infl_cr": "유입전환율", "ord_cr": "주문전환율", "eff": "효율"})
                st.dataframe(show[["AF코드", "BPU", "발송유형", "브랜드", "발송모수", "UV", "VISIT", "고객수",
                                  "주문건수", "거래액", "유입전환율", "주문전환율", "효율"]].style.format({
                    "발송모수": "{:,.0f}", "UV": "{:,.0f}", "VISIT": "{:,.0f}", "고객수": "{:,.0f}",
                    "주문건수": "{:,.0f}", "거래액": "{:,.0f}", "유입전환율": "{:.2%}", "주문전환율": "{:.2%}",
                    "효율": "{:,.2f}",
                }), use_container_width=True, hide_index=True)

            c1, c2 = st.columns(2)
            if c1.button("💾 저장하기", use_container_width=True):
                if final_rows is None or final_rows.empty:
                    st.warning("저장할 유효한 행이 없어요. AF코드와 일자(8자리)를 채워주세요.")
                else:
                    try:
                        old_for_merge = storage_load(BK)
                    except Exception as e:
                        st.error(f"⚠️ 기존 누적 데이터를 불러오지 못해 저장을 중단했어요 ({str(e)[:150].strip()}). "
                                 "이 상태로 저장하면 기존 데이터가 사라질 수 있어요 — 새로고침 후 다시 시도해 주세요.")
                    else:
                        merged = merge_store(old_for_merge, final_rows[STORE_COLS])
                        storage_save(BK, merged)
                        st.success(f"{len(final_rows):,}건 저장 완료 (누적 {len(merged):,}건)")
                        st.session_state.manual_editor_ver += 1
                        st.session_state.pop("results_override", None)
                        st.session_state.pop("results_override_id", None)
                        st.session_state.pop("plan_import_rows", None)
                        st.rerun()
            if c2.button("↺ 비우기", use_container_width=True):
                st.session_state.manual_editor_ver += 1
                st.session_state.pop("results_override", None)
                st.session_state.pop("results_override_id", None)
                st.session_state.pop("plan_import_rows", None)
                st.rerun()

    # ══════════════════════════════════════════════════════════
    # 데이터
    # ══════════════════════════════════════════════════════════
    else:
        st.title("데이터 · 다운로드")
        st.caption(f"누적 저장 {len(df):,}건 · 필터 적용 {len(f):,}건")
        st.download_button("⬇️ CSV 다운로드 (필터 적용)", f.drop(columns=["dt"]).to_csv(index=False).encode("utf-8-sig"),
                          file_name="daily_push_perf_filtered.csv", mime="text/csv")

        st.markdown("---")
        st.markdown("#### 전체 데이터 편집")
        st.caption("행을 수정·삭제하거나 새 행을 추가한 뒤 저장하면, 저장된 전체 데이터가 이 표 내용으로 바뀌어요 (필터와 무관).")
        date_search = st.text_input("일자 검색 (예: 20260629 또는 202606)", "", key="data_editor_date_search")
        full_df = df[STORE_COLS]
        if date_search.strip():
            mask = full_df["date"].astype(str).str.contains(re.escape(date_search.strip()), na=False)
            editor_source, other_rows = full_df[mask], full_df[~mask]
            st.caption(f"검색 결과 {len(editor_source):,}행 표시 중 (전체 {len(full_df):,}행 · 검색에 안 걸린 "
                      f"{len(other_rows):,}행은 화면엔 안 보이지만 저장해도 그대로 남아있어요).")
        else:
            editor_source, other_rows = full_df, full_df.iloc[0:0]
        if "data_editor_ver" not in st.session_state:
            st.session_state.data_editor_ver = 0
        edited_all = st.data_editor(editor_source, num_rows="dynamic", use_container_width=True, height=460,
                                    column_config=store_column_config(),
                                    key=f"data_editor_{st.session_state.data_editor_ver}_{hash(date_search.strip())}")
        if st.button("💾 변경사항 저장", use_container_width=True):
            clean = clean_manual_rows(edited_all)
            merged = pd.concat([other_rows, clean], ignore_index=True) if date_search.strip() else clean
            storage_save(BK, merged)
            st.success(f"{len(merged):,}건으로 저장했어요.")
            st.session_state.data_editor_ver += 1
            st.rerun()


if __name__ == "__main__":
    main()
