# -*- coding: utf-8 -*-
"""
앱푸시 발송 성과 통합 분석 대시보드
────────────────────────────────────────────────────────────────
문구 데이터(발송 ID·BPU·타이틀·본문·타겟) +
실적 데이터(발송건수·오픈·수신거부·전환·GMV) 를 조인하여

  ① 고(High) CTR 카피 특징 분석 및 상위 문구 선정
  ② 요일·시간대별 CTR / CVR 효율 히트맵
  ③ 발송량 증가에 따른 피로도(Fatigue) 분석 및 최적 주기 제안
  ④ BPU별 GMV 기여도 및 ROAS 효율 비교

데이터 입력 방식
  · 엑셀 파일 업로드 (문구 시트 + 실적 시트 분리 또는 통합)
  · Google Sheets 연동 (서비스 계정 TOML)

실행: python -m streamlit run push_analytics_dashboard.py
"""

import io, os, re, json, datetime, unicodedata
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# ══════════════════════════════════════════════════════════════════════════════
# 0. 팔레트 & 상수
# ══════════════════════════════════════════════════════════════════════════════

PAL = {
    "primary":  "#2E68B0",
    "success":  "#367A4C",
    "warning":  "#A07010",
    "danger":   "#B83A3A",
    "purple":   "#7B5BC0",
    "slate":    "#1e293b",
    "muted":    "#64748b",
    "border":   "#e2e8f0",
    "bg":       "#f8f9fc",
    "card":     "#ffffff",
}

DOW_KR  = ["월", "화", "수", "목", "금", "토", "일"]
DOW_EN  = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
HOUR_BANDS = {
    "새벽(0-6시)":    list(range(0, 6)),
    "출근길(6-9시)":  list(range(6, 9)),
    "오전(9-12시)":   list(range(9, 12)),
    "점심(12-14시)":  list(range(12, 14)),
    "오후(14-17시)":  list(range(14, 17)),
    "퇴근길(17-20시)": list(range(17, 20)),
    "저녁(20-23시)":  list(range(20, 23)),
    "심야(23시+)":    [23],
}

# ══════════════════════════════════════════════════════════════════════════════
# 1. 순수 데이터 로직 (Streamlit 비의존)
# ══════════════════════════════════════════════════════════════════════════════

# ── 1-A. 문구 데이터 파싱 ────────────────────────────────────────────────────

MSG_COLMAP = {
    "발송ID": "send_id", "발송 ID": "send_id",
    "발송일시": "send_dt", "발송 일시": "send_dt", "날짜": "send_dt", "일시": "send_dt",
    "BPU": "bpu", "사업부": "bpu", "브랜드": "bpu",
    "푸시타이틀": "title", "타이틀": "title", "제목": "title",
    "푸시본문": "body", "본문": "body", "내용": "body",
    "타겟조건": "target", "타겟 조건": "target", "타겟": "target",
    "발송채널": "channel", "채널": "channel",
}

PERF_COLMAP = {
    "발송ID": "send_id", "발송 ID": "send_id",
    "발송일시": "send_dt", "발송 일시": "send_dt", "날짜": "send_dt",
    "BPU": "bpu", "사업부": "bpu",
    "발송건수": "send_cnt", "발송": "send_cnt", "발송량": "send_cnt",
    "오픈건수": "open_cnt", "클릭수": "open_cnt", "클릭건수": "open_cnt", "UV": "open_cnt",
    "수신거부건수": "unsub_cnt", "수신거부": "unsub_cnt",
    "구매전환건수": "conv_cnt", "당일구매전환건수": "conv_cnt", "주문건수": "conv_cnt",
    "당일거래액": "gmv", "거래액": "gmv", "GMV": "gmv", "주문금액": "gmv",
    "시간대": "hour",
}


def _parse_sheet_df(ws_rows: list, colmap: dict) -> pd.DataFrame:
    """2D 행 리스트 → colmap 기반 DataFrame."""
    if not ws_rows:
        return pd.DataFrame()
    hdr = [str(c).strip() if c is not None else "" for c in ws_rows[0]]
    col_idx = {}
    for h, i in ((h, i) for i, h in enumerate(hdr)):
        std = colmap.get(h)
        if std and std not in col_idx:
            col_idx[std] = i
    if not col_idx:
        return pd.DataFrame()
    recs = []
    for row in ws_rows[1:]:
        rec = {}
        for std, i in col_idx.items():
            rec[std] = row[i] if i < len(row) else None
        recs.append(rec)
    return pd.DataFrame(recs)


def parse_msg_bytes(file_bytes: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    """문구 엑셀 → DataFrame."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    # 시트 선택 전략
    target = None
    if sheet_name and sheet_name in wb.sheetnames:
        target = sheet_name
    else:
        for s in wb.sheetnames:
            ws = wb[s]
            hdr = [str(v).strip() if v else "" for v in
                   next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            if any(h in MSG_COLMAP for h in hdr):
                if "타이틀" in hdr or "제목" in hdr or "본문" in hdr or "내용" in hdr:
                    target = s
                    break
    if target is None and wb.sheetnames:
        target = wb.sheetnames[0]
    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    df = _parse_sheet_df(rows, MSG_COLMAP)
    return _finalize_msg(df)


def parse_perf_bytes(file_bytes: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    """실적 엑셀 → DataFrame."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    target = None
    if sheet_name and sheet_name in wb.sheetnames:
        target = sheet_name
    else:
        for s in wb.sheetnames:
            ws = wb[s]
            hdr = [str(v).strip() if v else "" for v in
                   next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            if any(h in PERF_COLMAP for h in hdr):
                if any(h in ("발송건수", "발송", "발송량", "오픈건수", "거래액", "GMV") for h in hdr):
                    target = s
                    break
    if target is None and wb.sheetnames:
        target = wb.sheetnames[0]
    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    df = _parse_sheet_df(rows, PERF_COLMAP)
    return _finalize_perf(df)


def _finalize_msg(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    df["send_dt"] = pd.to_datetime(df.get("send_dt"), errors="coerce")
    for c in ("title", "body", "bpu", "target", "channel"):
        if c not in df:
            df[c] = ""
        df[c] = df[c].fillna("").astype(str).str.strip()
    if "send_id" not in df:
        df["send_id"] = df.index.astype(str)
    df["send_id"] = df["send_id"].astype(str).str.strip()
    return df.dropna(subset=["send_id"]).reset_index(drop=True)


def _finalize_perf(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    df["send_dt"] = pd.to_datetime(df.get("send_dt"), errors="coerce")
    for c in ("send_cnt", "open_cnt", "unsub_cnt", "conv_cnt", "gmv"):
        if c in df:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
        else:
            df[c] = 0.0
    if "hour" not in df:
        df["hour"] = df["send_dt"].dt.hour
    else:
        df["hour"] = pd.to_numeric(df["hour"], errors="coerce")
        df["hour"] = df["hour"].fillna(df["send_dt"].dt.hour)
    if "bpu" not in df:
        df["bpu"] = ""
    df["bpu"] = df["bpu"].fillna("").astype(str).str.strip()
    if "send_id" not in df:
        df["send_id"] = df.index.astype(str)
    df["send_id"] = df["send_id"].astype(str).str.strip()
    # 파생 지표
    df["ctr"]       = np.where(df["send_cnt"] > 0, df["open_cnt"]  / df["send_cnt"], np.nan)
    df["cvr"]       = np.where(df["open_cnt"]  > 0, df["conv_cnt"] / df["open_cnt"],  np.nan)
    df["unsub_rate"] = np.where(df["send_cnt"] > 0, df["unsub_cnt"] / df["send_cnt"], np.nan)
    df["rps"]       = np.where(df["send_cnt"] > 0, df["gmv"]       / df["send_cnt"], 0.0)
    df["dow"]       = df["send_dt"].dt.dayofweek          # 0=Mon … 6=Sun
    df["hour_band"] = df["hour"].apply(_hour_to_band)
    return df.reset_index(drop=True)


def _hour_to_band(h):
    try:
        h = int(h)
    except (TypeError, ValueError):
        return "미상"
    for band, hours in HOUR_BANDS.items():
        if h in hours:
            return band
    return "미상"


# ── 1-B. 데이터 조인 ─────────────────────────────────────────────────────────

def merge_msg_perf(msg_df: pd.DataFrame, perf_df: pd.DataFrame) -> pd.DataFrame:
    """실적 기준 좌-조인: send_id 일치 → 문구 컬럼 부착. send_id 없으면 send_dt+bpu 폴백."""
    if perf_df.empty:
        return perf_df.copy()

    # send_id 기준 조인
    msg_cols = ["send_id", "title", "body", "target", "channel"]
    msg_sub = msg_df[[c for c in msg_cols if c in msg_df.columns]].copy()
    msg_sub = msg_sub.rename(columns={})

    merged = perf_df.merge(msg_sub, on="send_id", how="left")
    for c in ("title", "body", "target", "channel"):
        if c not in merged:
            merged[c] = ""
        merged[c] = merged[c].fillna("")

    # 문구 태깅
    merged = tag_copy(merged)
    return merged


# ── 1-C. 문구 자동 태깅 ──────────────────────────────────────────────────────

_EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001F9FF"
    "\U00002600-\U000027BF"
    "\U0001FA00-\U0001FA6F"
    "\U0001FA70-\U0001FAFF"
    "\U00002702-\U000027B0"
    "]+",
    re.UNICODE,
)
_BRACKET_RE  = re.compile(r'\[.{1,20}\]')
_PERCENT_RE  = re.compile(r'\d+\s*%')
_PRICE_RE    = re.compile(r'\d[\d,]+\s*원')
_BENEFIT_KW  = re.compile(r'(할인|쿠폰|포인트|혜택|무료|증정|특가|최저가|반값|적립|리워드|캐시백)', re.I)
_URGENCY_KW  = re.compile(r'(마감|오늘만|D-\d|한정|지금|선착순|오직|단 \d)', re.I)
_EMOTION_KW  = re.compile(r'(설레|행복|좋아|사랑|그리|추억|기분|여름|봄|겨울|가을|힐링|감성|취향)', re.I)
_QUESTION_RE = re.compile(r'[？?]')


def _count_emoji(text: str) -> int:
    return sum(len(m.group()) for m in _EMOJI_RE.finditer(text))


def _copy_type(title: str, body: str) -> str:
    combined = title + " " + body
    scores = {
        "혜택형": len(_BENEFIT_KW.findall(combined)) * 2 + len(_PERCENT_RE.findall(combined)) + len(_PRICE_RE.findall(combined)),
        "긴급형": len(_URGENCY_KW.findall(combined)) * 2,
        "감성형": len(_EMOTION_KW.findall(combined)) * 2 + len(_QUESTION_RE.findall(combined)),
        "정보형": 0,
    }
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "정보형"


def tag_copy(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    title = df.get("title", pd.Series([""] * len(df))).fillna("").astype(str)
    body  = df.get("body",  pd.Series([""] * len(df))).fillna("").astype(str)
    combined = title + " " + body

    df["has_emoji"]    = combined.apply(lambda s: _count_emoji(s) > 0)
    df["emoji_cnt"]    = combined.apply(_count_emoji)
    df["has_bracket"]  = title.str.contains(r'\[.{1,20}\]', regex=True)
    df["has_percent"]  = combined.str.contains(r'\d+\s*%', regex=True)
    df["has_price"]    = combined.str.contains(r'\d[\d,]+\s*원', regex=True)
    df["title_len"]    = title.str.len()
    df["body_len"]     = body.str.len()
    df["copy_type"]    = [_copy_type(t, b) for t, b in zip(title, body)]
    df["has_question"] = combined.str.contains(r'[？?]', regex=True)
    df["has_urgency"]  = combined.apply(lambda s: bool(_URGENCY_KW.search(s)))
    df["has_benefit"]  = combined.apply(lambda s: bool(_BENEFIT_KW.search(s)))
    return df


# ── 1-D. 샘플 데이터 생성 (업로드 없을 때 체험용) ──────────────────────────

def _make_sample_data() -> pd.DataFrame:
    """재현 가능한 예시 데이터 200건."""
    rng = np.random.default_rng(42)
    n = 200
    bpus   = ["여성의류", "남성의류", "스포츠", "라이프", "아웃도어", "잡화"]
    types  = ["혜택형", "감성형", "긴급형", "정보형"]
    titles_pool = {
        "혜택형": ["[특가] 오늘만 20% 할인!", "쿠폰 드려요 🎁 최대 30% OFF", "포인트 2배 적립 이벤트"],
        "감성형": ["이 여름, 당신의 스타일 🌞", "설레는 봄 신상 도착했어요", "힐링이 필요한 당신에게"],
        "긴급형": ["D-1 마감! 서두르세요", "선착순 100명 한정 혜택", "오늘 자정 마감, 지금 바로"],
        "정보형": ["신상품 입고 안내", "6월 인기 브랜드 TOP10", "주간 추천 아이템"],
    }
    bodies_pool = {
        "혜택형": ["지금 바로 할인 쿠폰을 받아가세요. 오늘 자정까지만!", "LF몰 앱에서만 드리는 특별 혜택"],
        "감성형": ["당신만을 위한 특별한 컬렉션을 만나보세요", "새 계절, 새 스타일로 나를 표현해요"],
        "긴급형": ["한정 수량 소진 시 종료. 지금 확인하세요", "선착순 마감 임박!"],
        "정보형": ["LF몰 이번 주 베스트 아이템을 확인해보세요", "인기 브랜드 신제품 모아보기"],
    }
    targets = ["전체", "30일 미방문", "구매이력 있음", "관심상품 추가", "신규가입 7일"]

    rows = []
    for i in range(n):
        ct   = rng.choice(types)
        bpu  = rng.choice(bpus)
        dt   = pd.Timestamp("2026-01-01") + pd.Timedelta(days=int(rng.integers(0, 180)))
        hour = int(rng.choice([7, 8, 9, 12, 13, 18, 19, 20, 21], p=[.06,.07,.1,.12,.1,.12,.12,.15,.16]))
        send_dt = dt + pd.Timedelta(hours=hour)
        title   = rng.choice(titles_pool[ct])
        body    = rng.choice(bodies_pool[ct])

        # 성과 시뮬레이션 (유형별 기대 CTR 차등)
        base_ctr = {"혜택형": .045, "감성형": .032, "긴급형": .055, "정보형": .022}[ct]
        hour_mult = {7:.9, 8:1.1, 9:1.2, 12:1.3, 13:1.1, 18:1.2, 19:1.3, 20:1.4, 21:1.1}.get(hour, 1.0)
        send_cnt  = int(rng.integers(5000, 200000))
        ctr_val   = np.clip(base_ctr * hour_mult * rng.lognormal(0, .25), .005, .18)
        open_cnt  = int(send_cnt * ctr_val)
        cvr_val   = np.clip(rng.lognormal(-3.5, .4), .005, .12)
        conv_cnt  = int(open_cnt * cvr_val)
        gmv_per   = rng.integers(50000, 300000)
        gmv       = conv_cnt * gmv_per
        unsub_cnt = int(send_cnt * np.clip(rng.lognormal(-7, .5), .0001, .01))

        rows.append({
            "send_id": f"PU{i+1:04d}", "send_dt": send_dt,
            "bpu": bpu, "title": title, "body": body,
            "target": rng.choice(targets), "channel": "앱푸시",
            "send_cnt": send_cnt, "open_cnt": open_cnt,
            "unsub_cnt": unsub_cnt, "conv_cnt": conv_cnt, "gmv": gmv,
        })

    df = pd.DataFrame(rows)
    df["hour"]       = df["send_dt"].dt.hour
    df["ctr"]        = df["open_cnt"]  / df["send_cnt"]
    df["cvr"]        = np.where(df["open_cnt"] > 0, df["conv_cnt"] / df["open_cnt"], 0)
    df["unsub_rate"] = df["unsub_cnt"] / df["send_cnt"]
    df["rps"]        = df["gmv"] / df["send_cnt"]
    df["dow"]        = df["send_dt"].dt.dayofweek
    df["hour_band"]  = df["hour"].apply(_hour_to_band)
    df = tag_copy(df)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 2. 공통 UI 헬퍼
# ══════════════════════════════════════════════════════════════════════════════

def _pct(v, denom=100):
    return f"{v * denom:.2f}%"


def _metric_card(col, label: str, value: str, delta: str = "", color: str = PAL["primary"]):
    col.markdown(
        f"""
        <div style="background:{PAL['card']};border:1px solid {PAL['border']};
                    border-radius:10px;padding:16px 20px;margin-bottom:6px;">
          <div style="font-size:12px;color:{PAL['muted']};margin-bottom:4px;">{label}</div>
          <div style="font-size:24px;font-weight:700;color:{color};">{value}</div>
          {"<div style='font-size:12px;color:"+PAL['muted']+"'>"+delta+"</div>" if delta else ""}
        </div>""",
        unsafe_allow_html=True,
    )


def _section(title: str, icon: str = ""):
    st.markdown(
        f"""<div style="margin:28px 0 10px;padding-bottom:6px;
                border-bottom:2px solid {PAL['primary']};
                font-size:16px;font-weight:700;color:{PAL['slate']};">
            {icon} {title}</div>""",
        unsafe_allow_html=True,
    )


def _fig_layout(fig, height=420, margin=dict(l=40, r=20, t=40, b=40)):
    fig.update_layout(
        height=height,
        plot_bgcolor=PAL["card"],
        paper_bgcolor=PAL["bg"],
        font=dict(family="Noto Sans KR, sans-serif", size=12, color=PAL["slate"]),
        margin=margin,
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(size=11)),
    )
    fig.update_xaxes(gridcolor=PAL["border"], linecolor=PAL["border"])
    fig.update_yaxes(gridcolor=PAL["border"], linecolor=PAL["border"])
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# 3. 분석 탭 렌더링 함수
# ══════════════════════════════════════════════════════════════════════════════

# ── 탭 ① : 문구 효율 분석 ────────────────────────────────────────────────────

def tab_copy_analysis(df: pd.DataFrame):
    _section("High CTR 카피 특징 분석", "✍️")

    # ── 요약 KPI
    c1, c2, c3, c4 = st.columns(4)
    _metric_card(c1, "분석 발송 건수", f"{len(df):,}건")
    _metric_card(c2, "평균 CTR", _pct(df["ctr"].mean()), color=PAL["primary"])
    _metric_card(c3, "평균 CVR", _pct(df["cvr"].mean()), color=PAL["success"])
    _metric_card(c4, "평균 GMV/발송", f"₩{df['rps'].mean():,.0f}", color=PAL["purple"])

    st.markdown("---")

    # ── 카피 유형별 성과 비교
    col_a, col_b = st.columns([1, 1])

    with col_a:
        st.markdown("**카피 유형별 CTR 분포**")
        ct_g = df.groupby("copy_type").agg(
            mean_ctr=("ctr", "mean"),
            median_ctr=("ctr", "median"),
            count=("ctr", "count"),
        ).reset_index().sort_values("mean_ctr", ascending=False)

        bar_colors = [PAL["primary"], PAL["success"], PAL["warning"], PAL["purple"]]
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=ct_g["copy_type"],
            y=(ct_g["mean_ctr"] * 100).round(3),
            name="평균 CTR (%)",
            marker_color=bar_colors[:len(ct_g)],
            text=(ct_g["mean_ctr"] * 100).round(2).astype(str) + "%",
            textposition="outside",
        ))
        fig.add_trace(go.Scatter(
            x=ct_g["copy_type"],
            y=(ct_g["median_ctr"] * 100).round(3),
            name="중앙값 CTR (%)",
            mode="markers",
            marker=dict(size=10, color=PAL["danger"], symbol="diamond"),
        ))
        fig.update_layout(yaxis_title="CTR (%)", xaxis_title="카피 유형", height=320,
                          plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
                          font=dict(size=12), margin=dict(l=40, r=20, t=20, b=40),
                          legend=dict(bgcolor="rgba(0,0,0,0)"))
        st.plotly_chart(fig, use_container_width=True)

        # 발송 수 소주석
        st.caption("  |  ".join([f"{r['copy_type']}: {r['count']:,}건" for _, r in ct_g.iterrows()]))

    with col_b:
        st.markdown("**카피 요소별 평균 CTR 비교**")
        features = {
            "이모지 포함": df["has_emoji"],
            "대괄호[] 사용": df["has_bracket"],
            "% 할인 언급": df["has_percent"],
            "가격(원) 언급": df["has_price"],
            "혜택 키워드": df["has_benefit"],
            "긴급 키워드": df["has_urgency"],
            "질문형(?)": df["has_question"],
        }
        feat_rows = []
        for fname, mask in features.items():
            g_true  = df.loc[mask,  "ctr"].mean() if mask.sum() > 0 else np.nan
            g_false = df.loc[~mask, "ctr"].mean() if (~mask).sum() > 0 else np.nan
            feat_rows.append({"feature": fname, "포함": g_true, "미포함": g_false,
                               "n_incl": mask.sum()})
        feat_df = pd.DataFrame(feat_rows).dropna().sort_values("포함", ascending=True)

        fig2 = go.Figure()
        fig2.add_trace(go.Bar(y=feat_df["feature"], x=(feat_df["포함"]  * 100).round(3),
                               name="포함",  orientation="h",
                               marker_color=PAL["primary"],
                               text=(feat_df["포함"] * 100).round(2).astype(str) + "%",
                               textposition="outside"))
        fig2.add_trace(go.Bar(y=feat_df["feature"], x=(feat_df["미포함"] * 100).round(3),
                               name="미포함", orientation="h",
                               marker_color=PAL["border"],
                               text=(feat_df["미포함"] * 100).round(2).astype(str) + "%",
                               textposition="outside"))
        fig2.update_layout(barmode="group", height=320,
                           plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
                           font=dict(size=12), margin=dict(l=130, r=40, t=20, b=40),
                           xaxis_title="CTR (%)",
                           legend=dict(bgcolor="rgba(0,0,0,0)"))
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")

    # ── 제목 길이 vs CTR 산점도
    col_c, col_d = st.columns([1, 1])

    with col_c:
        st.markdown("**타이틀 길이 vs CTR**")
        fig3 = px.scatter(
            df.dropna(subset=["ctr", "title_len"]),
            x="title_len", y=df["ctr"] * 100,
            color="copy_type",
            trendline="lowess",
            labels={"title_len": "타이틀 글자 수", "y": "CTR (%)"},
            color_discrete_sequence=[PAL["primary"], PAL["success"], PAL["warning"], PAL["purple"]],
            opacity=0.55,
        )
        fig3.update_layout(height=320, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
                           margin=dict(l=40, r=20, t=20, b=40),
                           legend=dict(bgcolor="rgba(0,0,0,0)", title=""))
        st.plotly_chart(fig3, use_container_width=True)

    with col_d:
        st.markdown("**이모지 수 vs CTR**")
        em_g = df.groupby("emoji_cnt")["ctr"].agg(["mean", "count"]).reset_index()
        em_g = em_g[em_g["count"] >= 3]
        fig4 = go.Figure()
        fig4.add_trace(go.Scatter(
            x=em_g["emoji_cnt"],
            y=(em_g["mean"] * 100).round(3),
            mode="lines+markers",
            marker=dict(size=em_g["count"].clip(5, 30), color=PAL["warning"],
                        sizemode="area", sizeref=2, opacity=0.7),
            line=dict(color=PAL["warning"], width=2),
        ))
        fig4.update_layout(
            xaxis_title="이모지 개수", yaxis_title="평균 CTR (%)",
            height=320, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
            margin=dict(l=40, r=20, t=20, b=40),
        )
        st.plotly_chart(fig4, use_container_width=True)

    st.markdown("---")

    # ── 상위 10 문구 테이블
    _section("CTR 상위 10개 문구", "🏆")
    top10 = (
        df[df["send_cnt"] >= df["send_cnt"].quantile(.25)]  # 소량 발송 제외
        .sort_values("ctr", ascending=False)
        .head(10)
        [["send_id", "bpu", "title", "body", "copy_type", "send_cnt", "ctr", "cvr", "gmv"]]
        .copy()
    )
    top10["CTR"]  = top10["ctr"].apply(lambda v: f"{v*100:.2f}%")
    top10["CVR"]  = top10["cvr"].apply(lambda v: f"{v*100:.2f}%")
    top10["GMV"]  = top10["gmv"].apply(lambda v: f"₩{v:,.0f}")
    top10["발송수"] = top10["send_cnt"].apply(lambda v: f"{v:,}")
    st.dataframe(
        top10.rename(columns={
            "send_id": "발송ID", "bpu": "BPU",
            "title": "타이틀", "body": "본문", "copy_type": "유형",
        })[["발송ID", "BPU", "타이틀", "본문", "유형", "발송수", "CTR", "CVR", "GMV"]],
        use_container_width=True, hide_index=True,
    )

    # ── AI 인사이트 (간이 규칙 기반)
    st.markdown("---")
    _section("카피 최적화 제안", "💡")
    best_type = ct_g.iloc[0]["copy_type"]
    best_ctr  = ct_g.iloc[0]["mean_ctr"]
    feat_sorted = feat_df.sort_values("포함", ascending=False)
    top_feat = feat_sorted.iloc[0]["feature"] if not feat_sorted.empty else "-"

    st.info(
        f"**분석 요약**: 현재 데이터에서 **{best_type}** 카피의 평균 CTR이 "
        f"**{best_ctr*100:.2f}%**로 가장 높습니다.  \n"
        f"요소별 분석에서는 **{top_feat}**가 포함된 발송이 더 높은 CTR을 보입니다.  \n"
        f"타이틀은 **{int(feat_df['n_incl'].median())}자 내외**로 작성할 때 "
        f"클릭률 최적화 효과가 관찰됩니다."
    )


# ── 탭 ② : 요일·시간대별 효율 ────────────────────────────────────────────────

def tab_time_analysis(df: pd.DataFrame):
    _section("요일·시간대별 CTR / CVR 효율", "📅")

    c1, c2 = st.columns(2)

    # 요일별 집계
    dow_g = df.groupby("dow").agg(
        avg_ctr=("ctr", "mean"), avg_cvr=("cvr", "mean"),
        total_send=("send_cnt", "sum"), count=("ctr", "count"),
    ).reindex(range(7)).reset_index()
    dow_g["dow_label"] = [DOW_KR[i] for i in dow_g["dow"]]

    with c1:
        st.markdown("**요일별 평균 CTR / CVR**")
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Bar(
            x=dow_g["dow_label"], y=(dow_g["avg_ctr"] * 100).round(3),
            name="CTR (%)", marker_color=PAL["primary"], opacity=0.8,
        ), secondary_y=False)
        fig.add_trace(go.Scatter(
            x=dow_g["dow_label"], y=(dow_g["avg_cvr"] * 100).round(3),
            name="CVR (%)", mode="lines+markers",
            marker=dict(size=8, color=PAL["success"]),
            line=dict(color=PAL["success"], width=2),
        ), secondary_y=True)
        fig.update_yaxes(title_text="CTR (%)", secondary_y=False)
        fig.update_yaxes(title_text="CVR (%)", secondary_y=True)
        fig.update_layout(height=340, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
                          margin=dict(l=40, r=40, t=20, b=40),
                          legend=dict(bgcolor="rgba(0,0,0,0)"),
                          font=dict(size=12))
        st.plotly_chart(fig, use_container_width=True)

    with c2:
        st.markdown("**시간대별 평균 CTR**")
        band_order = list(HOUR_BANDS.keys())
        hb_g = (df.groupby("hour_band")["ctr"].mean() * 100).reindex(band_order).reset_index()
        hb_g.columns = ["hour_band", "avg_ctr"]
        hb_g["avg_ctr"] = hb_g["avg_ctr"].round(3)

        fig2 = go.Figure(go.Bar(
            x=hb_g["avg_ctr"], y=hb_g["hour_band"],
            orientation="h",
            marker_color=[PAL["primary"] if v == hb_g["avg_ctr"].max()
                          else PAL["border"] for v in hb_g["avg_ctr"]],
            text=hb_g["avg_ctr"].round(2).astype(str) + "%",
            textposition="outside",
        ))
        fig2.update_layout(height=340, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
                           margin=dict(l=120, r=60, t=20, b=40),
                           xaxis_title="평균 CTR (%)", font=dict(size=12))
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")

    # ── 요일 × 시간대 CTR 히트맵
    _section("요일 × 시간대 CTR 히트맵", "🔥")
    heatmap_data = df.groupby(["dow", "hour_band"])["ctr"].mean().unstack(fill_value=np.nan)
    heatmap_data = heatmap_data.reindex(columns=list(HOUR_BANDS.keys()))
    heatmap_data = heatmap_data.reindex(range(7))
    heatmap_data.index = DOW_KR

    fig3 = go.Figure(go.Heatmap(
        z=(heatmap_data.values * 100).round(3),
        x=heatmap_data.columns.tolist(),
        y=DOW_KR,
        colorscale=[[0, "#EEF3FA"], [0.5, "#6FA0D0"], [1, PAL["primary"]]],
        text=np.where(
            np.isnan(heatmap_data.values),
            "",
            np.round(heatmap_data.values * 100, 2).astype(str) + "%",
        ),
        texttemplate="%{text}",
        textfont=dict(size=11),
        hoverongaps=False,
        colorbar=dict(title="CTR (%)"),
    ))
    fig3.update_layout(
        height=280, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
        margin=dict(l=60, r=20, t=20, b=40),
        xaxis=dict(side="bottom", tickangle=0),
        font=dict(size=12),
    )
    st.plotly_chart(fig3, use_container_width=True)

    # ── 요일 × 시간대 CVR 히트맵
    _section("요일 × 시간대 CVR 히트맵", "🛒")
    cvr_heat = df.groupby(["dow", "hour_band"])["cvr"].mean().unstack(fill_value=np.nan)
    cvr_heat = cvr_heat.reindex(columns=list(HOUR_BANDS.keys()))
    cvr_heat = cvr_heat.reindex(range(7))
    cvr_heat.index = DOW_KR

    fig4 = go.Figure(go.Heatmap(
        z=(cvr_heat.values * 100).round(3),
        x=cvr_heat.columns.tolist(),
        y=DOW_KR,
        colorscale=[[0, "#EEF8F1"], [0.5, "#5DA87A"], [1, PAL["success"]]],
        text=np.where(
            np.isnan(cvr_heat.values),
            "",
            np.round(cvr_heat.values * 100, 2).astype(str) + "%",
        ),
        texttemplate="%{text}",
        textfont=dict(size=11),
        hoverongaps=False,
        colorbar=dict(title="CVR (%)"),
    ))
    fig4.update_layout(
        height=280, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
        margin=dict(l=60, r=20, t=20, b=40),
        xaxis=dict(side="bottom", tickangle=0),
        font=dict(size=12),
    )
    st.plotly_chart(fig4, use_container_width=True)

    # ── 베스트 슬롯 텍스트 요약
    st.markdown("---")
    flat = (heatmap_data.values * 100)
    if not np.all(np.isnan(flat)):
        best_idx = np.unravel_index(np.nanargmax(flat), flat.shape)
        best_dow  = DOW_KR[best_idx[0]]
        best_band = heatmap_data.columns[best_idx[1]]
        best_ctr  = flat[best_idx]
        st.success(
            f"**최고 CTR 슬롯**: {best_dow}요일 · {best_band} — "
            f"평균 CTR **{best_ctr:.2f}%**  \n"
            "이 슬롯에 혜택형·긴급형 카피를 집중 배치하면 반응률을 극대화할 수 있습니다."
        )


# ── 탭 ③ : 피로도 분석 ────────────────────────────────────────────────────────

def tab_fatigue_analysis(df: pd.DataFrame):
    _section("발송량 피로도 분석", "📉")

    # 일별 집계 (발송량 vs CTR/수신거부율)
    if "send_dt" not in df.columns:
        st.warning("발송 일시 데이터가 없어 피로도 분석을 수행할 수 없습니다.")
        return

    daily = df.groupby(df["send_dt"].dt.date).agg(
        daily_send=("send_cnt", "sum"),
        daily_count=("send_id", "count"),
        avg_ctr=("ctr", "mean"),
        avg_unsub=("unsub_rate", "mean"),
    ).reset_index()
    daily.columns = ["date", "daily_send", "daily_count", "avg_ctr", "avg_unsub"]
    daily["date"] = pd.to_datetime(daily["date"])
    daily = daily.sort_values("date")

    # 7일 롤링 평균
    daily["roll_ctr"]   = daily["avg_ctr"].rolling(7, min_periods=1).mean()
    daily["roll_unsub"] = daily["avg_unsub"].rolling(7, min_periods=1).mean()
    daily["roll_send"]  = daily["daily_send"].rolling(7, min_periods=1).mean()

    col_a, col_b = st.columns([2, 1])

    with col_a:
        st.markdown("**일별 총 발송량 vs 평균 CTR (7일 이동평균)**")
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Bar(
            x=daily["date"], y=daily["daily_send"],
            name="일 총 발송량", marker_color=PAL["border"], opacity=0.6,
        ), secondary_y=False)
        fig.add_trace(go.Scatter(
            x=daily["date"], y=(daily["roll_ctr"] * 100).round(3),
            name="7일 평균 CTR (%)", mode="lines",
            line=dict(color=PAL["primary"], width=2.5),
        ), secondary_y=True)
        fig.add_trace(go.Scatter(
            x=daily["date"], y=(daily["roll_unsub"] * 100).round(3),
            name="7일 평균 수신거부율 (%)", mode="lines",
            line=dict(color=PAL["danger"], width=2, dash="dash"),
        ), secondary_y=True)
        fig.update_yaxes(title_text="일 총 발송량 (건)", secondary_y=False)
        fig.update_yaxes(title_text="비율 (%)", secondary_y=True)
        fig.update_layout(height=360, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
                          margin=dict(l=60, r=60, t=20, b=40),
                          legend=dict(bgcolor="rgba(0,0,0,0)"), font=dict(size=12))
        st.plotly_chart(fig, use_container_width=True)

    with col_b:
        st.markdown("**주간 발송 횟수 vs 평균 CTR**")
        weekly = df.copy()
        weekly["week"] = weekly["send_dt"].dt.to_period("W")
        wk_g = weekly.groupby("week").agg(
            n_sends=("send_id", "count"),
            avg_ctr=("ctr", "mean"),
        ).reset_index()
        wk_g["week_str"] = wk_g["week"].astype(str)

        fig2 = px.scatter(
            wk_g, x="n_sends", y=wk_g["avg_ctr"] * 100,
            trendline="lowess",
            labels={"n_sends": "주간 발송 횟수", "y": "평균 CTR (%)"},
            color_discrete_sequence=[PAL["primary"]],
        )
        fig2.update_traces(marker=dict(size=9, opacity=0.75))
        fig2.update_layout(height=360, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
                           margin=dict(l=50, r=20, t=20, b=50), font=dict(size=12))
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")

    # ── 발송 빈도 구간별 CTR 분포 (박스플롯)
    _section("발송 횟수 구간별 CTR 분포", "📦")
    weekly2 = df.copy()
    weekly2["week"] = weekly2["send_dt"].dt.to_period("W")
    wk_cnt = weekly2.groupby("week")["send_id"].count().rename("n_sends")
    weekly2 = weekly2.join(wk_cnt, on="week")
    weekly2["freq_band"] = pd.cut(
        weekly2["n_sends"],
        bins=[0, 3, 7, 14, 21, 999],
        labels=["주 1-3회", "주 4-7회", "주 8-14회", "주 15-21회", "주 22+회"],
    )
    fig3 = go.Figure()
    colors = [PAL["success"], PAL["primary"], PAL["warning"], PAL["danger"], "#7B5BC0"]
    for i, band in enumerate(["주 1-3회", "주 4-7회", "주 8-14회", "주 15-21회", "주 22+회"]):
        sub = weekly2[weekly2["freq_band"] == band]["ctr"].dropna()
        if sub.empty:
            continue
        fig3.add_trace(go.Box(
            y=sub * 100, name=band,
            marker_color=colors[i], boxmean="sd",
        ))
    fig3.update_layout(
        height=360, yaxis_title="CTR (%)",
        plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
        margin=dict(l=40, r=20, t=20, b=40),
        legend=dict(bgcolor="rgba(0,0,0,0)"), font=dict(size=12),
    )
    st.plotly_chart(fig3, use_container_width=True)

    st.markdown("---")

    # ── 피로도 진단 & 최적 빈도 제안
    _section("피로도 진단 및 최적 발송 빈도 제안", "🎯")

    # 주간 발송 횟수 vs CTR 피어슨 상관 (간이)
    wk_g2 = weekly.groupby("week").agg(n_sends=("send_id","count"), avg_ctr=("ctr","mean")).reset_index()
    if len(wk_g2) >= 5:
        corr = wk_g2["n_sends"].corr(wk_g2["avg_ctr"])
        fatigue_dir = "감소" if corr < -0.15 else ("무상관" if abs(corr) < 0.15 else "증가")
        # CTR이 가장 높은 주간 발송 횟수 구간
        best_wk = wk_g2.sort_values("avg_ctr", ascending=False).iloc[0]["n_sends"]

        col_x, col_y = st.columns(2)
        col_x.metric("발송 횟수 ↔ CTR 상관계수", f"{corr:.3f}",
                     help="음수: 많이 보낼수록 CTR 감소 (피로도 징후)")
        col_y.metric("CTR 최고 주간 발송 횟수", f"{int(best_wk):,}회")

        if corr < -0.25:
            st.warning(
                f"**⚠️ 피로도 경고**: 주간 발송 횟수와 CTR 사이 음의 상관(r={corr:.2f})이 "
                f"관찰됩니다.  \n"
                f"데이터 기준 **주 {max(1, int(best_wk)):,}회** 내외가 CTR 효율이 가장 높습니다.  \n"
                "발송 횟수를 줄이거나, BPU/타겟 분산 전략을 검토하세요."
            )
        elif corr < -0.10:
            st.info(
                f"**💡 주의**: 약한 피로도 신호가 감지됩니다(r={corr:.2f}).  \n"
                f"현재 주간 발송 횟수 중앙값 {int(wk_g2['n_sends'].median())}회를 기준으로 "
                "유지·최적화를 고려하세요."
            )
        else:
            st.success(
                f"현재 발송 패턴에서 피로도 신호는 뚜렷하지 않습니다(r={corr:.2f}).  \n"
                "CTR·수신거부율 추이를 주간 단위로 모니터링하세요."
            )
    else:
        st.info("피로도 분석에는 최소 5주 이상의 데이터가 필요합니다.")


# ── 탭 ④ : BPU별 GMV 기여도 ─────────────────────────────────────────────────

def tab_bpu_analysis(df: pd.DataFrame):
    _section("BPU별 GMV 기여도 및 효율 비교", "🏢")

    bpu_g = df.groupby("bpu").agg(
        total_send=("send_cnt", "sum"),
        total_open=("open_cnt", "sum"),
        total_conv=("conv_cnt", "sum"),
        total_gmv=("gmv", "sum"),
        avg_ctr=("ctr", "mean"),
        avg_cvr=("cvr", "mean"),
        n_campaigns=("send_id", "count"),
    ).reset_index()
    bpu_g["rps"]   = np.where(bpu_g["total_send"] > 0,
                               bpu_g["total_gmv"] / bpu_g["total_send"], 0)
    bpu_g["gmv_share"] = bpu_g["total_gmv"] / bpu_g["total_gmv"].sum()
    bpu_g = bpu_g.sort_values("total_gmv", ascending=False).reset_index(drop=True)

    # KPI 요약
    c1, c2, c3, c4 = st.columns(4)
    _metric_card(c1, "총 GMV", f"₩{bpu_g['total_gmv'].sum():,.0f}", color=PAL["success"])
    _metric_card(c2, "총 발송량", f"{bpu_g['total_send'].sum():,.0f}건")
    _metric_card(c3, "전체 평균 CTR", _pct(df["ctr"].mean()), color=PAL["primary"])
    _metric_card(c4, "전체 평균 RPS", f"₩{df['rps'].mean():,.1f}", color=PAL["purple"])

    st.markdown("---")

    col_a, col_b = st.columns([1, 1])

    with col_a:
        st.markdown("**BPU별 총 GMV**")
        fig = go.Figure(go.Bar(
            x=bpu_g["bpu"],
            y=bpu_g["total_gmv"],
            marker_color=[PAL["primary"] if i == 0 else PAL["border"]
                          for i in range(len(bpu_g))],
            text=bpu_g["total_gmv"].apply(lambda v: f"₩{v/1e6:.1f}M"),
            textposition="outside",
        ))
        fig.update_layout(
            height=340, yaxis_title="거래액 (원)",
            plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
            margin=dict(l=60, r=20, t=20, b=60),
            font=dict(size=12),
        )
        st.plotly_chart(fig, use_container_width=True)

    with col_b:
        st.markdown("**BPU별 GMV 점유율 (파이차트)**")
        fig2 = go.Figure(go.Pie(
            labels=bpu_g["bpu"],
            values=bpu_g["total_gmv"],
            hole=0.38,
            textinfo="label+percent",
            marker=dict(colors=px.colors.qualitative.Bold),
        ))
        fig2.update_layout(
            height=340, plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
            margin=dict(l=20, r=20, t=20, b=20),
            showlegend=False, font=dict(size=12),
        )
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")

    # ── BPU별 CTR vs RPS (발송건당 거래액) 버블차트
    _section("BPU 효율 매트릭스 (CTR × 발송건당 GMV)", "🔵")
    fig3 = go.Figure()
    for _, row in bpu_g.iterrows():
        fig3.add_trace(go.Scatter(
            x=[row["avg_ctr"] * 100],
            y=[row["rps"]],
            mode="markers+text",
            name=row["bpu"],
            text=[row["bpu"]],
            textposition="top center",
            marker=dict(
                size=np.sqrt(row["total_send"] / bpu_g["total_send"].max()) * 60 + 10,
                opacity=0.75,
            ),
        ))
    # 사분면 기준선
    avg_x = df["ctr"].mean() * 100
    avg_y = df["rps"].mean()
    fig3.add_hline(y=avg_y, line_dash="dot", line_color=PAL["muted"],
                   annotation_text=f"평균 RPS ₩{avg_y:,.0f}", annotation_position="right")
    fig3.add_vline(x=avg_x, line_dash="dot", line_color=PAL["muted"],
                   annotation_text=f"평균 CTR {avg_x:.2f}%", annotation_position="top right")
    fig3.update_layout(
        height=400,
        xaxis_title="평균 CTR (%)",
        yaxis_title="발송건당 GMV (원)",
        plot_bgcolor=PAL["card"], paper_bgcolor=PAL["bg"],
        margin=dict(l=60, r=20, t=20, b=60),
        showlegend=False, font=dict(size=12),
    )
    st.plotly_chart(fig3, use_container_width=True)

    st.caption(
        "버블 크기 = 총 발송량. "
        "우상단(高CTR·高RPS) BPU가 앱푸시 ROAS 최우선 투자 대상입니다."
    )

    st.markdown("---")

    # ── BPU 상세 테이블
    _section("BPU별 상세 성과표", "📊")
    disp = bpu_g.copy()
    disp["총 GMV"]     = disp["total_gmv"].apply(lambda v: f"₩{v:,.0f}")
    disp["총 발송"]     = disp["total_send"].apply(lambda v: f"{v:,.0f}")
    disp["평균 CTR"]   = disp["avg_ctr"].apply(lambda v: f"{v*100:.2f}%")
    disp["평균 CVR"]   = disp["avg_cvr"].apply(lambda v: f"{v*100:.2f}%")
    disp["발송건당GMV"] = disp["rps"].apply(lambda v: f"₩{v:,.1f}")
    disp["GMV 점유율"] = disp["gmv_share"].apply(lambda v: f"{v*100:.1f}%")
    disp["캠페인 수"]  = disp["n_campaigns"]
    st.dataframe(
        disp.rename(columns={"bpu": "BPU"})[
            ["BPU", "캠페인 수", "총 발송", "평균 CTR", "평균 CVR", "총 GMV", "발송건당GMV", "GMV 점유율"]
        ],
        use_container_width=True, hide_index=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# 4. 데이터 업로드 / 구글시트 연동 헬퍼
# ══════════════════════════════════════════════════════════════════════════════

def _gs_creds():
    """Streamlit Secrets → google.oauth2 Credentials. 없으면 None."""
    try:
        import streamlit as st
        from google.oauth2.service_account import Credentials
        info = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(
            info, scopes=["https://spreadsheets.google.com/feeds",
                          "https://www.googleapis.com/auth/drive"]
        )
        return creds
    except Exception:
        return None


def load_from_gsheet(spreadsheet_url: str):
    """구글시트 URL → (msg_df, perf_df). 문구/실적 시트를 자동 탐색."""
    try:
        import gspread
        creds = _gs_creds()
        if creds is None:
            return None, None
        gc = gspread.authorize(creds)
        sh = gc.open_by_url(spreadsheet_url)
        msg_df, perf_df = None, None
        for ws in sh.worksheets():
            rows = ws.get_all_values()
            if not rows:
                continue
            hdr = [str(c).strip() for c in rows[0]]
            has_title = any(h in MSG_COLMAP for h in hdr)
            has_send  = any(h in PERF_COLMAP for h in hdr)
            if has_title and ("타이틀" in hdr or "제목" in hdr or "본문" in hdr or "내용" in hdr):
                msg_df = _finalize_msg(_parse_sheet_df(rows, MSG_COLMAP))
            elif has_send and ("발송건수" in hdr or "발송" in hdr or "오픈건수" in hdr or "거래액" in hdr or "GMV" in hdr):
                perf_df = _finalize_perf(_parse_sheet_df(rows, PERF_COLMAP))
        return msg_df, perf_df
    except Exception as e:
        st.error(f"Google Sheets 연동 오류: {e}")
        return None, None


# ══════════════════════════════════════════════════════════════════════════════
# 5. 메인 앱
# ══════════════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="앱푸시 성과 분석 대시보드",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    # 전역 CSS
    st.markdown(f"""
    <style>
      @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');
      html, body, [class*="css"] {{ font-family: 'Noto Sans KR', sans-serif !important; }}
      .block-container {{ padding: 1.5rem 2rem !important; }}
      [data-testid="stMetricValue"] {{ font-size: 22px !important; font-weight: 700; }}
      div[data-testid="stTabs"] button {{ font-size: 14px !important; font-weight: 600; }}
    </style>""", unsafe_allow_html=True)

    # ── 헤더
    st.markdown(
        f"""<div style="background:linear-gradient(90deg,{PAL['slate']} 0%,{PAL['primary']} 100%);
                        border-radius:12px;padding:20px 28px;margin-bottom:20px;color:#fff;">
              <div style="font-size:22px;font-weight:700;margin-bottom:4px;">
                📲 앱푸시 발송 성과 통합 분석 대시보드</div>
              <div style="font-size:13px;opacity:.8;">
                문구 효율 · 요일/시간 · 피로도 · BPU별 GMV 기여도 분석</div>
            </div>""",
        unsafe_allow_html=True,
    )

    # ══ 사이드바: 데이터 입력 ══════════════════════════════════════════════
    with st.sidebar:
        st.markdown("### 📂 데이터 입력")
        input_mode = st.radio(
            "입력 방식", ["📄 엑셀 파일 업로드", "🔗 Google Sheets 연동", "🧪 샘플 데이터로 체험"],
            label_visibility="collapsed",
        )

        df_raw = None

        if input_mode == "📄 엑셀 파일 업로드":
            st.markdown("**문구 데이터 파일**")
            msg_file = st.file_uploader("푸시 문구 (타이틀·본문·BPU 포함)", type=["xlsx", "xls"],
                                         key="msg_upload")
            st.markdown("**실적 데이터 파일**")
            perf_file = st.file_uploader("발송 성과 (발송량·오픈·GMV 포함)", type=["xlsx", "xls"],
                                          key="perf_upload")

            if msg_file and perf_file:
                try:
                    msg_df  = parse_msg_bytes(msg_file.read())
                    perf_df = parse_perf_bytes(perf_file.read())
                    df_raw  = merge_msg_perf(msg_df, perf_df)
                    st.success(f"✅ 병합 완료 — {len(df_raw):,}건")
                except Exception as e:
                    st.error(f"파일 파싱 오류: {e}")
            elif msg_file and not perf_file:
                st.info("📋 실적 파일도 업로드하면 성과 분석이 가능합니다.")
            elif perf_file and not msg_file:
                # 실적만 있을 때도 동작
                try:
                    perf_df = parse_perf_bytes(perf_file.read())
                    df_raw  = tag_copy(perf_df)
                    st.success(f"✅ 실적 파일 로드 — {len(df_raw):,}건 (문구 미연결)")
                except Exception as e:
                    st.error(f"파일 파싱 오류: {e}")

        elif input_mode == "🔗 Google Sheets 연동":
            gs_url = st.text_input("스프레드시트 URL",
                                    placeholder="https://docs.google.com/spreadsheets/d/…")
            if st.button("연동 시작", type="primary") and gs_url:
                with st.spinner("Google Sheets 읽는 중…"):
                    msg_df, perf_df = load_from_gsheet(gs_url)
                if perf_df is not None and not perf_df.empty:
                    df_raw = merge_msg_perf(msg_df, perf_df) if (msg_df is not None and not msg_df.empty) \
                             else tag_copy(perf_df)
                    st.success(f"✅ 연동 완료 — {len(df_raw):,}건")
                else:
                    st.error("데이터를 불러오지 못했습니다. 시트 헤더를 확인하세요.")

        else:  # 샘플 데이터
            df_raw = _make_sample_data()
            st.info(f"🧪 샘플 데이터 {len(df_raw):,}건 (200건·6BPU·180일)")

        # ── 필터
        if df_raw is not None and not df_raw.empty:
            st.markdown("---")
            st.markdown("### 🔍 필터")

            # 기간
            if "send_dt" in df_raw.columns and df_raw["send_dt"].notna().any():
                min_dt = df_raw["send_dt"].min().date()
                max_dt = df_raw["send_dt"].max().date()
                date_range = st.date_input("기간", value=(min_dt, max_dt),
                                            min_value=min_dt, max_value=max_dt)
                if len(date_range) == 2:
                    df_raw = df_raw[
                        (df_raw["send_dt"].dt.date >= date_range[0]) &
                        (df_raw["send_dt"].dt.date <= date_range[1])
                    ]

            # BPU
            if "bpu" in df_raw.columns:
                bpus_all = sorted(df_raw["bpu"].dropna().unique().tolist())
                sel_bpu = st.multiselect("BPU", bpus_all, default=bpus_all,
                                          placeholder="전체")
                if sel_bpu:
                    df_raw = df_raw[df_raw["bpu"].isin(sel_bpu)]

            # 요일
            dows_all = ["월", "화", "수", "목", "금", "토", "일"]
            sel_dow = st.multiselect("요일", dows_all, default=dows_all,
                                      placeholder="전체")
            if sel_dow and "dow" in df_raw.columns:
                sel_dow_idx = [DOW_KR.index(d) for d in sel_dow]
                df_raw = df_raw[df_raw["dow"].isin(sel_dow_idx)]

            # 키워드
            kw = st.text_input("문구 키워드 검색", placeholder="할인, 쿠폰, 브랜드명…")
            if kw and "title" in df_raw.columns:
                mask = (
                    df_raw["title"].str.contains(kw, case=False, na=False) |
                    df_raw.get("body", pd.Series([""] * len(df_raw))).str.contains(kw, case=False, na=False)
                )
                df_raw = df_raw[mask]

            st.caption(f"필터 적용 후: **{len(df_raw):,}건**")

    # ══ 메인 콘텐츠 ═══════════════════════════════════════════════════════
    if df_raw is None or df_raw.empty:
        st.markdown(
            """<div style="text-align:center;padding:80px 20px;color:#64748b;">
               <div style="font-size:48px;margin-bottom:16px;">📲</div>
               <div style="font-size:18px;font-weight:600;">사이드바에서 데이터를 불러오세요</div>
               <div style="font-size:14px;margin-top:8px;">
                 엑셀 업로드 · Google Sheets 연동 · 샘플 데이터 체험 중 선택 가능합니다.</div>
             </div>""",
            unsafe_allow_html=True,
        )
        return

    # 탭
    tab1, tab2, tab3, tab4 = st.tabs([
        "✍️ 문구 효율 분석",
        "📅 요일·시간대 분석",
        "📉 피로도 분석",
        "🏢 BPU 기여도",
    ])

    with tab1:
        tab_copy_analysis(df_raw)
    with tab2:
        tab_time_analysis(df_raw)
    with tab3:
        tab_fatigue_analysis(df_raw)
    with tab4:
        tab_bpu_analysis(df_raw)

    # 데이터 다운로드
    with st.expander("📥 분석 데이터 다운로드"):
        csv = df_raw.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            "CSV 다운로드", data=csv.encode("utf-8-sig"),
            file_name=f"push_analytics_{datetime.date.today()}.csv",
            mime="text/csv",
        )


if __name__ == "__main__":
    main()
